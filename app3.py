
# ==============================================================================
# ENTERPRISE LOGISTICS, DISPATCH ENGINE & SALES AUTOMATION SUITE
# ENHANCED DISPATCH PLANNER WITH DYNAMIC IN-APP DATABASE BUILDER & CRUD OPERATIONS
# ==============================================================================

import datetime
import io
import json
import os
import re
import smtplib
import sqlite3
import urllib.parse
import zipfile
from email.message import EmailMessage

from fpdf import FPDF
import openpyxl
import pandas as pd
import pytz
import streamlit as st
import streamlit.components.v1 as components

# ==============================================================================
# SECTION 1: GLOBAL CONFIGURATION, TIMEZONE (IST) & METADATA
# ==============================================================================

st.set_page_config(
    page_title="Enterprise Logistics & Sales Automation Hub",
    page_icon="🚚",
    layout="wide",
    initial_sidebar_state="expanded"
)

IST = pytz.timezone("Asia/Kolkata")

def get_ist_now():
    return datetime.datetime.now(IST)

def get_ist_date_str():
    return get_ist_now().strftime("%Y-%m-%d")

def get_ist_time_str():
    return get_ist_now().strftime("%H:%M:%S")

def get_ist_timestamp_full():
    return get_ist_now().strftime("%Y-%m-%d %H:%M:%S")

def get_ist_file_suffix():
    return get_ist_now().strftime("%H%M%S")

DB_NAME = "enterprise_logistics_sales_hub.db"

# ==============================================================================
# SECTION 2: 8 COMPLETE ENTERPRISE COLOR PALETTES
# ==============================================================================

THEMES = {
    "💼 Classic Enterprise Navy": {
        "icon": "💼", "bg": "#f4f6f9", "text": "#1f2937", "card_bg": "#ffffff", "border": "#cbd5e1",
        "btn_bg": "#1e3a8a", "btn_hover": "#1d4ed8", "primary": "#2563eb", "input_bg": "#ffffff", "input_text": "#1f2937"
    },
    "⚙️ Classic Default Theme": {
        "icon": "⚙️", "bg": "#f4f6f9", "text": "#1f2937", "card_bg": "#ffffff", "border": "#cbd5e1",
        "btn_bg": "#1e3a8a", "btn_hover": "#1d4ed8", "primary": "#2563eb", "input_bg": "#ffffff", "input_text": "#1f2937"
    },
    "🌙 Modern Dark ERP": {
        "icon": "🌙", "bg": "#0b0f19", "text": "#f3f4f6", "card_bg": "#1f2937", "border": "#374151",
        "btn_bg": "#374151", "btn_hover": "#4b5563", "primary": "#3b82f6", "input_bg": "#111827", "input_text": "#f3f4f6"
    },
    "📊 Corporate Slate": {
        "icon": "📊", "bg": "#eef2f5", "text": "#0f172a", "card_bg": "#ffffff", "border": "#94a3b8",
        "btn_bg": "#475569", "btn_hover": "#334155", "primary": "#0284c7", "input_bg": "#ffffff", "input_text": "#0f172a"
    },
    "☀️ Clean Light Minimal": {
        "icon": "☀️", "bg": "#ffffff", "text": "#111827", "card_bg": "#f9fafb", "border": "#d1d5db",
        "btn_bg": "#0f172a", "btn_hover": "#1e293b", "primary": "#10b981", "input_bg": "#ffffff", "input_text": "#111827"
    },
    "⚡ Cyber Blue": {
        "icon": "⚡", "bg": "#f0fdfa", "text": "#042f2e", "card_bg": "#ccfbf1", "border": "#5eead4",
        "btn_bg": "#0d9488", "btn_hover": "#0f766e", "primary": "#14b8a6", "input_bg": "#ffffff", "input_text": "#042f2e"
    },
    "🌲 Emerald Corporate": {
        "icon": "🌲", "bg": "#f0fdf4", "text": "#14532d", "card_bg": "#dcfce7", "border": "#86efac",
        "btn_bg": "#16a34a", "btn_hover": "#15803d", "primary": "#22c55e", "input_bg": "#ffffff", "input_text": "#14532d"
    },
    "🍇 Executive Burgundy": {
        "icon": "🍇", "bg": "#fdf2f8", "text": "#500724", "card_bg": "#fce7f3", "border": "#f472b6",
        "btn_bg": "#db2777", "btn_hover": "#be185d", "primary": "#ec4899", "input_bg": "#ffffff", "input_text": "#500724"
    },
    "🪙 Titanium Charcoal": {
        "icon": "🪙", "bg": "#18181b", "text": "#fafafa", "card_bg": "#27272a", "border": "#52525b",
        "btn_bg": "#52525b", "btn_hover": "#71717a", "primary": "#e4e4e7", "input_bg": "#09090b", "input_text": "#fafafa"
    }
}

# ==============================================================================
# SECTION 3: SESSION STATE INITIALIZATION
# ==============================================================================

GLOBAL_DEFAULTS = {
    "selected_theme": "🌙 Modern Dark ERP",
    "fg_code": "FG500014",
    "col_map": "36:FG500014AJ\n37:FG500014AK",
    "agency_override": "101:36:FG500014N01\n101:37:FG500014N02",
    "route": "22",
    "email_user": st.secrets.get("email", {}).get("sender_email", "") if hasattr(st, "secrets") else "",
    "email_pass": st.secrets.get("email", {}).get("app_password", "") if hasattr(st, "secrets") else "",
    "recipient": st.secrets.get("email", {}).get("recipient_email", "") if hasattr(st, "secrets") else "",
    "whatsapp_num": "919876543210",
    "processed_files": [],
    "comparison_summary": [],
    "skipped_rows_log": [],
    "anomaly_logs": [],
    "unmapped_current_batch": [],
    "kpi_data": {
        "input_qty": 0.0,
        "gen_qty": 0.0,
        "valid_count": 0,
        "missing_count": 0,
        "skipped_count": 0
    }
}

for d_key, d_val in GLOBAL_DEFAULTS.items():
    if d_key not in st.session_state:
        st.session_state[d_key] = d_val

current_theme = THEMES.get(st.session_state.selected_theme, THEMES["💼 Classic Enterprise Navy"])

st.markdown(
    f"""
    <style>
        .stApp {{
            background-color: {current_theme['bg']};
            color: {current_theme['text']};
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        }}
        h1, h2, h3, h4, h5, h6, p, span, label, .stMarkdown {{
            color: {current_theme['text']} !important;
        }}
        input, textarea, select {{
            background-color: {current_theme['input_bg']} !important;
            color: {current_theme['input_text']} !important;
            border: 1px solid {current_theme['border']} !important;
            border-radius: 4px !important;
        }}
        .stButton>button {{
            width: 100%;
            height: 38px;
            background-color: {current_theme['btn_bg']} !important;
            color: #ffffff !important;
            font-size: 13px !important;
            font-weight: 600 !important;
            border-radius: 4px;
            border: 1px solid {current_theme['border']};
            transition: all 0.2s ease-in-out;
        }}
        .stButton>button:hover {{
            background-color: {current_theme['btn_hover']} !important;
            color: #ffffff !important;
        }}
        button[kind="primary"] {{
            background-color: {current_theme['primary']} !important;
            color: #ffffff !important;
        }}
        div[data-testid="stExpander"] {{
            background-color: {current_theme['card_bg']};
            border: 1px solid {current_theme['border']};
            border-radius: 6px;
            margin-bottom: 12px;
        }}
        div[data-testid="stDataFrame"] {{
            border: 1px solid {current_theme['border']};
            border-radius: 6px;
            background-color: {current_theme['card_bg']};
        }}
    </style>
    """,
    unsafe_allow_html=True
)

# ==============================================================================
# SECTION 4: UNIFIED DATABASE ARCHITECTURE (ALL CORE TABLES)
# ==============================================================================

def get_db_connection():
    conn = sqlite3.connect(DB_NAME)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_all_enterprise_databases():
    conn = get_db_connection()
    cur = conn.cursor()

    # 1. Route-Agency-DR Master
    cur.execute("""
        CREATE TABLE IF NOT EXISTS unique_routes_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name TEXT,
            route_no TEXT,
            agency_no TEXT,
            dr_code TEXT,
            created_at TEXT,
            UNIQUE(route_no, agency_no, dr_code)
        )
    """)

    # 2. Uploaded Input File Archive
    cur.execute("""
        CREATE TABLE IF NOT EXISTS uploaded_files_archive (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name TEXT UNIQUE,
            upload_timestamp TEXT,
            total_records INTEGER,
            file_size_kb REAL,
            batch_status TEXT
        )
    """)

    # 3. Pending Orders Database
    cur.execute("""
        CREATE TABLE IF NOT EXISTS pending_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT,
            order_no TEXT,
            route_no TEXT,
            agency_no TEXT,
            dr_code TEXT,
            fg_code TEXT,
            bags_qty REAL,
            weight_mt REAL,
            order_ref TEXT,
            status TEXT DEFAULT 'Pending',
            uploaded_at TEXT
        )
    """)

    # 4. Transporter Fleet Master
    cur.execute("""
        CREATE TABLE IF NOT EXISTS fleet_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vehicle_no TEXT UNIQUE,
            vehicle_type TEXT,
            capacity_bags INTEGER,
            capacity_mt REAL,
            transporter_name TEXT,
            driver_name TEXT,
            driver_phone TEXT,
            status TEXT DEFAULT 'Available'
        )
    """)

    # 5. Plant Loading Bays Master
    cur.execute("""
        CREATE TABLE IF NOT EXISTS loading_bays (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bay_no TEXT UNIQUE,
            bay_name TEXT,
            status TEXT DEFAULT 'Open'
        )
    """)

    # 6. Trip Loading Slips Master
    cur.execute("""
        CREATE TABLE IF NOT EXISTS trip_loading_slips (
            trip_id TEXT PRIMARY KEY,
            trip_date TEXT,
            route_no TEXT,
            vehicle_no TEXT,
            transporter_name TEXT,
            driver_name TEXT,
            driver_phone TEXT,
            loading_bay TEXT,
            total_bags REAL,
            total_weight_mt REAL,
            capacity_utilization_pct REAL,
            status TEXT,
            created_at TEXT
        )
    """)

    # 7. Trip Order Items Manifest Sequence
    cur.execute("""
        CREATE TABLE IF NOT EXISTS trip_order_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            trip_id TEXT,
            order_no TEXT,
            agency_no TEXT,
            route_no TEXT,
            dr_code TEXT,
            fg_code TEXT,
            allocated_bags REAL,
            allocated_weight_mt REAL,
            delivery_seq INTEGER,
            status TEXT DEFAULT 'Assigned',
            FOREIGN KEY (trip_id) REFERENCES trip_loading_slips(trip_id) ON DELETE CASCADE
        )
    """)

    # 8. Daily Dispatch Sale Register
    cur.execute("""
        CREATE TABLE IF NOT EXISTS daily_dispatch_register (
            register_id INTEGER PRIMARY KEY AUTOINCREMENT,
            dispatch_date TEXT,
            trip_id TEXT,
            vehicle_no TEXT,
            transporter_name TEXT,
            route_no TEXT,
            agency_no TEXT,
            order_no TEXT,
            dr_code TEXT,
            fg_code TEXT,
            dispatched_bags REAL,
            dispatched_weight_mt REAL,
            bay_no TEXT,
            dispatched_at TEXT
        )
    """)

    # 9. Partial / Remaining Pending Dispatch Database
    cur.execute("""
        CREATE TABLE IF NOT EXISTS partial_dispatch_ledger (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            trip_id TEXT,
            source_file TEXT,
            order_no TEXT,
            route_no TEXT,
            agency_no TEXT,
            dr_code TEXT,
            fg_code TEXT,
            original_bags REAL,
            dispatched_bags REAL,
            remaining_bags REAL,
            status TEXT DEFAULT 'Partial Pending',
            created_at TEXT
        )
    """)

    # 10. Unmapped Missing DR Fallback Ledger
    cur.execute("""
        CREATE TABLE IF NOT EXISTS unmapped_missing_dr_ledger (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name TEXT,
            route_no TEXT,
            agency_no TEXT,
            dr_code TEXT,
            created_at TEXT,
            UNIQUE(route_no, agency_no)
        )
    """)

    # 11. Generated Output Files Storage Ledger (BLOB Storage)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS output_files_ledger (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name TEXT UNIQUE,
            file_type TEXT,
            file_data BLOB,
            created_at TEXT
        )
    """)

    # 12. Traceability Ledger
    cur.execute("""
        CREATE TABLE IF NOT EXISTS input_output_traceability (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_timestamp TEXT,
            input_file_name TEXT,
            input_file_blob BLOB,
            total_input_qty REAL,
            generated_output_file TEXT,
            output_type TEXT,
            version_no INTEGER,
            created_at TEXT
        )
    """)

    # 13. Discrepancy Audit Ledger
    cur.execute("""
        CREATE TABLE IF NOT EXISTS discrepancy_audit_ledger (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_timestamp TEXT,
            file_name TEXT,
            agency_no TEXT,
            dr_code TEXT,
            fg_code TEXT,
            input_qty REAL,
            generated_qty REAL,
            difference REAL,
            logged_at TEXT
        )
    """)

    # 14. System Audit History Logs
    cur.execute("""
        CREATE TABLE IF NOT EXISTS history_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            files_count INTEGER,
            total_qty REAL,
            status TEXT,
            remarks TEXT
        )
    """)

    # Default Fleet Seed
    cur.execute("SELECT COUNT(*) FROM fleet_master")
    if cur.fetchone()[0] == 0:
        default_fleet = [
            ('PB-10-AZ-1122', '10 Wheeler Truck', 400, 20.0, 'National Logistics', 'Gurpreet Singh', '9876543210', 'Available'),
            ('PB-08-BX-4455', '12 Wheeler Multi-Axle', 600, 30.0, 'Speedway Cargo', 'Baljit Sharma', '9812345678', 'Available'),
            ('PB-29-CD-9900', 'Canter / Eicher', 200, 10.0, 'Punjab Roadlines', 'Ramesh Kumar', '9823456789', 'Available'),
            ('PB-11-GH-3321', '14 Wheeler Heavy', 800, 40.0, 'Apex Transporters', 'Jarnail Singh', '9834567890', 'Available')
        ]
        cur.executemany("""
            INSERT INTO fleet_master (vehicle_no, vehicle_type, capacity_bags, capacity_mt, transporter_name, driver_name, driver_phone, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, default_fleet)

    # Default Bays Seed
    cur.execute("SELECT COUNT(*) FROM loading_bays")
    if cur.fetchone()[0] == 0:
        default_bays = [
            ('BAY-01', 'North Plant Main Gate', 'Open'),
            ('BAY-02', 'Storage Silo Bay 2', 'Open'),
            ('BAY-03', 'Express Bulk Bay 3', 'Open')
        ]
        cur.executemany("INSERT INTO loading_bays (bay_no, bay_name, status) VALUES (?, ?, ?)", default_bays)

    conn.commit()
    conn.close()

init_all_enterprise_databases()

# ==============================================================================
# SECTION 5: UTILITY & EXPORT GENERATOR ENGINES
# ==============================================================================

def to_excel_download_bytes(df: pd.DataFrame, sheet_name="DataSheet") -> bytes:
    output_stream = io.BytesIO()
    with pd.ExcelWriter(output_stream, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return output_stream.getvalue()

def build_pdf_loading_slip(trip_data: dict, items_df: pd.DataFrame) -> bytes:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    pdf.set_font("Arial", "B", 16)
    pdf.cell(190, 8, "ENTERPRISE DISPATCH & LOADING SLIP", ln=True, align="C")
    pdf.set_font("Arial", "", 9)
    pdf.cell(190, 5, f"Official Gate Pass & Manifest | Generated (IST): {get_ist_timestamp_full()}", ln=True, align="C")
    pdf.ln(4)

    pdf.set_font("Arial", "B", 9)
    pdf.cell(95, 6, f"Trip ID: {trip_data.get('trip_id', '')}", border=1)
    pdf.cell(95, 6, f"Trip Date: {trip_data.get('trip_date', '')}", border=1, ln=True)
    pdf.cell(95, 6, f"Vehicle No: {trip_data.get('vehicle_no', '')}", border=1)
    pdf.cell(95, 6, f"Route No: {trip_data.get('route_no', '')}", border=1, ln=True)
    pdf.cell(95, 6, f"Transporter: {trip_data.get('transporter_name', '')}", border=1)
    pdf.cell(95, 6, f"Loading Bay: {trip_data.get('loading_bay', '')}", border=1, ln=True)
    pdf.cell(95, 6, f"Driver Name: {trip_data.get('driver_name', '')}", border=1)
    pdf.cell(95, 6, f"Driver Phone: {trip_data.get('driver_phone', '')}", border=1, ln=True)
    pdf.ln(4)

    pdf.set_font("Arial", "B", 9)
    pdf.cell(12, 6, "Seq", border=1, align="C")
    pdf.cell(32, 6, "Agency No", border=1, align="C")
    pdf.cell(40, 6, "DR Code", border=1, align="C")
    pdf.cell(46, 6, "FG Code", border=1, align="C")
    pdf.cell(30, 6, "Bags Qty", border=1, align="C")
    pdf.cell(30, 6, "Weight (MT)", border=1, ln=True, align="C")

    pdf.set_font("Arial", "", 8)
    for _, it in items_df.iterrows():
        pdf.cell(12, 5, str(it.get("delivery_seq", "")), border=1, align="C")
        pdf.cell(32, 5, str(it.get("agency_no", "")), border=1, align="C")
        pdf.cell(40, 5, str(it.get("dr_code", "")), border=1, align="C")
        pdf.cell(46, 5, str(it.get("fg_code", "")), border=1, align="C")
        pdf.cell(30, 5, f"{float(it.get('allocated_bags', 0)):,.0f}", border=1, align="R")
        pdf.cell(30, 5, f"{float(it.get('allocated_weight_mt', 0)):,.2f}", border=1, ln=True, align="R")

    pdf.set_font("Arial", "B", 9)
    pdf.cell(130, 6, "TOTAL LOAD MANIFEST", border=1, align="R")
    pdf.cell(30, 6, f"{float(trip_data.get('total_bags', 0)):,.0f}", border=1, align="R")
    pdf.cell(30, 6, f"{float(trip_data.get('total_weight_mt', 0)):,.2f}", border=1, ln=True, align="R")

    pdf.ln(12)
    pdf.set_font("Arial", "", 9)
    pdf.cell(63, 6, "Driver Sign: ________________", ln=False)
    pdf.cell(63, 6, "Security Gate In: ________________", ln=False)
    pdf.cell(63, 6, "Supervisor Sign: ________________", ln=True)

    return bytes(pdf.output())

# ==============================================================================
# SECTION 6: SIDEBAR NAVIGATION ENGINE
# ==============================================================================

with st.sidebar:
    st.image("https://img.icons8.com/color/96/delivery-truck.png", width=55)
    st.title("Logistics Master Suite")

    main_menu = st.radio(
        "Navigation",
        [
            "⚡ Inbound Demand & Sales Order Engine",
            "🚚 Route Dispatch Trip Planner",
            "📦 Live Inventory Stock & ERP Demand Matcher",
            "📋 Loading Slips & Active Trips",
            "📖 Daily Dispatch Sale Register",
            "🧩 Partial / Split Dispatch Database",
            "⏳ Pending Orders Ledger",
            "🗄️ File Upload Archive",
            "📋 Master DB & Unmapped Ledger",
            "🚛 Fleet & Loading Bay Master",
            "🔍 Traceability & Audit Ledgers",
            "📊 Executive KPI & Visual Analytics",
            "🎯 Universal Date & Multi-Field Filter Center",
            "⚡ Smart Multi-Truck Load Optimizer Pro",     
            "🗄️ In-App Database Builder & Dynamic Linker" # NEW MODULE ADDED
        ]
    )
    st.markdown("---")
    st.subheader("🎨 Interface Theme Engine")
    theme_choice = st.selectbox(
        "Select Theme",
        list(THEMES.keys()),
        index=list(THEMES.keys()).index(st.session_state.selected_theme)
    )
    if theme_choice != st.session_state.selected_theme:
        st.session_state.selected_theme = theme_choice
        st.rerun()

    st.markdown("---")
    st.subheader("📱 Notification Config")
    st.session_state.whatsapp_num = st.text_input("WhatsApp Alert Mobile No", value=st.session_state.whatsapp_num)

# ==============================================================================
# ENTERPRISE LOGISTICS, DISPATCH ENGINE & SALES AUTOMATION SUITE
# PART 2: INBOUND DEMAND ENGINE, TRIP PLANNER & LOADING SLIPS
# ==============================================================================

# ==============================================================================
# MODULE 1: INBOUND DEMAND & SALES ORDER AUTOMATION ENGINE
# ==============================================================================

if main_menu == "⚡ Inbound Demand & Sales Order Engine":
    st.title("⚡ Enterprise Inbound Demand & Sales Order Processing Engine")
    st.markdown("Upload multiple **Demand Workbooks** to execute DR auto-lookup, eliminate duplicate orders, generate structured `Output.xlsx` files, and sync pending demand.")

    with st.expander("⚙️ SKU, Route & Multi-Channel Dispatch Settings", expanded=False):
        c1, c2, c3 = st.columns(3)
        with c1:
            st.session_state.fg_code = st.text_input("Default Fallback FG Code", value=st.session_state.fg_code)
            st.session_state.route = st.text_input("Default Route Fallback", value=st.session_state.route)
        with c2:
            st.session_state.col_map = st.text_area("Column Index Mapping (Col:FG)", value=st.session_state.col_map, height=80)
        with c3:
            st.session_state.agency_override = st.text_area("Agency SKU Overrides (Agency:Col:FG)", value=st.session_state.agency_override, height=80)

        st.markdown("---")
        c4, c5 = st.columns(2)
        with c4:
            st.session_state.email_user = st.text_input("Sender Gmail ID", value=st.session_state.email_user)
            st.session_state.email_pass = st.text_input("Gmail App Password", type="password", value=st.session_state.email_pass)
        with c5:
            st.session_state.recipient = st.text_input("Recipient Email", value=st.session_state.recipient)

    # Parse Mappings
    col_map_dict = {}
    for line in st.session_state.col_map.split("\n"):
        if ":" in line:
            parts = line.split(":")
            if parts[0].strip().isdigit():
                col_map_dict[int(parts[0].strip())] = parts[1].strip()

    agency_override_dict = {}
    for line in st.session_state.agency_override.split("\n"):
        parts = line.split(":")
        if len(parts) == 3 and parts[0].strip().isdigit() and parts[1].strip().isdigit():
            agency_override_dict[(int(parts[0].strip()), int(parts[1].strip()))] = parts[2].strip()

    uploaded_files = st.file_uploader(
        "Upload Inbound Demand Excel Workbooks",
        type=["xlsx", "xls"],
        accept_multiple_files=True
    )

    if uploaded_files and st.button("🚀 Process Batch Orders & Ingest to Pending Database", type="primary"):
        st.session_state.processed_files = []
        st.session_state.comparison_summary = []
        st.session_state.skipped_rows_log = []
        st.session_state.anomaly_logs = []
        st.session_state.unmapped_current_batch = []

        total_in_qty = 0.0
        total_gen_qty = 0.0
        total_valid = 0
        total_missing = 0
        total_skipped = 0

        conn = get_db_connection()
        cur = conn.cursor()

        try:
            with open("Output.xlsx", "rb") as f:
                template_bytes = f.read()
        except FileNotFoundError:
            template_bytes = None

        pending_records_to_insert = []
        master_routes_to_insert = []
        unmapped_records_to_insert = []
        traceability_records = []

        batch_ts = get_ist_timestamp_full()
        today_date = get_ist_date_str()
        time_suffix = get_ist_file_suffix()

        for up_file in uploaded_files:
            short_fname = up_file.name
            if short_fname.lower() == "output.xlsx":
                continue

            # Anti-Duplicate Guard: Skip if file was already uploaded and archived
            cur.execute("SELECT id FROM uploaded_files_archive WHERE file_name=?", (short_fname,))
            if cur.fetchone():
                st.warning(f"⚠️ '{short_fname}' pehle se process ho chuki hai. Duplicate upload skip kiya gaya.")
                continue

            f_bytes = up_file.getvalue()

            try:
                wb_raw = openpyxl.load_workbook(io.BytesIO(f_bytes), data_only=True)
                ws_raw = wb_raw.active
                data_matrix = []
                for row_cells in ws_raw.iter_rows(values_only=True):
                    data_matrix.append(list(row_cells))
                df_input = pd.DataFrame(data_matrix)
            except Exception as e:
                try:
                    df_input = pd.read_excel(io.BytesIO(f_bytes), header=None)
                except Exception as ex:
                    st.error(f"Error reading '{short_fname}': {str(ex)}")
                    continue

            # 1. Detect FG Header Row & Column
            fg_row, fg_col = -1, -1
            for r in range(min(df_input.shape[0], 30)):
                for c in range(df_input.shape[1]):
                    cell_val = str(df_input.iloc[r, c]).strip().upper()
                    if "FG" in cell_val or "PRODUCT" in cell_val or "SKU" in cell_val:
                        fg_row, fg_col = r, c
                        break
                if fg_row != -1:
                    break

            if fg_row == -1:
                st.error(f"❌ '{short_fname}' me FG / SKU header row detect nahi hui.")
                continue

            # 2. Strict Total Column Cutoff (Formula & Summary Header Guard)
            total_col = df_input.shape[1]
            for c_s in range(fg_col, df_input.shape[1]):
                val_header = str(df_input.iloc[fg_row, c_s]).strip().upper()
                if any(kw in val_header for kw in ["TOTAL", "SUM", "NET", "TTL", "GR. TOTAL", "GRAND"]):
                    total_col = c_s
                    break
                is_sum_col = False
                for r_scan in range(max(0, fg_row - 5), min(fg_row + 5, df_input.shape[0])):
                    cell_txt = str(df_input.iloc[r_scan, c_s]).strip().upper()
                    if any(kw in cell_txt for kw in ["TOTAL", "SUM", "GRAND TOTAL"]):
                        is_sum_col = True
                        break
                if is_sum_col:
                    total_col = c_s
                    break

            # 3. Route Number Detection
            route_num = st.session_state.route if st.session_state.route != "" else "22"
            for r in range(fg_row):
                for c in range(min(total_col, 20)):
                    val = str(df_input.iloc[r, c]).replace('.0', '').strip()
                    if val != "" and 1 <= len(val) <= 4 and any(ch.isdigit() for ch in val):
                        if not any(prefix in val.upper() for prefix in ["DR", "FG", "OR", "SO", "RT-", "TOTAL"]):
                            route_num = val
                            break

            resolved_route = "".join(ch for ch in str(route_num) if ch.isalnum() or ch in ('-', '_'))

            # 4. Agency & DR Columns
            agency_col = -1
            for c_s in range(fg_col - 1, -1, -1):
                col_samples = df_input.iloc[fg_row + 1: fg_row + 15, c_s].dropna().astype(str).str.replace(r'\.0$', '', regex=True)
                if col_samples.str.isdigit().sum() >= 2:
                    agency_col = c_s
                    break
            if agency_col == -1:
                agency_col = fg_col - 1 if fg_col > 0 else 0

            dr_code_col = -1
            for c_s in range(fg_col - 1, -1, -1):
                col_samples = df_input.iloc[fg_row + 1: fg_row + 15, c_s].dropna().astype(str).str.upper()
                if col_samples.str.startswith("DR").sum() >= 1:
                    dr_code_col = c_s
                    break

            valid_cols = [(c, str(df_input.iloc[fg_row, c]).strip()) for c in range(fg_col, total_col)]

            # Output Workbooks
            wb_valid = openpyxl.load_workbook(io.BytesIO(template_bytes)) if template_bytes else openpyxl.Workbook()
            ws_valid = wb_valid["Order Data"] if "Order Data" in wb_valid.sheetnames else wb_valid.active
            wb_missing = openpyxl.load_workbook(io.BytesIO(template_bytes)) if template_bytes else openpyxl.Workbook()
            ws_missing = wb_missing["Order Data"] if "Order Data" in wb_missing.sheetnames else wb_missing.active

            valid_r_idx, missing_r_idx = 6, 6
            valid_order_no, missing_order_no = 1, 1
            valid_items_cnt, missing_items_cnt = 0, 0
            agency_counts_valid, agency_counts_missing = {}, {}
            file_input_qty = 0.0

            for r in range(fg_row + 1, df_input.shape[0]):
                # SUMMARY ROW EXCLUSION
                row_raw_values = [str(val).strip().upper() for val in df_input.iloc[r, :min(total_col, 15)] if pd.notna(val)]
                if any(any(kw in cell_str for kw in ["TOTAL", "SUM", "GRAND TOTAL", "GR. TOTAL", "NET TOTAL", "TOTAL QTY"]) for cell_str in row_raw_values):
                    continue

                agency_raw = df_input.iloc[r, agency_col]
                if pd.isna(agency_raw) or str(agency_raw).strip() in ["", "nan", "None", "0"]:
                    continue

                agency_str = str(agency_raw).replace(".0", "").strip()
                if not agency_str.isdigit() or not (1 <= len(agency_str) <= 6):
                    total_skipped += 1
                    st.session_state.skipped_rows_log.append({
                        "File Name": short_fname, "Row": r + 1, "Agency": str(agency_raw), "Reason": "Non-numeric agency"
                    })
                    continue

                agency_val = int(agency_str)

                clean_dr = ""
                if dr_code_col >= 0:
                    raw_dr = df_input.iloc[r, dr_code_col]
                    if pd.notna(raw_dr) and "DR" in str(raw_dr).upper():
                        clean_dr = str(raw_dr).strip()

                if not clean_dr:
                    cur.execute(
                        "SELECT dr_code FROM unique_routes_master WHERE route_no=? AND agency_no=?",
                        (resolved_route, str(agency_val))
                    )
                    match = cur.fetchone()
                    if match:
                        clean_dr = match[0]
                    else:
                        clean_dr = f"NEW_CUST_{agency_val}"
                        unmapped_records_to_insert.append((short_fname, resolved_route, str(agency_val), clean_dr, batch_ts))
                        st.session_state.unmapped_current_batch.append({
                            "File Name": short_fname, "Route": resolved_route, "Agency": agency_val, "Fallback DR": clean_dr
                        })

                is_valid_dr = clean_dr.upper().startswith("DR")
                if is_valid_dr:
                    master_routes_to_insert.append((short_fname, resolved_route, str(agency_val), clean_dr, batch_ts))
                    agency_counts_valid[agency_val] = agency_counts_valid.get(agency_val, 0) + 1
                    seq_num = agency_counts_valid[agency_val]
                    ref_code = f"RT-{resolved_route}-{agency_val}-{today_date}" if seq_num == 1 else f"RT-{resolved_route}-{agency_val}-{today_date}-{seq_num}"
                    target_ws, curr_r, order_id_to_write = ws_valid, valid_r_idx, valid_order_no
                else:
                    agency_counts_missing[agency_val] = agency_counts_missing.get(agency_val, 0) + 1
                    seq_num = agency_counts_missing[agency_val]
                    ref_code = f"RT-{resolved_route}-{agency_val}-{today_date}-NEW" if seq_num == 1 else f"RT-{resolved_route}-{agency_val}-{today_date}-NEW-{seq_num}"
                    target_ws, curr_r, order_id_to_write = ws_missing, missing_r_idx, missing_order_no

                item_seq_id = 10
                row_items_added = 0
                for c_idx, fg_val in valid_cols:
                    q_val = df_input.iloc[r, c_idx]
                    if pd.notna(q_val) and str(q_val).strip() != "":
                        q_str = str(q_val).strip()
                        if q_str.startswith("=") or "SUM(" in q_str.upper():
                            continue
                        try:
                            f_qty = float(q_str)
                            if f_qty > 0:
                                current_fg = fg_val if fg_val.startswith("FG") else col_map_dict.get(c_idx, st.session_state.fg_code)
                                if (agency_val, c_idx) in agency_override_dict:
                                    current_fg = agency_override_dict[(agency_val, c_idx)]

                                # Order uniqueness check
                                cur.execute("""
                                    SELECT id FROM pending_orders 
                                    WHERE route_no=? AND agency_no=? AND fg_code=? AND status='Pending' AND bags_qty=?
                                """, (resolved_route, str(agency_val), current_fg, f_qty))
                                if not cur.fetchone():
                                    pending_records_to_insert.append((
                                        short_fname, f"ORD-{agency_val}-{r}", resolved_route, str(agency_val),
                                        clean_dr, current_fg, f_qty, round(f_qty * 0.05, 2), ref_code, "Pending", batch_ts
                                    ))

                                total_in_qty += f_qty
                                total_gen_qty += f_qty
                                file_input_qty += f_qty
                                row_items_added += 1

                                target_ws.cell(row=curr_r, column=2, value=order_id_to_write)
                                target_ws.cell(row=curr_r, column=3, value="OR")
                                target_ws.cell(row=curr_r, column=4, value="SO20")
                                target_ws.cell(row=curr_r, column=5, value=10)
                                target_ws.cell(row=curr_r, column=6, value=20)
                                target_ws.cell(row=curr_r, column=7, value=clean_dr)
                                target_ws.cell(row=curr_r, column=8, value=clean_dr)
                                target_ws.cell(row=curr_r, column=9, value=ref_code)
                                target_ws.cell(row=curr_r, column=10, value=today_date)
                                target_ws.cell(row=curr_r, column=11, value=today_date)
                                target_ws.cell(row=curr_r, column=15, value=item_seq_id)
                                target_ws.cell(row=curr_r, column=16, value=current_fg)
                                target_ws.cell(row=curr_r, column=19, value=f_qty)
                                target_ws.cell(row=curr_r, column=20, value="Bag")
                                target_ws.cell(row=curr_r, column=22, value=2100)
                                target_ws.cell(row=curr_r, column=26, value=resolved_route)
                                target_ws.cell(row=curr_r, column=27, value=agency_val)

                                item_seq_id += 10
                                curr_r += 1
                        except ValueError:
                            pass

                if row_items_added > 0:
                    if is_valid_dr:
                        valid_r_idx = curr_r
                        valid_order_no += 1
                        valid_items_cnt += row_items_added
                        total_valid += 1
                    else:
                        missing_r_idx = curr_r
                        missing_order_no += 1
                        missing_items_cnt += row_items_added
                        total_missing += 1

            if valid_items_cnt > 0:
                buf_v = io.BytesIO()
                wb_valid.save(buf_v)
                out_name_v = f"{resolved_route}_{today_date}_{time_suffix}_Valid.xlsx"
                st.session_state.processed_files.append({
                    "name": f"{short_fname} (Valid DR)", "data": buf_v.getvalue(), "filename": out_name_v, "orders": valid_items_cnt
                })
                cur.execute(
                    "INSERT OR REPLACE INTO output_files_ledger (file_name, file_type, file_data, created_at) VALUES (?, 'Valid DR', ?, ?)",
                    (out_name_v, buf_v.getvalue(), batch_ts)
                )
                traceability_records.append((batch_ts, short_fname, f_bytes, file_input_qty, out_name_v, "Valid DR", 1, batch_ts))

            if missing_items_cnt > 0:
                buf_m = io.BytesIO()
                wb_missing.save(buf_m)
                out_name_m = f"{resolved_route}_{today_date}_{time_suffix}_Missing_DR.xlsx"
                st.session_state.processed_files.append({
                    "name": f"{short_fname} (Missing DR)", "data": buf_m.getvalue(), "filename": out_name_m, "orders": missing_items_cnt
                })
                cur.execute(
                    "INSERT OR REPLACE INTO output_files_ledger (file_name, file_type, file_data, created_at) VALUES (?, 'Missing DR', ?, ?)",
                    (out_name_m, buf_m.getvalue(), batch_ts)
                )
                traceability_records.append((batch_ts, short_fname, f_bytes, file_input_qty, out_name_m, "Missing DR", 1, batch_ts))

            cur.execute(
                """
                INSERT OR REPLACE INTO uploaded_files_archive (file_name, upload_timestamp, total_records, file_size_kb, batch_status)
                VALUES (?, ?, ?, ?, 'Processed')
            """,
                (short_fname, batch_ts, valid_items_cnt + missing_items_cnt, round(len(f_bytes) / 1024, 2))
            )

        if pending_records_to_insert:
            cur.executemany(
                """
                INSERT INTO pending_orders (source_file, order_no, route_no, agency_no, dr_code, fg_code, bags_qty, weight_mt, order_ref, status, uploaded_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
                pending_records_to_insert
            )

        if master_routes_to_insert:
            cur.executemany(
                "INSERT OR IGNORE INTO unique_routes_master (file_name, route_no, agency_no, dr_code, created_at) VALUES (?, ?, ?, ?, ?)",
                master_routes_to_insert
            )

        if unmapped_records_to_insert:
            cur.executemany(
                "INSERT OR IGNORE INTO unmapped_missing_dr_ledger (file_name, route_no, agency_no, dr_code, created_at) VALUES (?, ?, ?, ?, ?)",
                unmapped_records_to_insert
            )

        if traceability_records:
            cur.executemany(
                """
                INSERT INTO input_output_traceability (batch_timestamp, input_file_name, input_file_blob, total_input_qty, generated_output_file, output_type, version_no, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
                traceability_records
            )

        cur.execute(
            "INSERT INTO history_logs (timestamp, files_count, total_qty, status, remarks) VALUES (?, ?, ?, 'Success', ?)",
            (batch_ts, len(uploaded_files), total_in_qty, f"Clean Qty Extracted: {total_in_qty}")
        )

        conn.commit()
        conn.close()

        st.session_state.kpi_data = {
            "input_qty": total_in_qty, "gen_qty": total_gen_qty, "valid_count": total_valid,
            "missing_count": total_missing, "skipped_count": total_skipped
        }

        if len(st.session_state.processed_files) > 0:
            st.success(f"🎉 Batch execution completed! Clean Qty: {total_in_qty:,.0f} Bags extracted. Created {len(st.session_state.processed_files)} output workbooks & saved {len(pending_records_to_insert)} pending orders.")
        else:
            st.warning("⚠️ Koi valid demand row extract nahi ho saki.")

    # Multi-Channel Export Panel
    if st.session_state.processed_files:
        st.markdown("---")
        st.subheader("📥 Multi-Channel Export & Notification Hub")
        kpi = st.session_state.kpi_data

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for item in st.session_state.processed_files:
                zf.writestr(item["filename"], item["data"])

        c_exp1, c_exp2, c_exp3, c_exp4, c_exp5, c_exp6, c_exp7 = st.columns(7)
        with c_exp1:
            st.download_button("📦 Download ZIP", zip_buf.getvalue(), f"Batch_Orders_{get_ist_date_str()}.zip", "application/zip")
        with c_exp2:
            try:
                pdf_rep = FPDF()
                pdf_rep.add_page()
                pdf_rep.set_font("Arial", "B", 16)
                pdf_rep.cell(190, 10, "Sales Orders Execution Invoice", ln=True, align="C")
                pdf_rep.set_font("Arial", "", 10)
                pdf_rep.cell(190, 6, f"Generated On (IST): {get_ist_timestamp_full()}", ln=True, align="C")
                pdf_rep.ln(6)
                pdf_rep.set_font("Arial", "B", 10)
                pdf_rep.cell(100, 8, "Metric Description", 1)
                pdf_rep.cell(90, 8, "Value", 1, ln=True)
                pdf_rep.set_font("Arial", "", 10)
                pdf_rep.cell(100, 7, "Total Input Quantity", 1)
                pdf_rep.cell(90, 7, f"{kpi['input_qty']:,.0f} Bags", 1, ln=True)
                pdf_rep.cell(100, 7, "Valid Orders Count", 1)
                pdf_rep.cell(90, 7, str(kpi["valid_count"]), 1, ln=True)
                pdf_rep.cell(100, 7, "Fallback (NEW_CUST) Count", 1)
                pdf_rep.cell(90, 7, str(kpi["missing_count"]), 1, ln=True)
                pdf_rep.cell(100, 7, "Skipped Rows", 1)
                pdf_rep.cell(90, 7, str(kpi["skipped_count"]), 1, ln=True)
                st.download_button("📄 PDF Report", bytes(pdf_rep.output()), f"Execution_Invoice_{get_ist_date_str()}.pdf", "application/pdf")
            except Exception as e:
                st.error(f"PDF Error: {e}")
        with c_exp3:
            txt_summary = f"SALES EXECUTION SUMMARY - {get_ist_timestamp_full()}\nTotal Bags: {kpi['input_qty']}\nValid Orders: {kpi['valid_count']}\nFallback Orders: {kpi['missing_count']}"
            st.download_button("📄 Summary TXT", txt_summary.encode("utf-8"), f"Summary_{get_ist_date_str()}.txt", "text/plain")
        with c_exp4:
            json_dump = json.dumps({"timestamp": get_ist_timestamp_full(), "kpi": kpi}, indent=4)
            st.download_button("💾 Backup JSON", json_dump.encode("utf-8"), f"Audit_{get_ist_date_str()}.json", "application/json")
        with c_exp5:
            components.html("""<button onclick="parent.window.print()" style="width:100%; height:38px; background:#2563eb; color:white; border:none; border-radius:4px; font-weight:600; cursor:pointer;">🖨️ Print View</button>""", height=45)
        with c_exp6:
            if st.button("📧 Email Reports"):
                if st.session_state.email_user and st.session_state.email_pass and st.session_state.recipient:
                    try:
                        msg = EmailMessage()
                        msg["Subject"] = f"🚀 Sales Demand Execution Report - {get_ist_date_str()}"
                        msg["From"] = st.session_state.email_user
                        msg["To"] = st.session_state.recipient
                        msg.set_content(f"Daily Demand Batch processed successfully on {get_ist_timestamp_full()}.\nTotal Bags: {kpi['input_qty']}")
                        for item in st.session_state.processed_files:
                            msg.add_attachment(item["data"], maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=item["filename"])
                        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
                            smtp.login(st.session_state.email_user, st.session_state.email_pass)
                            smtp.send_message(msg)
                        st.success("✅ Email dispatched!")
                    except Exception as e:
                        st.error(f"Email Failed: {e}")
                else:
                    st.warning("⚠️ Enter email credentials in settings!")
        with c_exp7:
            wa_text = f"Sales Batch Processed! Total Bags: {kpi['input_qty']} | Valid: {kpi['valid_count']} | Fallbacks: {kpi['missing_count']}"
            st.markdown(f'<a href="https://wa.me/{st.session_state.whatsapp_num}?text={urllib.parse.quote(wa_text)}" target="_blank"><button style="width:100%; height:38px; background:#25D366; color:white; border:none; border-radius:4px; font-weight:600;">📱 WhatsApp</button></a>', unsafe_allow_html=True)

        st.markdown("##### Individual File Downloads:")
        for idx_f, f_itm in enumerate(st.session_state.processed_files):
            st.download_button(f"📥 Download {f_itm['name']}", f_itm["data"], f_itm["filename"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_indiv_{idx_f}")

# ==============================================================================
# MODULE 2: ROUTE DISPATCH TRIP PLANNER (SKU-LEVEL PARTIAL DISPATCH & OVERLOAD CONFIRMATION)
# ==============================================================================

elif main_menu == "🚚 Route Dispatch Trip Planner":
    st.title("🚚 Route Dispatch Planning & Vehicle Allocation")
    st.markdown("Intelligent vehicle assignment with **SKU-level partial modifications** and **Overload confirmation protocols**.")

    conn = get_db_connection()
    df_pending = pd.read_sql("SELECT * FROM pending_orders WHERE status='Pending'", conn)

    if df_pending.empty:
        st.info("ℹ️ No pending order demand found. Upload demand workbooks in Module 1.")
    else:
        route_summary = df_pending.groupby("route_no").agg({
            "agency_no": "nunique", "bags_qty": "sum", "weight_mt": "sum"
        }).reset_index().rename(columns={"agency_no": "Total Agencies", "bags_qty": "Total Bags", "weight_mt": "Total MT"})

        st.subheader("📦 Pending Demand Clusters Grouped by Route")
        st.dataframe(route_summary, use_container_width=True)

        st.markdown("---")
        col_p1, col_p2 = st.columns(2)
        with col_p1:
            sel_route = st.selectbox("1. Select Route for Dispatch Trip", route_summary["route_no"].tolist())
            route_df = df_pending[df_pending["route_no"] == str(sel_route)].copy()

            avail_fleet = pd.read_sql("SELECT * FROM fleet_master WHERE status='Available'", conn)
            avail_bays = pd.read_sql("SELECT * FROM loading_bays WHERE status='Open'", conn)

            fleet_opts = [f"{r['vehicle_no']} | {r['vehicle_type']} (Cap: {r['capacity_bags']} Bags / {r['capacity_mt']} MT)" for _, r in avail_fleet.iterrows()]
            sel_vehicle = st.selectbox("2. Assign Available Vehicle", fleet_opts if fleet_opts else ["No Vehicles Available"])
            sel_bay = st.selectbox("3. Assign Loading Bay", [f"{r['bay_no']} - {r['bay_name']}" for _, r in avail_bays.iterrows()])

        with col_p2:
            st.markdown("##### 📋 4. Filter Agencies in Route:")
            agencies = route_df["agency_no"].unique().tolist()
            selected_agencies = st.multiselect("Select Agencies to allocate:", agencies, default=agencies)

        st.markdown("---")
        st.subheader("📦 5. Agency & SKU Level Quantity Modification (Include/Exclude/Partial Split):")
        st.markdown("Har Agency ke specific SKU ko include karein ya quantity modify karein (e.g. 30 ki jagah 20 ya 35 bhej sakte hain). Jo quantity bachegi wo automatically **Partial Dispatch Database** me chali jayegi.")

        filtered_route_df = route_df[route_df["agency_no"].isin(selected_agencies)].copy()
        
        # Prepare Interactive Editor Table
        filtered_route_df["Include in Trip"] = True
        filtered_route_df["Dispatch Bags"] = filtered_route_df["bags_qty"]
        
        display_cols = ["id", "Include in Trip", "agency_no", "dr_code", "fg_code", "bags_qty", "Dispatch Bags", "order_no"]
        
        edited_orders = st.data_editor(
            filtered_route_df[display_cols],
            column_config={
                "id": st.column_config.NumberColumn("Item ID", disabled=True),
                "agency_no": st.column_config.TextColumn("Agency", disabled=True),
                "dr_code": st.column_config.TextColumn("DR Code", disabled=True),
                "fg_code": st.column_config.TextColumn("FG Code", disabled=True),
                "bags_qty": st.column_config.NumberColumn("Original Pending Bags", disabled=True),
                "Include in Trip": st.column_config.CheckboxColumn("Include in Vehicle?"),
                "Dispatch Bags": st.column_config.NumberColumn("Dispatch Qty (Bags)", min_value=0.0, step=1.0)
            },
            hide_index=True,
            use_container_width=True,
            key="trip_sku_editor"
        )

        # Calculate Trip Aggregates
        active_trip_items = edited_orders[edited_orders["Include in Trip"] == True].copy()
        trip_bags = active_trip_items["Dispatch Bags"].sum()
        trip_mt = round(trip_bags * 0.05, 2)

        # Dynamic Color Status Badging
        partial_agencies = []
        for ag, group in edited_orders.groupby("agency_no"):
            orig = group["bags_qty"].sum()
            disp = group[group["Include in Trip"] == True]["Dispatch Bags"].sum()
            if disp < orig:
                partial_agencies.append(str(ag))

        if partial_agencies:
            st.markdown(f"""
                <div style="background-color: #fef3c7; border: 1px solid #f59e0b; padding: 10px; border-radius: 6px; margin: 10px 0;">
                    <span style="color: #b45309; font-weight: bold;">⚠️ PARTIAL / SPLIT AGENCIES DETECTED:</span> 
                    <span style="background: #fde68a; color: #92400e; padding: 2px 8px; border-radius: 4px; font-weight: bold;">
                        {', '.join(partial_agencies)}
                    </span> 
                    <br><small style="color: #78350f;">In agencies ka remaining balance automatically alag Partial Dispatch Database me update ho jayega.</small>
                </div>
            """, unsafe_allow_html=True)

        st.markdown("---")
        # Overload Handling Logic
        cap_bags = 0
        is_overloaded = False
        if sel_vehicle != "No Vehicles Available":
            v_num = sel_vehicle.split(" | ")[0]
            v_info = avail_fleet[avail_fleet["vehicle_no"] == v_num].iloc[0]
            cap_bags = int(v_info["capacity_bags"])
            util_pct = (trip_bags / cap_bags * 100) if cap_bags > 0 else 0.0
            
            c_met1, c_met2 = st.columns(2)
            c_met1.metric("Total Dispatch Bags", f"{trip_bags:,.0f} / {cap_bags} Bags", f"{util_pct:.1f}% Capacity Utilization")
            c_met2.metric("Total Tonnage (MT)", f"{trip_mt:,.2f} MT")

            if trip_bags > cap_bags:
                is_overloaded = True
                st.error(f"🚨 **VEHICLE OVERLOAD WARNING:** Truck capacity exceeded by {trip_bags - cap_bags:,.0f} bags ({trip_mt - v_info['capacity_mt']:.2f} MT)!")
                confirm_overload = st.checkbox("⚠️ Check this box to CONFIRM and OVERRIDE vehicle capacity limits for this trip.")
            else:
                confirm_overload = True
                if util_pct < 70:
                    st.warning("⚠️ **Low Utilization Warning:** Truck capacity 70% se kam hai.")
                else:
                    st.success("🟢 **Optimal Load Allocation!**")

        st.markdown("---")
        submit_disabled = (sel_vehicle == "No Vehicles Available") or active_trip_items.empty or (is_overloaded and not confirm_overload)
        
        if st.button("🚀 Confirm Trip & Generate Loading Slip", type="primary", disabled=submit_disabled):
            cur = conn.cursor()
            now_ist = get_ist_now()
            trip_id = f"TRIP-{sel_route}-{now_ist.strftime('%Y%m%d%H%M%S')}"
            v_num = sel_vehicle.split(" | ")[0]
            v_info = avail_fleet[avail_fleet["vehicle_no"] == v_num].iloc[0]
            bay_code = sel_bay.split(" - ")[0]

            # 1. Insert Trip Master
            cur.execute("""
                INSERT INTO trip_loading_slips (trip_id, trip_date, route_no, vehicle_no, transporter_name, driver_name, driver_phone, loading_bay, total_bags, total_weight_mt, capacity_utilization_pct, status, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Planned', ?)
            """, (trip_id, now_ist.strftime("%Y-%m-%d"), str(sel_route), v_num, v_info["transporter_name"], v_info["driver_name"], v_info["driver_phone"], bay_code, trip_bags, trip_mt, round((trip_bags/v_info["capacity_bags"]*100), 2), now_ist.strftime("%Y-%m-%d %H:%M:%S")))

            # 2. Process Items & Split Partials
            seq = 1
            for _, r_val in edited_orders.iterrows():
                item_id = r_val["id"]
                orig_qty = float(r_val["bags_qty"])
                inc = r_val["Include in Trip"]
                disp_qty = float(r_val["Dispatch Bags"]) if inc else 0.0

                # Fetch original item details
                item_row = df_pending[df_pending["id"] == item_id].iloc[0]

                if inc and disp_qty > 0:
                    cur.execute("""
                        INSERT INTO trip_order_items (trip_id, order_no, agency_no, route_no, dr_code, fg_code, allocated_bags, allocated_weight_mt, delivery_seq, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Assigned')
                    """, (trip_id, item_row["order_no"], item_row["agency_no"], item_row["route_no"], item_row["dr_code"], item_row["fg_code"], disp_qty, round(disp_qty*0.05, 2), seq))
                    seq += 1

                # Handle Partial Remaining Logic
                rem_qty = orig_qty - disp_qty
                if rem_qty > 0:
                    # Update pending order with remaining quantity
                    cur.execute("UPDATE pending_orders SET bags_qty=?, weight_mt=?, status='Pending' WHERE id=?", (rem_qty, round(rem_qty*0.05, 2), item_id))
                    # Record in Partial Dispatch Database
                    cur.execute("""
                        INSERT INTO partial_dispatch_ledger (trip_id, source_file, order_no, route_no, agency_no, dr_code, fg_code, original_bags, dispatched_bags, remaining_bags, status, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Partial Pending', ?)
                    """, (trip_id, item_row["source_file"], item_row["order_no"], item_row["route_no"], item_row["agency_no"], item_row["dr_code"], item_row["fg_code"], orig_qty, disp_qty, rem_qty, now_ist.strftime("%Y-%m-%d %H:%M:%S")))
                elif rem_qty == 0 and inc:
                    cur.execute("UPDATE pending_orders SET status='Assigned' WHERE id=?", (item_id,))
                elif rem_qty < 0: # If user modified to more than original
                    cur.execute("UPDATE pending_orders SET status='Assigned' WHERE id=?", (item_id,))
                    cur.execute("""
                        INSERT INTO partial_dispatch_ledger (trip_id, source_file, order_no, route_no, agency_no, dr_code, fg_code, original_bags, dispatched_bags, remaining_bags, status, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Excess Dispatched', ?)
                    """, (trip_id, item_row["source_file"], item_row["order_no"], item_row["route_no"], item_row["agency_no"], item_row["dr_code"], item_row["fg_code"], orig_qty, disp_qty, 0, now_ist.strftime("%Y-%m-%d %H:%M:%S")))

            cur.execute("UPDATE fleet_master SET status='Assigned to Trip' WHERE vehicle_no=?", (v_num,))
            conn.commit()
            st.success(f"🎉 Trip '{trip_id}' Created Successfully! Partials updated in Partial Dispatch Database.")
            st.rerun()
    conn.close()

# ==============================================================================
# MODULE 3: LOADING SLIPS & ACTIVE TRIPS
# ==============================================================================

elif main_menu == "📋 Loading Slips & Active Trips":
    st.title("📋 Trip Slips & Active Vehicle Dispatches")

    conn = get_db_connection()
    df_trips = pd.read_sql("SELECT * FROM trip_loading_slips ORDER BY created_at DESC", conn)

    search_q = st.text_input("🔍 Search Trips (Trip ID, Vehicle, Route, Transporter):", "")
    if search_q:
        df_trips = df_trips[df_trips.apply(lambda r: r.astype(str).str.contains(search_q, case=False).any(), axis=1)]

    st.dataframe(df_trips, use_container_width=True)

    c_t1, c_t2 = st.columns([1, 2])
    with c_t1:
        if not df_trips.empty:
            st.download_button("📥 Export Trips to Excel", to_excel_download_bytes(df_trips, "Trips"), "Dispatch_Trips.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with c_t2:
        with st.expander("🗑️ Multi-Select Delete Trips"):
            del_t_ids = st.multiselect("Select Trip IDs to Delete:", df_trips["trip_id"].tolist() if not df_trips.empty else [])
            if st.button("Delete Selected Trips"):
                cur = conn.cursor()
                for tid in del_t_ids:
                    v = cur.execute("SELECT vehicle_no FROM trip_loading_slips WHERE trip_id=?", (tid,)).fetchone()
                    if v:
                        cur.execute("UPDATE fleet_master SET status='Available' WHERE vehicle_no=?", (v[0],))
                    cur.execute("DELETE FROM trip_order_items WHERE trip_id=?", (tid,))
                    cur.execute("DELETE FROM trip_loading_slips WHERE trip_id=?", (tid,))
                conn.commit()
                st.success("Trips deleted & vehicles released.")
                st.rerun()

    if not df_trips.empty:
        st.markdown("---")
        st.subheader("📄 Inspect Manifest, Generate PDF & Gate Out")
        sel_trip = st.selectbox("Select Trip ID:", df_trips["trip_id"].tolist())
        trip_row = df_trips[df_trips["trip_id"] == sel_trip].iloc[0]
        items_df = pd.read_sql("SELECT * FROM trip_order_items WHERE trip_id=? ORDER BY delivery_seq ASC", conn, params=(sel_trip,))

        st.dataframe(items_df, use_container_width=True)

        btn_c1, btn_c2, btn_c3 = st.columns(3)
        with btn_c1:
            pdf_slip = build_pdf_loading_slip(trip_row.to_dict(), items_df)
            st.download_button("📄 Download PDF Loading Slip", pdf_slip, f"Loading_Slip_{sel_trip}.pdf", "application/pdf")
        with btn_c2:
            wa_text = f"Enterprise Dispatch: Trip {trip_row['trip_id']} | Vehicle {trip_row['vehicle_no']} | Route {trip_row['route_no']} | Bags {trip_row['total_bags']}"
            st.markdown(f'<a href="https://wa.me/{st.session_state.whatsapp_num}?text={urllib.parse.quote(wa_text)}" target="_blank"><button style="width:100%; height:38px; background:#25D366; color:white; border:none; border-radius:4px; font-weight:600;">📱 WhatsApp Alert</button></a>', unsafe_allow_html=True)
        with btn_c3:
            if trip_row["status"] != "Dispatched" and st.button("🏁 Gate Out / Dispatched", type="primary"):
                cur = conn.cursor()
                for _, it in items_df.iterrows():
                    cur.execute("""
                        INSERT INTO daily_dispatch_register (dispatch_date, trip_id, vehicle_no, transporter_name, route_no, agency_no, order_no, dr_code, fg_code, dispatched_bags, dispatched_weight_mt, bay_no, dispatched_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (trip_row["trip_date"], trip_row["trip_id"], trip_row["vehicle_no"], trip_row["transporter_name"], trip_row["route_no"], it["agency_no"], it["order_no"], it["dr_code"], it["fg_code"], it["allocated_bags"], it["allocated_weight_mt"], trip_row["loading_bay"], get_ist_timestamp_full()))
                cur.execute("UPDATE trip_loading_slips SET status='Dispatched' WHERE trip_id=?", (sel_trip,))
                cur.execute("UPDATE fleet_master SET status='Available' WHERE vehicle_no=?", (trip_row["vehicle_no"],))
                conn.commit()
                st.success("✅ Dispatched & recorded in Daily Dispatch Register!")
                st.rerun()
    conn.close()
# ==============================================================================
# ENTERPRISE LOGISTICS, DISPATCH ENGINE & SALES AUTOMATION SUITE
# PART 3: REGISTERS, PARTIALS, PENDING, ARCHIVE, MASTER DB, FLEET, AUDIT, KPI
# & IN-APP DYNAMIC DATABASE BUILDER / CRUD MANAGER
# ==============================================================================

# ==============================================================================
# MODULE 4: DAILY DISPATCH SALE REGISTER
# ==============================================================================

if main_menu == "📖 Daily Dispatch Sale Register":
    st.title("📖 Daily Dispatch Sale Register Database")

    conn = get_db_connection()
    df_reg = pd.read_sql("SELECT * FROM daily_dispatch_register ORDER BY register_id DESC", conn)

    search_r = st.text_input("🔍 Search Register (Agency, Route, Vehicle, DR Code, Order No):", "")
    if search_r:
        df_reg = df_reg[df_reg.apply(lambda r: r.astype(str).str.contains(search_r, case=False).any(), axis=1)]

    st.markdown("##### ✏️ Cell Editing Table (Direct edit karein aur neeche Save button dabayein):")
    edited_reg = st.data_editor(df_reg, use_container_width=True, num_rows="dynamic", key="editor_reg")

    c1, c2, c3 = st.columns([1, 1, 2])
    with c1:
        if st.button("💾 Save Register Changes", type="primary"):
            conn = get_db_connection()
            cur = conn.cursor()
            for _, row in edited_reg.iterrows():
                cur.execute("""
                    UPDATE daily_dispatch_register 
                    SET dispatch_date=?, trip_id=?, vehicle_no=?, transporter_name=?, route_no=?, agency_no=?, order_no=?, dr_code=?, fg_code=?, dispatched_bags=?, dispatched_weight_mt=?, bay_no=?
                    WHERE register_id=?
                """, (row['dispatch_date'], row['trip_id'], row['vehicle_no'], row['transporter_name'], str(row['route_no']), str(row['agency_no']), str(row['order_no']), str(row['dr_code']), str(row['fg_code']), float(row['dispatched_bags']), float(row['dispatched_weight_mt']), str(row['bay_no']), row['register_id']))
            conn.commit()
            conn.close()
            st.success("✅ Register changes successfully saved!")
            st.rerun()

    with c2:
        if not df_reg.empty:
            st.download_button("📥 Export Register to Excel", to_excel_download_bytes(df_reg, "DailyRegister"), "Daily_Dispatch_Sale_Register.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with c3:
        with st.expander("🗑️ Delete Options (Multiple IDs / Route / Date / Purge)"):
            del_mode = st.radio("Choose Delete Mode:", ["Select by IDs", "By Route No", "By Dispatch Date", "⚠️ Purge Complete Register"], horizontal=True)
            if del_mode == "Select by IDs":
                del_r_ids = st.multiselect("Select Register IDs to Delete:", df_reg["register_id"].tolist() if not df_reg.empty else [])
                if st.button("Delete Selected IDs"):
                    cur = conn.cursor()
                    cur.executemany("DELETE FROM daily_dispatch_register WHERE register_id=?", [(i,) for i in del_r_ids])
                    conn.commit()
                    st.success("Selected records deleted.")
                    st.rerun()
            elif del_mode == "By Route No":
                r_to_del = st.text_input("Enter Route No to delete from Register:")
                if st.button("Delete by Route"):
                    cur = conn.cursor()
                    cur.execute("DELETE FROM daily_dispatch_register WHERE route_no=?", (r_to_del,))
                    conn.commit()
                    st.success(f"Route {r_to_del} records deleted.")
                    st.rerun()
            elif del_mode == "By Dispatch Date":
                d_to_del = st.text_input("Enter Date (YYYY-MM-DD) to delete:")
                if st.button("Delete by Date"):
                    cur = conn.cursor()
                    cur.execute("DELETE FROM daily_dispatch_register WHERE dispatch_date=?", (d_to_del,))
                    conn.commit()
                    st.success(f"Date {d_to_del} records deleted.")
                    st.rerun()
            elif del_mode == "⚠️ Purge Complete Register":
                if st.button("🔥 Confirm Wipe Register"):
                    cur = conn.cursor()
                    cur.execute("DELETE FROM daily_dispatch_register")
                    cur.execute("DELETE FROM sqlite_sequence WHERE name='daily_dispatch_register'")
                    conn.commit()
                    st.success("Register completely wiped!")
                    st.rerun()
    conn.close()

# ==============================================================================
# MODULE 5: PARTIAL / SPLIT DISPATCH DATABASE
# ==============================================================================

elif main_menu == "🧩 Partial / Split Dispatch Database":
    st.title("🧩 Partial & Split Dispatch Database")
    st.markdown("Jo orders kisi trip me **partially dispatch** hue hain ya jinka remaining balance bach gaya hai, unka complete historical audit ledger.")

    conn = get_db_connection()
    df_part = pd.read_sql("SELECT * FROM partial_dispatch_ledger ORDER BY id DESC", conn)

    search_part = st.text_input("🔍 Search Partial Ledger (Trip ID, Agency, Route, FG Code):", "")
    if search_part:
        df_part = df_part[df_part.apply(lambda r: r.astype(str).str.contains(search_part, case=False).any(), axis=1)]

    st.markdown("##### ✏️ Cell Editing Partial Table:")
    edited_part = st.data_editor(df_part, use_container_width=True, num_rows="dynamic", key="editor_part")

    c1, c2, c3 = st.columns([1, 1, 2])
    with c1:
        if st.button("💾 Save Partial Changes", type="primary"):
            conn = get_db_connection()
            cur = conn.cursor()
            for _, row in edited_part.iterrows():
                cur.execute("""
                    UPDATE partial_dispatch_ledger 
                    SET trip_id=?, route_no=?, agency_no=?, dr_code=?, fg_code=?, original_bags=?, dispatched_bags=?, remaining_bags=?, status=?
                    WHERE id=?
                """, (row['trip_id'], str(row['route_no']), str(row['agency_no']), str(row['dr_code']), str(row['fg_code']), float(row['original_bags']), float(row['dispatched_bags']), float(row['remaining_bags']), str(row['status']), row['id']))
            conn.commit()
            conn.close()
            st.success("✅ Partial ledger changes saved!")
            st.rerun()

    with c2:
        if not df_part.empty:
            st.download_button("📥 Export Partials to Excel", to_excel_download_bytes(df_part, "Partials"), "Partial_Dispatch_Ledger.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with c3:
        with st.expander("🗑️ Delete Partial Records"):
            del_part_ids = st.multiselect("Select Partial IDs to Delete:", df_part["id"].tolist() if not df_part.empty else [])
            if st.button("Delete Selected Partials"):
                cur = conn.cursor()
                cur.executemany("DELETE FROM partial_dispatch_ledger WHERE id=?", [(i,) for i in del_part_ids])
                conn.commit()
                st.success("Selected partial records deleted.")
                st.rerun()
            if st.button("🔥 Purge Partial Database"):
                cur = conn.cursor()
                cur.execute("DELETE FROM partial_dispatch_ledger")
                cur.execute("DELETE FROM sqlite_sequence WHERE name='partial_dispatch_ledger'")
                conn.commit()
                st.success("Partial ledger wiped.")
                st.rerun()
    conn.close()

# ==============================================================================
# MODULE 6: PENDING ORDERS LEDGER
# ==============================================================================

elif main_menu == "⏳ Pending Orders Ledger":
    st.title("⏳ Pending Orders Database")

    conn = get_db_connection()
    df_p = pd.read_sql("SELECT * FROM pending_orders ORDER BY id DESC", conn)

    search_p = st.text_input("🔍 Search Pending Orders (Order No, Route, Agency, FG Code, DR Code):", "")
    if search_p:
        df_p = df_p[df_p.apply(lambda r: r.astype(str).str.contains(search_p, case=False).any(), axis=1)]

    st.markdown("##### ✏️ Cell Editing Table (Direct edit karein aur neeche Save button dabayein):")
    edited_p = st.data_editor(df_p, use_container_width=True, num_rows="dynamic", key="editor_pending")

    c1, c2, c3 = st.columns([1, 1, 2])
    with c1:
        if st.button("💾 Save Pending Order Changes", type="primary"):
            conn = get_db_connection()
            cur = conn.cursor()
            for _, row in edited_p.iterrows():
                cur.execute("""
                    UPDATE pending_orders 
                    SET order_no=?, route_no=?, agency_no=?, dr_code=?, fg_code=?, bags_qty=?, weight_mt=?, status=?
                    WHERE id=?
                """, (str(row['order_no']), str(row['route_no']), str(row['agency_no']), str(row['dr_code']), str(row['fg_code']), float(row['bags_qty']), float(row['weight_mt']), str(row['status']), row['id']))
            conn.commit()
            conn.close()
            st.success("✅ Pending orders successfully saved!")
            st.rerun()

    with c2:
        if not df_p.empty:
            st.download_button("📥 Export Pending to Excel", to_excel_download_bytes(df_p, "PendingOrders"), "Pending_Orders.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with c3:
        with st.expander("🗑️ Delete Options (Multiple IDs / Route / Source File / Purge)"):
            del_p_mode = st.radio("Choose Delete Mode:", ["Select by IDs", "By Route No", "By Source File", "⚠️ Purge All Pending"], horizontal=True)
            if del_p_mode == "Select by IDs":
                del_p_ids = st.multiselect("Select Order IDs to Delete:", df_p["id"].tolist() if not df_p.empty else [])
                if st.button("Delete Selected Orders"):
                    cur = conn.cursor()
                    cur.executemany("DELETE FROM pending_orders WHERE id=?", [(i,) for i in del_p_ids])
                    conn.commit()
                    st.success("Selected orders deleted.")
                    st.rerun()
            elif del_p_mode == "By Route No":
                r_del = st.text_input("Enter Route No to delete from Pending:")
                if st.button("Delete Pending by Route"):
                    cur = conn.cursor()
                    cur.execute("DELETE FROM pending_orders WHERE route_no=?", (r_del,))
                    conn.commit()
                    st.success(f"Route {r_del} pending orders deleted.")
                    st.rerun()
            elif del_p_mode == "By Source File":
                f_del = st.selectbox("Select File to purge:", df_p["source_file"].unique().tolist() if not df_p.empty else ["None"])
                if st.button("Delete File Orders"):
                    cur = conn.cursor()
                    cur.execute("DELETE FROM pending_orders WHERE source_file=?", (f_del,))
                    conn.commit()
                    st.success(f"File {f_del} orders removed.")
                    st.rerun()
            elif del_p_mode == "⚠️ Purge All Pending":
                if st.button("🔥 Confirm Wipe All Pending"):
                    cur = conn.cursor()
                    cur.execute("DELETE FROM pending_orders")
                    cur.execute("DELETE FROM sqlite_sequence WHERE name='pending_orders'")
                    conn.commit()
                    st.success("All pending orders wiped!")
                    st.rerun()
    conn.close()

# ==============================================================================
# MODULE 7: FILE UPLOAD ARCHIVE
# ==============================================================================

elif main_menu == "🗄️ File Upload Archive":
    st.title("🗄️ Uploaded Input File Archive")

    conn = get_db_connection()
    df_a = pd.read_sql("SELECT * FROM uploaded_files_archive ORDER BY id DESC", conn)

    search_a = st.text_input("🔍 Search Archive Logs:", "")
    if search_a:
        df_a = df_a[df_a.apply(lambda r: r.astype(str).str.contains(search_a, case=False).any(), axis=1)]

    st.dataframe(df_a, use_container_width=True)

    c1, c2 = st.columns([1, 2])
    with c1:
        if not df_a.empty:
            st.download_button("📥 Export Archive to Excel", to_excel_download_bytes(df_a, "ArchiveLogs"), "Uploaded_Archive_Logs.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with c2:
        with st.expander("🗑️ Delete Archive Options"):
            del_a_ids = st.multiselect("Select File IDs to Delete:", df_a["id"].tolist() if not df_a.empty else [])
            if st.button("Delete Selected Archive Logs"):
                cur = conn.cursor()
                cur.executemany("DELETE FROM uploaded_files_archive WHERE id=?", [(i,) for i in del_a_ids])
                conn.commit()
                st.success("Logs deleted.")
                st.rerun()
            if st.button("🔥 Purge All Archive"):
                cur = conn.cursor()
                cur.execute("DELETE FROM uploaded_files_archive")
                cur.execute("DELETE FROM sqlite_sequence WHERE name='uploaded_files_archive'")
                conn.commit()
                st.success("Archive wiped.")
                st.rerun()
    conn.close()

# ==============================================================================
# MODULE 8: MASTER DB & UNMAPPED LEDGER (WITH BULK EXCEL IMPORTER)
# ==============================================================================

elif main_menu == "📋 Master DB & Unmapped Ledger":
    st.title("📋 Unique Master Mapping DB & Unmapped Fallback Ledger")

    conn = get_db_connection()
    df_m = pd.read_sql("SELECT * FROM unique_routes_master ORDER BY id DESC", conn)
    df_u = pd.read_sql("SELECT * FROM unmapped_missing_dr_ledger ORDER BY id DESC", conn)

    t1, t2 = st.tabs(["📋 Unique Master Mapping DB", "🚨 Unmapped Missing DR Ledger"])

    with t1:
        # DATA IMPORT OPTION FOR MASTER ROUTES
        with st.expander("📥 Import Master Routes via Excel / CSV File"):
            up_master = st.file_uploader("Upload Master Routes File (Must contain route_no, agency_no, dr_code)", type=["xlsx", "csv"], key="master_file_import")
            if up_master and st.button("🚀 Process & Import Master File"):
                try:
                    if up_master.name.endswith('.csv'):
                        df_imp = pd.read_csv(up_master)
                    else:
                        df_imp = pd.read_excel(up_master)
                    
                    df_imp.columns = [str(c).strip().lower().replace(' ', '_') for c in df_imp.columns]
                    if all(req in df_imp.columns for req in ['route_no', 'agency_no', 'dr_code']):
                        imp_records = []
                        now_ts = get_ist_timestamp_full()
                        for _, row in df_imp.iterrows():
                            imp_records.append((up_master.name, str(row['route_no']).strip(), str(row['agency_no']).replace('.0','').strip(), str(row['dr_code']).strip(), now_ts))
                        
                        cur = conn.cursor()
                        cur.executemany("INSERT OR REPLACE INTO unique_routes_master (file_name, route_no, agency_no, dr_code, created_at) VALUES (?, ?, ?, ?, ?)", imp_records)
                        conn.commit()
                        st.success(f"✅ Successfully imported {len(imp_records)} master records!")
                        st.rerun()
                    else:
                        st.error("❌ Required columns missing! File must have: 'route_no', 'agency_no', 'dr_code'")
                except Exception as ex:
                    st.error(f"Import Error: {str(ex)}")

        sm = st.text_input("🔍 Search Master Database:", "", key="search_m")
        if sm:
            df_m = df_m[df_m.apply(lambda r: r.astype(str).str.contains(sm, case=False).any(), axis=1)]
        
        st.markdown("##### ✏️ Cell Editing Table (Direct edit karein aur neeche Save button dabayein):")
        edited_m = st.data_editor(df_m, use_container_width=True, num_rows="dynamic", key="editor_master")

        c1, c2, c3 = st.columns([1, 1, 2])
        with c1:
            if st.button("💾 Save Master DB Changes", type="primary"):
                cur = conn.cursor()
                for _, row in edited_m.iterrows():
                    cur.execute("""
                        UPDATE unique_routes_master 
                        SET route_no=?, agency_no=?, dr_code=?
                        WHERE id=?
                    """, (str(row['route_no']), str(row['agency_no']), str(row['dr_code']), row['id']))
                conn.commit()
                st.success("✅ Master DB changes saved!")
                st.rerun()
        with c2:
            st.download_button("📥 Export Master to Excel", to_excel_download_bytes(df_m, "MasterRoutes"), "Unique_Routes_Master.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with c3:
            with st.expander("🗑️ Delete Master Options"):
                del_m = st.multiselect("Select IDs to Delete:", df_m["id"].tolist() if not df_m.empty else [])
                if st.button("Delete Master Records"):
                    cur = conn.cursor()
                    cur.executemany("DELETE FROM unique_routes_master WHERE id=?", [(i,) for i in del_m])
                    conn.commit()
                    st.success("Master records deleted.")
                    st.rerun()
                if st.button("🔥 Purge All Master DB"):
                    cur = conn.cursor()
                    cur.execute("DELETE FROM unique_routes_master")
                    cur.execute("DELETE FROM sqlite_sequence WHERE name='unique_routes_master'")
                    conn.commit()
                    st.success("Master DB wiped.")
                    st.rerun()

        with st.expander("➕ Add Single Record to Master DB"):
            m_c1, m_c2, m_c3 = st.columns(3)
            with m_c1: man_r = st.text_input("Route No", "10")
            with m_c2: man_a = st.text_input("Agency No", "")
            with m_c3: man_dr = st.text_input("DR Code", "")
            if st.button("Save Master Record"):
                if man_r and man_a and man_dr:
                    cur = conn.cursor()
                    cur.execute("INSERT OR REPLACE INTO unique_routes_master (file_name, route_no, agency_no, dr_code, created_at) VALUES ('Manual_Entry', ?, ?, ?, ?)", (man_r, man_a, man_dr, get_ist_timestamp_full()))
                    conn.commit()
                    st.success("Saved!")
                    st.rerun()

    with t2:
        su = st.text_input("🔍 Search Unmapped Ledger:", "", key="search_u")
        if su:
            df_u = df_u[df_u.apply(lambda r: r.astype(str).str.contains(su, case=False).any(), axis=1)]
        st.dataframe(df_u, use_container_width=True)

        c1, c2 = st.columns([1, 2])
        with c1:
            st.download_button("📥 Export Unmapped to Excel", to_excel_download_bytes(df_u, "UnmappedDR"), "Unmapped_Missing_DR.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with c2:
            with st.expander("🗑️ Delete Unmapped Options"):
                del_u = st.multiselect("Select Unmapped IDs:", df_u["id"].tolist() if not df_u.empty else [])
                if st.button("Delete Unmapped Records"):
                    cur = conn.cursor()
                    cur.executemany("DELETE FROM unmapped_missing_dr_ledger WHERE id=?", [(i,) for i in del_u])
                    conn.commit()
                    st.success("Unmapped records deleted.")
                    st.rerun()
                if st.button("🔥 Purge Unmapped Ledger"):
                    cur = conn.cursor()
                    cur.execute("DELETE FROM unmapped_missing_dr_ledger")
                    cur.execute("DELETE FROM sqlite_sequence WHERE name='unmapped_missing_dr_ledger'")
                    conn.commit()
                    st.success("Unmapped Ledger wiped.")
                    st.rerun()
    conn.close()

# ==============================================================================
# MODULE 9: FLEET & LOADING BAY MASTER (WITH BULK EXCEL IMPORTER)
# ==============================================================================

elif main_menu == "🚛 Fleet & Loading Bay Master":
    st.title("🚛 Transporter Fleet Master & Loading Bay Configurations")

    conn = get_db_connection()
    df_f = pd.read_sql("SELECT * FROM fleet_master", conn)
    df_b = pd.read_sql("SELECT * FROM loading_bays", conn)

    tab1, tab2 = st.tabs(["🚛 Fleet Master", "🏭 Loading Bays"])

    with tab1:
        # DATA IMPORT OPTION FOR FLEET MASTER
        with st.expander("📥 Import Vehicle Master List via Excel / CSV File"):
            up_fleet = st.file_uploader("Upload Fleet Master List (Columns: vehicle_no, vehicle_type, capacity_bags, capacity_mt, transporter_name, driver_name, driver_phone)", type=["xlsx", "csv"], key="fleet_import_file")
            if up_fleet and st.button("🚀 Process & Import Fleet Master"):
                try:
                    if up_fleet.name.endswith('.csv'):
                        df_fl_imp = pd.read_csv(up_fleet)
                    else:
                        df_fl_imp = pd.read_excel(up_fleet)
                    
                    df_fl_imp.columns = [str(c).strip().lower().replace(' ', '_') for c in df_fl_imp.columns]
                    if all(req in df_fl_imp.columns for req in ['vehicle_no', 'capacity_bags']):
                        cur = conn.cursor()
                        for _, row in df_fl_imp.iterrows():
                            v_no = str(row['vehicle_no']).strip()
                            v_type = str(row.get('vehicle_type', '10 Wheeler Truck')).strip()
                            cap_b = int(row['capacity_bags'])
                            cap_m = float(row.get('capacity_mt', cap_b * 0.05))
                            trans = str(row.get('transporter_name', 'National Logistics')).strip()
                            driver = str(row.get('driver_name', 'Driver')).strip()
                            phone = str(row.get('driver_phone', '9876543210')).replace('.0','').strip()
                            
                            cur.execute("""
                                INSERT OR REPLACE INTO fleet_master (vehicle_no, vehicle_type, capacity_bags, capacity_mt, transporter_name, driver_name, driver_phone, status)
                                VALUES (?, ?, ?, ?, ?, ?, ?, 'Available')
                            """, (v_no, v_type, cap_b, cap_m, trans, driver, phone))
                        conn.commit()
                        st.success(f"✅ Successfully imported {len(df_fl_imp)} vehicles into Fleet Master!")
                        st.rerun()
                    else:
                        st.error("❌ Required columns missing! File must have at least: 'vehicle_no', 'capacity_bags'")
                except Exception as ex:
                    st.error(f"Fleet Import Error: {str(ex)}")

        st.markdown("##### ✏️ Cell Editing Fleet Table (Direct edit karein aur neeche Save button dabayein):")
        edited_f = st.data_editor(df_f, use_container_width=True, num_rows="dynamic", key="editor_fleet")

        c1, c2, c3 = st.columns([1, 1, 2])
        with c1:
            if st.button("💾 Save Fleet Changes", type="primary"):
                cur = conn.cursor()
                for _, row in edited_f.iterrows():
                    cur.execute("""
                        UPDATE fleet_master 
                        SET vehicle_no=?, vehicle_type=?, capacity_bags=?, capacity_mt=?, transporter_name=?, driver_name=?, driver_phone=?, status=?
                        WHERE id=?
                    """, (str(row['vehicle_no']), str(row['vehicle_type']), int(row['capacity_bags']), float(row['capacity_mt']), str(row['transporter_name']), str(row['driver_name']), str(row['driver_phone']), str(row['status']), row['id']))
                conn.commit()
                st.success("✅ Fleet master changes saved!")
                st.rerun()
        with c2:
            st.download_button("📥 Export Fleet to Excel", to_excel_download_bytes(df_f, "Fleet"), "Fleet_Master.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with c3:
            with st.expander("🗑️ Delete Vehicles"):
                del_f = st.multiselect("Select Vehicle IDs:", df_f["id"].tolist() if not df_f.empty else [])
                if st.button("Delete Vehicles"):
                    cur = conn.cursor()
                    cur.executemany("DELETE FROM fleet_master WHERE id=?", [(i,) for i in del_f])
                    conn.commit()
                    st.rerun()

        with st.expander("➕ Add Single Vehicle"):
            v1, v2, v3 = st.columns(3)
            with v1:
                v_num = st.text_input("Vehicle No (e.g. PB-10-AZ-9988)")
                v_type = st.selectbox("Vehicle Type", ["10 Wheeler Truck", "12 Wheeler Multi-Axle", "14 Wheeler Heavy", "Canter / Eicher"])
            with v2:
                v_bags = st.number_input("Capacity (Bags)", value=500, step=50)
                v_mt = st.number_input("Capacity (MT)", value=25.0, step=1.0)
            with v3:
                v_trans = st.text_input("Transporter Name")
                v_driver = st.text_input("Driver Name")
                v_phone = st.text_input("Driver Phone")
            if st.button("Save Vehicle"):
                if v_num:
                    cur = conn.cursor()
                    cur.execute("INSERT OR REPLACE INTO fleet_master (vehicle_no, vehicle_type, capacity_bags, capacity_mt, transporter_name, driver_name, driver_phone, status) VALUES (?, ?, ?, ?, ?, ?, ?, 'Available')", (v_num, v_type, v_bags, v_mt, v_trans, v_driver, v_phone))
                    conn.commit()
                    st.rerun()

    with tab2:
        st.markdown("##### ✏️ Cell Editing Loading Bays Table:")
        edited_b = st.data_editor(df_b, use_container_width=True, num_rows="dynamic", key="editor_bays")

        c1, c2, c3 = st.columns([1, 1, 2])
        with c1:
            if st.button("💾 Save Bays Changes", type="primary"):
                cur = conn.cursor()
                for _, row in edited_b.iterrows():
                    cur.execute("""
                        UPDATE loading_bays 
                        SET bay_no=?, bay_name=?, status=?
                        WHERE id=?
                    """, (str(row['bay_no']), str(row['bay_name']), str(row['status']), row['id']))
                conn.commit()
                st.success("✅ Loading bays changes saved!")
                st.rerun()
        with c2:
            st.download_button("📥 Export Bays to Excel", to_excel_download_bytes(df_b, "Bays"), "Loading_Bays.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with c3:
            with st.expander("🗑️ Delete Bays"):
                del_b = st.multiselect("Select Bay IDs:", df_b["id"].tolist() if not df_b.empty else [])
                if st.button("Delete Bays"):
                    cur = conn.cursor()
                    cur.executemany("DELETE FROM loading_bays WHERE id=?", [(i,) for i in del_b])
                    conn.commit()
                    st.rerun()

        with st.expander("➕ Add Bay"):
            b1, b2 = st.columns(2)
            with b1: bay_no = st.text_input("Bay Code (e.g. BAY-04)")
            with b2: bay_name = st.text_input("Location Name")
            if st.button("Save Bay"):
                if bay_no:
                    cur = conn.cursor()
                    cur.execute("INSERT OR REPLACE INTO loading_bays (bay_no, bay_name, status) VALUES (?, ?, 'Open')", (bay_no, bay_name))
                    conn.commit()
                    st.rerun()
    conn.close()

# ==============================================================================
# MODULE 10: TRACEABILITY & AUDIT LEDGERS
# ==============================================================================

elif main_menu == "🔍 Traceability & Audit Ledgers":
    st.title("🔍 Input-Output Traceability, Discrepancies & Stored Outputs")

    conn = get_db_connection()
    df_trace = pd.read_sql("SELECT id, batch_timestamp, input_file_name, total_input_qty, generated_output_file, output_type, version_no FROM input_output_traceability ORDER BY id DESC", conn)
    df_stored = pd.read_sql("SELECT id, file_name, file_type, created_at FROM output_files_ledger ORDER BY id DESC", conn)
    df_disc = pd.read_sql("SELECT * FROM discrepancy_audit_ledger ORDER BY id DESC", conn)

    t_tab1, t_tab2, t_tab3 = st.tabs(["🔗 Input-Output Traceability", "📦 Stored Output Files", "🔍 Discrepancy Audit"])

    with t_tab1:
        st.dataframe(df_trace, use_container_width=True)
        if not df_trace.empty:
            st.download_button("📥 Export Traceability to Excel", to_excel_download_bytes(df_trace, "Traceability"), "Traceability_Ledger.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            with st.expander("🗑️ Multi-Delete Traceability"):
                del_tr = st.multiselect("Select IDs:", df_trace["id"].tolist())
                if st.button("Delete Traceability"):
                    cur = conn.cursor()
                    cur.executemany("DELETE FROM input_output_traceability WHERE id=?", [(i,) for i in del_tr])
                    conn.commit()
                    st.rerun()

    with t_tab2:
        st.dataframe(df_stored, use_container_width=True)
        if not df_stored.empty:
            sel_out_id = st.number_input("Enter Stored Output ID to Download", min_value=1, step=1)
            if st.button("Download Stored File"):
                cur = conn.cursor()
                res = cur.execute("SELECT file_name, file_data FROM output_files_ledger WHERE id=?", (sel_out_id,)).fetchone()
                if res:
                    st.download_button(f"💾 Save {res[0]}", res[1], res[0], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                else:
                    st.warning("File not found.")

    with t_tab3:
        st.dataframe(df_disc, use_container_width=True)
        if df_disc.empty:
            st.success("🟢 No discrepancies logged.")
    conn.close()

# ==============================================================================
# MODULE 11: EXECUTIVE KPI & VISUAL ANALYTICS
# ==============================================================================

elif main_menu == "📊 Executive KPI & Visual Analytics":
    st.title("📊 Supply Chain & Dispatch KPI Analytics Dashboard")

    conn = get_db_connection()
    df_trips = pd.read_sql("SELECT * FROM trip_loading_slips", conn)
    df_pending = pd.read_sql("SELECT * FROM pending_orders", conn)
    df_reg = pd.read_sql("SELECT * FROM daily_dispatch_register", conn)
    df_part = pd.read_sql("SELECT * FROM partial_dispatch_ledger", conn)
    conn.close()

    col_k1, col_k2, col_k3, col_k4 = st.columns(4)
    tot_trips = len(df_trips)
    tot_dispatched_bags = df_reg["dispatched_bags"].sum() if not df_reg.empty else 0.0
    avg_util = df_trips["capacity_utilization_pct"].mean() if not df_trips.empty else 0.0
    active_pending_bags = df_pending[df_pending["status"] == "Pending"]["bags_qty"].sum() if not df_pending.empty else 0.0

    col_k1.metric("Total Trips Planned", tot_trips)
    col_k2.metric("Dispatched Bags", f"{tot_dispatched_bags:,.0f}")
    col_k3.metric("Avg Fleet Utilization", f"{avg_util:.1f}%")
    col_k4.metric("Active Pending Load", f"{active_pending_bags:,.0f} Bags")

    st.markdown("---")
    c_ch1, c_ch2 = st.columns(2)
    with c_ch1:
        st.markdown("##### 🚛 Route-wise Dispatched Load (MT)")
        if not df_reg.empty:
            st.bar_chart(df_reg.groupby("route_no")["dispatched_weight_mt"].sum())
        else:
            st.info("No dispatched register records.")
    with c_ch2:
        st.markdown("##### 🧩 Partial / Split Orders Breakdown")
        if not df_part.empty:
            st.bar_chart(df_part.groupby("agency_no")["remaining_bags"].sum())
        else:
            st.info("No partial orders recorded yet.")

# ==============================================================================
# MODULE 12: IN-APP DATABASE BUILDER & DYNAMIC LINKER (NEW ADDED MODULE)
# ==============================================================================

elif main_menu == "🗄️ In-App Database Builder & Dynamic Linker":
    st.title("🗄️ In-App Dynamic Database Builder & Universal CRUD")
    st.markdown("Create custom tables inside the application, add columns dynamically, insert/modify/delete records, and manage relational database schemas.")

    conn = get_db_connection()
    cur = conn.cursor()

    tab_db1, tab_db2 = st.tabs(["🏗️ Create Table & Add Columns", "📋 Universal Table CRUD & Manager"])

    with tab_db1:
        st.markdown("##### ➕ Create New Custom Table")
        new_tbl_name = st.text_input("New Table Name (e.g. vendor_master, quality_check)", "").strip().lower()
        new_tbl_name = re.sub(r'[^a-z0-9_]', '', new_tbl_name)
        
        link_to_existing = st.checkbox("Link with Existing Table (Foreign Key)")
        parent_table = ""
        if link_to_existing:
            parent_table = st.selectbox("Select Parent Table", ["pending_orders", "fleet_master", "unique_routes_master"])

        if st.button("🏗️ Build & Create Table", type="primary"):
            if new_tbl_name:
                try:
                    if link_to_existing and parent_table:
                        cur.execute(f"""
                            CREATE TABLE IF NOT EXISTS {new_tbl_name} (
                                id INTEGER PRIMARY KEY AUTOINCREMENT,
                                linked_parent_id INTEGER,
                                custom_key TEXT,
                                custom_value TEXT,
                                notes TEXT,
                                created_at TEXT,
                                FOREIGN KEY (linked_parent_id) REFERENCES {parent_table}(id) ON DELETE CASCADE
                            )
                        """)
                    else:
                        cur.execute(f"""
                            CREATE TABLE IF NOT EXISTS {new_tbl_name} (
                                id INTEGER PRIMARY KEY AUTOINCREMENT,
                                custom_key TEXT,
                                custom_value TEXT,
                                notes TEXT,
                                created_at TEXT
                            )
                        """)
                    conn.commit()
                    st.success(f"✅ Table '{new_tbl_name}' created successfully!")
                    st.rerun()
                except Exception as ex:
                    st.error(f"Error: {ex}")

        st.markdown("---")
        st.markdown("##### ➕ Add New Column to Existing Table")
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'")
        all_tables_list = [row[0] for row in cur.fetchall()]
        
        col_target_tbl = st.selectbox("Select Table to Alter", all_tables_list, key="alter_tbl_sel")
        new_col_name = st.text_input("New Column Name (e.g. priority_level, remarks)").strip().lower()
        new_col_name = re.sub(r'[^a-z0-9_]', '', new_col_name)
        new_col_type = st.selectbox("Column Data Type", ["TEXT", "REAL", "INTEGER"])

        if st.button("➕ Add Column", type="primary"):
            if col_target_tbl and new_col_name:
                try:
                    cur.execute(f"ALTER TABLE {col_target_tbl} ADD COLUMN {new_col_name} {new_col_type}")
                    conn.commit()
                    st.success(f"✅ Column '{new_col_name}' ({new_col_type}) added to '{col_target_tbl}' successfully!")
                    st.rerun()
                except Exception as ex:
                    st.error(f"Error adding column: {ex}")

    with tab_db2:
        st.markdown("##### 📋 View, Edit, Modify & Delete Records")
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'")
        all_tables_list = [row[0] for row in cur.fetchall()]
        
        selected_tbl = st.selectbox("Select Table to View/Edit", all_tables_list, key="crud_tbl_sel")

        if selected_tbl:
            df_selected = pd.read_sql(f"SELECT * FROM {selected_tbl}", conn)
            edited_custom_tbl = st.data_editor(df_selected, use_container_width=True, num_rows="dynamic", key=f"crud_{selected_tbl}")

            col_act1, col_act2, col_act3 = st.columns(3)
            with col_act1:
                if st.button(f"💾 Save Changes in `{selected_tbl}`", type="primary"):
                    try:
                        for _, row in edited_custom_tbl.iterrows():
                            row_dict = row.to_dict()
                            row_id = row_dict.get('id')
                            
                            if pd.notna(row_id):
                                cols_to_update = [k for k in row_dict.keys() if k != 'id']
                                set_clause = ", ".join([f"{c}=?" for c in cols_to_update])
                                vals = [row_dict[c] for c in cols_to_update] + [row_id]
                                cur.execute(f"UPDATE {selected_tbl} SET {set_clause} WHERE id=?", vals)
                            else:
                                cols_to_insert = [k for k in row_dict.keys() if k != 'id' and pd.notna(row_dict[k])]
                                placeholders = ", ".join(["?" for _ in cols_to_insert])
                                cols_str = ", ".join(cols_to_insert)
                                vals = [row_dict[c] for c in cols_to_insert]
                                cur.execute(f"INSERT INTO {selected_tbl} ({cols_str}) VALUES ({placeholders})", vals)
                        
                        conn.commit()
                        st.success(f"✅ Table `{selected_tbl}` updated!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Failed: {e}")

            with col_act2:
                if not df_selected.empty:
                    st.download_button("📥 Export Table (.xlsx)", to_excel_download_bytes(df_selected, selected_tbl), f"{selected_tbl}_export.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            with col_act3:
                if st.button(f"🔥 Drop Table `{selected_tbl}`", type="secondary"):
                    if selected_tbl not in ["pending_orders", "fleet_master", "unique_routes_master"]:
                        cur.execute(f"DROP TABLE IF EXISTS {selected_tbl}")
                        conn.commit()
                        st.warning(f"⚠️ Table `{selected_tbl}` dropped.")
                        st.rerun()
                    else:
                        st.error("❌ Core system tables cannot be dropped!")

    conn.close()

# ==============================================================================
# SECTION: THREE INTEGRATED EXTENSIONS (OPTIMIZER, FILTER & LIVE STOCK)
# (CLEAN & NON-DUPLICATE INTEGRATED SUITE - APPENDED AT END OF FILE)
# ==============================================================================

# ------------------------------------------------------------------------------
# 1. SMART MULTI-TRUCK LOAD OPTIMIZER PRO
# ------------------------------------------------------------------------------
if main_menu == "⚡ Smart Multi-Truck Load Optimizer Pro":
    st.title("⚡ Smart Multi-Vehicle Dispatch Optimizer Pro")
    st.markdown("Automated **Bin-Packing Algorithm** jo pending route load ko truck capacity ke hisab se exact fit batches (Trip-1, Trip-2...) me auto-split karta hai.")

    conn_opt = get_db_connection()
    cur_opt = conn_opt.cursor()

    df_p_all = pd.read_sql("SELECT * FROM pending_orders WHERE status='Pending'", conn_opt)
    df_fleet_all = pd.read_sql("SELECT * FROM fleet_master WHERE status='Available'", conn_opt)
    df_bays_all = pd.read_sql("SELECT * FROM loading_bays WHERE status='Open'", conn_opt)

    if df_p_all.empty:
        st.info("ℹ️ Koi pending demand nahi mili. Pehle Module 1 me demand files upload karein.")
    else:
        route_list = sorted(df_p_all["route_no"].unique().tolist())
        
        c_r1, c_r2, c_r3 = st.columns([1, 1, 1])
        with c_r1:
            sel_opt_route = st.selectbox("1. Select Route to Auto-Optimize:", route_list, key="opt_route_sel_unified")
        
        route_orders = df_p_all[df_p_all["route_no"] == str(sel_opt_route)].copy()
        tot_route_bags = route_orders["bags_qty"].sum()
        tot_route_mt = round(tot_route_bags * 0.05, 2)
        tot_agencies = route_orders["agency_no"].nunique()

        with c_r2:
            st.metric("Total Route Load", f"{tot_route_bags:,.0f} Bags", f"{tot_route_mt:.2f} MT")
        with c_r3:
            st.metric("Agencies in Route", f"{tot_agencies} Dealers")

        st.markdown("---")
        st.subheader("🚛 2. Select Vehicles for Smart Load Distribution")

        if df_fleet_all.empty:
            st.warning("⚠️ Koi 'Available' vehicle nahi mili. Fleet Master me status check karein.")
        else:
            avail_truck_choices = [
                f"{r['vehicle_no']} | {r['vehicle_type']} (Max: {r['capacity_bags']} Bags)" 
                for _, r in df_fleet_all.iterrows()
            ]
            
            sel_trucks_multi = st.multiselect(
                "Assign Available Trucks for this Route:",
                avail_truck_choices,
                default=avail_truck_choices[:min(3, len(avail_truck_choices))],
                key="opt_trucks_multi_unified"
            )

            if sel_trucks_multi and st.button("🤖 Run Auto-Load Balancing & Pack Trucks", type="primary", key="btn_run_pack_unified"):
                agency_grp = route_orders.groupby(["agency_no", "dr_code"]).agg({
                    "bags_qty": "sum"
                }).reset_index().sort_values(by="bags_qty", ascending=False)

                trucks_pool = []
                for t_str in sel_trucks_multi:
                    v_num = t_str.split(" | ")[0]
                    v_row = df_fleet_all[df_fleet_all["vehicle_no"] == v_num].iloc[0]
                    trucks_pool.append({
                        "vehicle_no": v_num,
                        "vehicle_type": v_row["vehicle_type"],
                        "transporter_name": v_row["transporter_name"],
                        "driver_name": v_row["driver_name"],
                        "driver_phone": v_row["driver_phone"],
                        "capacity_bags": int(v_row["capacity_bags"]),
                        "allocated_bags": 0.0,
                        "allocated_agencies": [],
                        "items": []
                    })

                unallocated_agencies = []
                for _, ag_row in agency_grp.iterrows():
                    ag_no = ag_row["agency_no"]
                    ag_dr = ag_row["dr_code"]
                    ag_bags = float(ag_row["bags_qty"])

                    placed = False
                    for trk in trucks_pool:
                        if (trk["allocated_bags"] + ag_bags) <= trk["capacity_bags"]:
                            trk["allocated_bags"] += ag_bags
                            trk["allocated_agencies"].append(ag_no)
                            sku_items = route_orders[route_orders["agency_no"] == ag_no].copy()
                            for _, itm in sku_items.iterrows():
                                trk["items"].append(itm)
                            placed = True
                            break
                    if not placed:
                        unallocated_agencies.append(ag_no)

                st.session_state["optimized_plan"] = {
                    "trucks": trucks_pool,
                    "unallocated": unallocated_agencies,
                    "route": sel_opt_route
                }

            if "optimized_plan" in st.session_state and st.session_state["optimized_plan"]["route"] == sel_opt_route:
                plan = st.session_state["optimized_plan"]
                st.markdown("---")
                st.subheader("📋 Generated Optimal Trip Manifests")

                trip_tabs = st.tabs([f"🚚 Trip {i+1}: {t['vehicle_no']} ({t['allocated_bags']:,.0f}/{t['capacity_bags']} Bags)" for i, t in enumerate(plan["trucks"])])

                for idx, tab_ui in enumerate(trip_tabs):
                    trk_info = plan["trucks"][idx]
                    with tab_ui:
                        util_pct = (trk_info["allocated_bags"] / trk_info["capacity_bags"] * 100) if trk_info["capacity_bags"] > 0 else 0
                        
                        m_c1, m_c2, m_c3 = st.columns(3)
                        m_c1.metric("Load Assigned", f"{trk_info['allocated_bags']:,.0f} Bags", f"{util_pct:.1f}% Utilization")
                        m_c2.metric("Total Weight", f"{trk_info['allocated_bags'] * 0.05:.2f} MT")
                        m_c3.metric("Agencies Covered", f"{len(trk_info['allocated_agencies'])} Dealers")

                        if util_pct > 100:
                            st.error("🚨 Overloaded!")
                        elif util_pct >= 85:
                            st.success("🟢 Perfect Utilization!")
                        else:
                            st.warning("🟡 Capacity Under-utilized")

                        if trk_info["items"]:
                            trip_items_df = pd.DataFrame(trk_info["items"])[["agency_no", "dr_code", "fg_code", "bags_qty", "order_no"]]
                            st.dataframe(trip_items_df, use_container_width=True)

                            bay_sel = st.selectbox(f"Select Loading Bay for {trk_info['vehicle_no']}:", [f"{r['bay_no']} - {r['bay_name']}" for _, r in df_bays_all.iterrows()], key=f"bay_sel_opt_{idx}")

                            if st.button(f"🚀 Confirm & Lock Load for Trip {idx+1} ({trk_info['vehicle_no']})", key=f"btn_cfm_trip_unified_{idx}"):
                                now_dt = get_ist_now()
                                trip_code = f"TRIP-{sel_opt_route}-{now_dt.strftime('%Y%m%d%H%M%S')}-{idx+1}"
                                bay_id = bay_sel.split(" - ")[0]

                                cur_opt.execute("""
                                    INSERT INTO trip_loading_slips (trip_id, trip_date, route_no, vehicle_no, transporter_name, driver_name, driver_phone, loading_bay, total_bags, total_weight_mt, capacity_utilization_pct, status, created_at)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Planned', ?)
                                """, (
                                    trip_code, now_dt.strftime("%Y-%m-%d"), str(sel_opt_route),
                                    trk_info["vehicle_no"], trk_info["transporter_name"], trk_info["driver_name"],
                                    trk_info["driver_phone"], bay_id, trk_info["allocated_bags"],
                                    round(trk_info["allocated_bags"] * 0.05, 2), round(util_pct, 2),
                                    now_dt.strftime("%Y-%m-%d %H:%M:%S")
                                ))

                                d_seq = 1
                                for itm_obj in trk_info["items"]:
                                    cur_opt.execute("""
                                        INSERT INTO trip_order_items (trip_id, order_no, agency_no, route_no, dr_code, fg_code, allocated_bags, allocated_weight_mt, delivery_seq, status)
                                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Assigned')
                                    """, (trip_code, itm_obj["order_no"], itm_obj["agency_no"], itm_obj["route_no"], itm_obj["dr_code"], itm_obj["fg_code"], itm_obj["bags_qty"], round(float(itm_obj["bags_qty"])*0.05, 2), d_seq))
                                    
                                    cur_opt.execute("UPDATE pending_orders SET status='Assigned' WHERE id=?", (itm_obj["id"],))
                                    d_seq += 1

                                cur_opt.execute("UPDATE fleet_master SET status='Assigned to Trip' WHERE vehicle_no=?", (trk_info["vehicle_no"],))
                                conn_opt.commit()
                                st.success(f"🎉 {trip_code} successfully created!")
                                del st.session_state["optimized_plan"]
                                st.rerun()

                if plan["unallocated"]:
                    st.markdown("---")
                    st.warning(f"⚠️ **Remaining Unpacked Load:** {len(plan['unallocated'])} Agencies could not fit in the selected trucks.")

    conn_opt.close()


# ------------------------------------------------------------------------------
# 2. DYNAMIC AUTO-DISCOVERY DATE & MULTI-FIELD FILTER CENTER
# ------------------------------------------------------------------------------
if main_menu == "🎯 Universal Date & Multi-Field Filter Center":
    st.title("🎯 Dynamic Auto-Discovery Date & Multi-Field Filter Center")
    st.markdown("Sabhi modules aur database tables ko dynamically auto-detect karke **Date Range (From Date ➔ To Date)** aur **Column Dropdowns** ke sath filter karein.")

    conn_flt_all = get_db_connection()
    cur_flt_all = conn_flt_all.cursor()

    cur_flt_all.execute("""
        SELECT name FROM sqlite_master 
        WHERE type='table' 
          AND name NOT LIKE 'sqlite_%' 
        ORDER BY name ASC
    """)
    all_active_tables = [r[0] for r in cur_flt_all.fetchall()]

    if not all_active_tables:
        st.warning("⚠️ Database me koi table nahi mili.")
    else:
        sel_table_name = st.selectbox("1. Select Module / Database Table to Filter:", all_active_tables, key="unified_tbl_selector")

        cur_flt_all.execute(f"PRAGMA table_info({sel_table_name})")
        tbl_columns = [col[1] for col in cur_flt_all.fetchall()]

        df_target_raw = pd.read_sql(f"SELECT * FROM {sel_table_name}", conn_flt_all)

        if df_target_raw.empty:
            st.info(f"ℹ️ Table `{sel_table_name}` me abhi koi data nahi hai.")
        else:
            df_filtered_view = df_target_raw.copy()

            # Date Column Auto-Detection
            date_col_options = [
                c for c in tbl_columns 
                if any(k in c.lower() for k in ["date", "time", "created", "logged", "upload", "at", "dispatch"])
            ]

            st.markdown("---")
            st.subheader("📅 Date Range Filtering")

            c_dt1, c_dt2, c_dt3 = st.columns([1, 1, 1])

            if date_col_options:
                with c_dt1:
                    chosen_date_col = st.selectbox("Detected Date Column:", date_col_options, key=f"u_dt_col_{sel_table_name}")

                df_filtered_view["_temp_eval_date"] = pd.to_datetime(df_filtered_view[chosen_date_col].astype(str).str[:10], errors="coerce")
                valid_date_series = df_filtered_view["_temp_eval_date"].dropna()

                if not valid_date_series.empty:
                    min_dt_val = valid_date_series.min().date()
                    max_dt_val = valid_date_series.max().date()

                    with c_dt2:
                        from_dt = st.date_input("From Date (IST):", min_dt_val, key=f"u_from_dt_{sel_table_name}")
                    with c_dt3:
                        to_dt = st.date_input("To Date (IST):", max_dt_val, key=f"u_to_dt_{sel_table_name}")

                    if from_dt and to_dt:
                        if from_dt <= to_dt:
                            mask_date_range = (df_filtered_view["_temp_eval_date"].dt.date >= from_dt) & (df_filtered_view["_temp_eval_date"].dt.date <= to_dt)
                            df_filtered_view = df_filtered_view[mask_date_range]
                        else:
                            st.error("⚠️ 'From Date' must be before or equal to 'To Date'.")
                else:
                    st.info("ℹ️ Is column me valid date format nahi mila.")

                df_filtered_view.drop(columns=["_temp_eval_date"], errors="ignore", inplace=True)
            else:
                st.info("ℹ️ Is table me koi date column detect nahi hua.")

            # Dynamic Column Multi-Select Dropdowns
            st.markdown("##### 🔍 Multi-Column Dynamic Dropdowns:")
            usable_filter_cols = [
                c for c in tbl_columns 
                if c not in ["id", "file_blob", "input_file_blob", "file_data", "file_hash"]
            ]

            if usable_filter_cols:
                display_flt_cols = usable_filter_cols[:6]
                flt_grid = st.columns(min(len(display_flt_cols), 3))

                for f_idx, f_col in enumerate(display_flt_cols):
                    with flt_grid[f_idx % 3]:
                        distinct_vals = sorted([str(x) for x in df_target_raw[f_col].dropna().unique() if str(x).strip() != ""])
                        if len(distinct_vals) > 0 and len(distinct_vals) <= 100:
                            chosen_flts = st.multiselect(
                                f"Filter {f_col.replace('_', ' ').title()}:",
                                distinct_vals,
                                key=f"u_dyn_flt_{sel_table_name}_{f_col}"
                            )
                            if chosen_flts:
                                df_filtered_view = df_filtered_view[df_filtered_view[f_col].astype(str).isin(chosen_flts)]

            # Global Text Search
            global_search_q = st.text_input("🔎 Search text across all fields:", "", key=f"u_dyn_search_{sel_table_name}")
            if global_search_q:
                df_filtered_view = df_filtered_view[df_filtered_view.apply(
                    lambda r: r.astype(str).str.contains(global_search_q, case=False).any(), axis=1
                )]

            # Dynamic Metrics Row
            st.markdown("---")
            k1, k2, k3, k4 = st.columns(4)
            k1.metric("Total Records", f"{len(df_filtered_view):,} of {len(df_target_raw):,}")

            numeric_cols = df_filtered_view.select_dtypes(include=['float', 'int']).columns.tolist()
            numeric_cols = [c for c in numeric_cols if c not in ['id', 'delivery_seq', 'version_no', 'pack_size_kg']]

            if len(numeric_cols) >= 1:
                col_n1 = numeric_cols[0]
                k2.metric(f"Sum {col_n1.title()}", f"{df_filtered_view[col_n1].sum():,.2f}")
            else:
                k2.metric("Columns", len(tbl_columns))

            if len(numeric_cols) >= 2:
                col_n2 = numeric_cols[1]
                k3.metric(f"Sum {col_n2.title()}", f"{df_filtered_view[col_n2].sum():,.2f}")
            else:
                match_p = (len(df_filtered_view) / len(df_target_raw) * 100) if len(df_target_raw) > 0 else 0
                k3.metric("Match Rate", f"{match_p:.1f}%")

            k4.metric("Status", "🟢 Live Ready")

            # Table Result & Exports
            st.markdown(f"##### 📋 Filtered Records in `{sel_table_name}`:")
            st.dataframe(df_filtered_view, use_container_width=True)

            c_e1, c_e2 = st.columns(2)
            with c_e1:
                st.download_button(
                    "📥 Export Filtered Data to Excel (.xlsx)",
                    to_excel_download_bytes(df_filtered_view, sel_table_name),
                    f"{sel_table_name}_Filtered_{get_ist_date_str()}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"u_btn_xl_{sel_table_name}"
                )
            with c_e2:
                st.download_button(
                    "📄 Export Filtered Data to CSV (.csv)",
                    df_filtered_view.to_csv(index=False).encode('utf-8'),
                    f"{sel_table_name}_Filtered_{get_ist_date_str()}.csv",
                    "text/csv",
                    key=f"u_btn_csv_{sel_table_name}"
                )

    conn_flt_all.close()


# ------------------------------------------------------------------------------
# 3. LIVE INVENTORY STOCK & ERP DEMAND RECONCILIATION ENGINE
# ------------------------------------------------------------------------------
if main_menu == "📦 Live Inventory Stock & ERP Demand Matcher":
    st.title("📦 Live Plant Stock, SKU Packaging & Demand Reconciliation Engine")
    st.markdown("Upload any **ERP / SAP MB52 Stock Sheet**, auto-calculate **Stock vs Pending Demand**, track **50 KG vs 25 KG Packaging**, and reconcile live warehouse buffers.")

    conn_stk_unified = get_db_connection()
    cur_stk_unified = conn_stk_unified.cursor()

    cur_stk_unified.execute("""
        CREATE TABLE IF NOT EXISTS plant_inventory_stock (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            upload_batch_id TEXT,
            source_file TEXT,
            file_hash TEXT,
            material_code TEXT,
            material_desc TEXT,
            pack_size_kg INTEGER DEFAULT 50,
            unit_weight_mt REAL DEFAULT 0.05,
            plant_stock_qty REAL DEFAULT 0.0,
            safety_stock_qty REAL DEFAULT 100.0,
            allocated_qty REAL DEFAULT 0.0,
            stock_date TEXT,
            created_at TEXT,
            UNIQUE(material_code, stock_date)
        )
    """)
    conn_stk_unified.commit()

    tab_stk1, tab_stk2, tab_stk3 = st.tabs([
        "📥 Upload & Auto-Detect Stock File",
        "⚖️ Live Stock vs Order Demand Balance",
        "✏️ Dynamic Stock Grid, Formulas & Schema Manager"
    ])

    # TAB 1: Stock Upload
    with tab_stk1:
        st.subheader("📥 Upload Plant Stock Inventory (Any ERP / SAP Layout)")
        up_stock_file = st.file_uploader("Upload Stock Excel / CSV File:", type=["xlsx", "xls", "csv"], key="u_stock_uploader_input")

        if up_stock_file:
            import hashlib
            file_bytes = up_stock_file.getvalue()
            file_hash = hashlib.md5(file_bytes).hexdigest()

            cur_stk_unified.execute("SELECT source_file, created_at FROM plant_inventory_stock WHERE file_hash=?", (file_hash,))
            dup_record = cur_stk_unified.fetchone()
            if dup_record:
                st.warning(f"⚠️ Duplicate File: Yeh file pehle upload ho chuki hai (`{dup_record[0]}`).")

            try:
                if up_stock_file.name.endswith(".csv"):
                    df_raw_stock = pd.read_csv(io.BytesIO(file_bytes))
                else:
                    df_raw_stock = pd.read_excel(io.BytesIO(file_bytes))

                st.dataframe(df_raw_stock.head(5), use_container_width=True)

                col_names = [str(c).strip() for c in df_raw_stock.columns]
                mat_guess = next((c for c in col_names if any(k in c.upper() for k in ["MAT", "FG", "SKU", "ITEM", "CODE", "PRODUCT"])), col_names[0])
                qty_guess = next((c for c in col_names if any(k in c.upper() for k in ["QTY", "STOCK", "UNRESTRICTED", "BAGS", "BALANCE", "AVAIL"])), col_names[-1])
                desc_guess = next((c for c in col_names if any(k in c.upper() for k in ["DESC", "NAME", "SPEC", "TEXT"])), "None")

                c_m1, c_m2, c_m3 = st.columns(3)
                with c_m1: sel_mat_col = st.selectbox("Select Material Column:", col_names, index=col_names.index(mat_guess), key="u_mat_col_pick")
                with c_m2: sel_qty_col = st.selectbox("Select Stock Qty Column:", col_names, index=col_names.index(qty_guess), key="u_qty_col_pick")
                with c_m3: sel_desc_col = st.selectbox("Select Description (Optional):", ["None"] + col_names, index=(["None"] + col_names).index(desc_guess), key="u_desc_col_pick")

                if st.button("🚀 Ingest & Process Stock into Live Inventory", type="primary", key="u_btn_stock_ingest"):
                    now_ts = get_ist_timestamp_full()
                    today_dt = get_ist_date_str()
                    batch_id = f"STK-{get_ist_now().strftime('%Y%m%d%H%M%S')}"

                    stock_entries = []
                    for _, s_row in df_raw_stock.iterrows():
                        raw_mat = str(s_row[sel_mat_col]).strip()
                        if pd.isna(raw_mat) or raw_mat in ["", "nan", "None", "0", "Total"]:
                            continue

                        raw_qty = s_row[sel_qty_col]
                        try:
                            clean_qty = float(re.sub(r'[^\d.]', '', str(raw_qty)))
                        except Exception:
                            clean_qty = 0.0

                        mat_desc = str(s_row[sel_desc_col]).strip() if sel_desc_col != "None" else raw_mat

                        # Packaging Auto-Detection (25 KG vs 50 KG)
                        combined_text = (raw_mat + " " + mat_desc).upper()
                        if any(k in combined_text for k in ["25KG", "25 KG", "25-KG"]):
                            pack_kg, unit_mt = 25, 0.025
                        else:
                            pack_kg, unit_mt = 50, 0.050

                        stock_entries.append((
                            batch_id, up_stock_file.name, file_hash, raw_mat, mat_desc,
                            pack_kg, unit_mt, clean_qty, 100.0, 0.0, today_dt, now_ts
                        ))

                    if stock_entries:
                        cur_stk_unified.executemany("""
                            INSERT OR REPLACE INTO plant_inventory_stock (
                                upload_batch_id, source_file, file_hash, material_code, material_desc,
                                pack_size_kg, unit_weight_mt, plant_stock_qty, safety_stock_qty, allocated_qty, stock_date, created_at
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, stock_entries)
                        conn_stk_unified.commit()
                        st.success(f"✅ Ingested {len(stock_entries)} records with 25/50 KG detection!")
                        st.rerun()
            except Exception as ex:
                st.error(f"Upload error: {ex}")

    # TAB 2: Reconciliation
    with tab_stk2:
        st.subheader("⚖️ Stock vs Pending Demand Reconciliation")

        df_inv = pd.read_sql("SELECT * FROM plant_inventory_stock ORDER BY id DESC", conn_stk_unified)
        df_ord = pd.read_sql("SELECT fg_code, SUM(bags_qty) as total_demand_bags FROM pending_orders WHERE status='Pending' GROUP BY fg_code", conn_stk_unified)

        if df_inv.empty:
            st.info("ℹ️ Koi stock upload nahi hua hai.")
        else:
            df_reconcile = pd.merge(df_inv, df_ord, left_on="material_code", right_on="fg_code", how="left")
            df_reconcile["total_demand_bags"] = df_reconcile["total_demand_bags"].fillna(0.0)
            df_reconcile["net_available_bags"] = df_reconcile["plant_stock_qty"] - df_reconcile["total_demand_bags"]
            df_reconcile["total_stock_mt"] = df_reconcile["plant_stock_qty"] * df_reconcile["unit_weight_mt"]
            df_reconcile["demand_mt"] = df_reconcile["total_demand_bags"] * df_reconcile["unit_weight_mt"]
            df_reconcile["net_mt"] = df_reconcile["net_available_bags"] * df_reconcile["unit_weight_mt"]

            def get_health(r):
                if r["net_available_bags"] < 0: return "🚨 SHORTAGE"
                elif r["net_available_bags"] < r["safety_stock_qty"]: return "⚠️ LOW SAFETY"
                return "🟢 HEALTHY"

            df_reconcile["Status"] = df_reconcile.apply(get_health, axis=1)

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Plant Stock", f"{df_reconcile['plant_stock_qty'].sum():,.0f} Bags", f"{df_reconcile['total_stock_mt'].sum():,.2f} MT")
            m2.metric("Pending Demand", f"{df_reconcile['total_demand_bags'].sum():,.0f} Bags", f"{df_reconcile['demand_mt'].sum():,.2f} MT")
            m3.metric("Free Stock", f"{df_reconcile['net_available_bags'].sum():,.0f} Bags", f"{df_reconcile['net_mt'].sum():,.2f} MT")
            m4.metric("Shortage SKUs", int((df_reconcile["net_available_bags"] < 0).sum()))

            display_cols = [
                "material_code", "material_desc", "pack_size_kg", "unit_weight_mt",
                "plant_stock_qty", "total_demand_bags", "net_available_bags", "safety_stock_qty",
                "total_stock_mt", "demand_mt", "Status", "stock_date"
            ]
            st.dataframe(df_reconcile[display_cols], use_container_width=True)

            st.download_button(
                "📥 Export Reconciliation (.xlsx)",
                to_excel_download_bytes(df_reconcile[display_cols], "StockVsOrder"),
                f"Stock_Reconciliation_{get_ist_date_str()}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="u_btn_dl_recon_xl"
            )

    # TAB 3: Dynamic Grid
    with tab_stk3:
        st.subheader("✏️ Dynamic Stock Grid & Column/Formula Engine")
        df_grid = pd.read_sql("SELECT * FROM plant_inventory_stock", conn_stk_unified)

        with st.expander("➕ Add Dynamic Column"):
            c1, c2, c3 = st.columns(3)
            with c1: new_c = re.sub(r'[^a-z0-9_]', '', st.text_input("New Column Name:", key="u_col_name_input").strip().lower())
            with c2: new_t = st.selectbox("Type:", ["TEXT", "REAL", "INTEGER"], key="u_col_type_select")
            with c3:
                if st.button("Add Column", key="u_btn_add_col") and new_c:
                    cur_stk_unified.execute(f"ALTER TABLE plant_inventory_stock ADD COLUMN {new_c} {new_t}")
                    conn_stk_unified.commit()
                    st.rerun()

        edited_stock = st.data_editor(df_grid, use_container_width=True, num_rows="dynamic", key="u_stock_grid_editor")

        if st.button("💾 Commit Grid Edits", type="primary", key="u_btn_save_stock_grid"):
            for _, r in edited_stock.iterrows():
                r_dict = r.to_dict()
                r_id = r_dict.get("id")
                r_dict["unit_weight_mt"] = 0.025 if int(r_dict.get("pack_size_kg", 50)) == 25 else 0.050
                if pd.notna(r_id):
                    up_cols = [k for k in r_dict.keys() if k != "id"]
                    set_str = ", ".join([f"{c}=?" for c in up_cols])
                    cur_stk_unified.execute(f"UPDATE plant_inventory_stock SET {set_str} WHERE id=?", [r_dict[c] for c in up_cols] + [r_id])
                else:
                    ins_cols = [k for k in r_dict.keys() if k != "id" and pd.notna(r_dict[k])]
                    cur_stk_unified.execute(f"INSERT INTO plant_inventory_stock ({', '.join(ins_cols)}) VALUES ({', '.join(['?' for _ in ins_cols])})", [r_dict[c] for c in ins_cols])
            conn_stk_unified.commit()
            st.success("Changes saved!")
            st.rerun()

    conn_stk_unified.close()
    # ==============================================================================
# SECTION: NEXT-GEN ENTERPRISE UNIVERSAL DATA HUB & DYNAMIC FILTER ENGINE
# (FULLY LINKED TO ALL MODULES + RUNTIME CUSTOM FILTER BUILDER + ZERO DUPLICATES)
# ==============================================================================

def render_advanced_universal_data_hub(is_full_page=True, key_scope="univ_hub"):
    conn_hub = get_db_connection()
    cur_hub = conn_hub.cursor()

    # 1. Fetch All Tables Dynamically (Existing + Future Custom DB Tables)
    cur_hub.execute("""
        SELECT name FROM sqlite_master 
        WHERE type='table' 
          AND name NOT LIKE 'sqlite_%' 
        ORDER BY name ASC
    """)
    available_tables = [r[0] for r in cur_hub.fetchall()]

    if not available_tables:
        st.warning("⚠️ Database me abhi koi tables uplabdh nahi hain.")
        conn_hub.close()
        return

    if is_full_page:
        st.title("🎯 Enterprise Universal Data Hub & Dynamic Filter Center")
        st.markdown("Sabhi modules, plant stock, sales orders aur dispatch registers ko **Date Range**, **Custom Filter Builder**, aur **Live Search** ke sath inspect aur export karein.")

    # Friendly Display Names Mapping
    TABLE_DISPLAY_NAMES = {
        "daily_dispatch_register": "📖 Daily Dispatch Sale Register",
        "plant_inventory_stock": "📦 Plant Stock & Inventory Ledger (50KG/25KG)",
        "pending_orders": "⏳ Pending Orders Database",
        "partial_dispatch_ledger": "🧩 Partial / Split Dispatch Database",
        "trip_loading_slips": "📋 Trip Loading Slips & Gate Passes",
        "trip_order_items": "📑 Trip Order Item Manifests",
        "unique_routes_master": "🗺️ Route-Agency-DR Master DB",
        "fleet_master": "🚛 Transporter Fleet Master",
        "loading_bays": "🏭 Plant Loading Bays",
        "input_output_traceability": "🔍 Input-Output Traceability Ledger",
        "uploaded_files_archive": "🗄️ File Upload Archive History",
        "unmapped_missing_dr_ledger": "🚨 Unmapped Missing DR Ledger"
    }

    def format_table_label(t):
        return TABLE_DISPLAY_NAMES.get(t, f"🗄️ Custom Table: {t}")

    # Table Selector
    sel_tbl = st.selectbox(
        "1. Select Database Table / Module to Inspect:",
        available_tables,
        format_func=format_table_label,
        key=f"{key_scope}_tbl_selector"
    )

    # Schema Analysis
    cur_hub.execute(f"PRAGMA table_info({sel_tbl})")
    col_records = cur_hub.fetchall()
    all_columns = [c[1] for c in col_records]
    
    # Load Table Data
    df_hub_raw = pd.read_sql(f"SELECT * FROM {sel_tbl}", conn_hub)

    if df_hub_raw.empty:
        st.info(f"ℹ️ Table `{sel_tbl}` create ho chuki hai par isme abhi koi records nahi hain.")
        st.dataframe(pd.DataFrame(columns=all_columns), use_container_width=True)
        conn_hub.close()
        return

    df_hub_active = df_hub_raw.copy()

    # --------------------------------------------------------------------------
    # 2. SMART DATE RANGE ENGINE
    # --------------------------------------------------------------------------
    date_candidates = [
        c for c in all_columns 
        if any(k in c.lower() for k in ["date", "time", "created", "logged", "upload", "at", "dispatch"])
    ]

    st.markdown("---")
    st.subheader("📅 1. Date Range Filtering")

    if date_candidates:
        c_dt1, c_dt2, c_dt3 = st.columns([1, 1, 1])
        with c_dt1:
            active_date_col = st.selectbox("Active Date Column:", date_candidates, key=f"{key_scope}_dt_col_{sel_tbl}")

        # Standardize and clean string dates for evaluation
        df_hub_active["_temp_eval_dt"] = pd.to_datetime(df_hub_active[active_date_col].astype(str).str[:10], errors="coerce")
        valid_dates = df_hub_active["_temp_eval_dt"].dropna()

        if not valid_dates.empty:
            with c_dt2:
                from_dt_pick = st.date_input("From Date (IST):", valid_dates.min().date(), key=f"{key_scope}_from_{sel_tbl}")
            with c_dt3:
                to_dt_pick = st.date_input("To Date (IST):", valid_dates.max().date(), key=f"{key_scope}_to_{sel_tbl}")

            if from_dt_pick and to_dt_pick:
                if from_dt_pick <= to_dt_pick:
                    date_mask = (df_hub_active["_temp_eval_dt"].dt.date >= from_dt_pick) & (df_hub_active["_temp_eval_dt"].dt.date <= to_dt_pick)
                    df_hub_active = df_hub_active[date_mask]
                else:
                    st.error("⚠️ 'From Date' must be before or equal to 'To Date'.")

        df_hub_active.drop(columns=["_temp_eval_dt"], errors="ignore", inplace=True)
    else:
        st.info("ℹ️ Is table me date field nahi hai (Static Master).")

    # --------------------------------------------------------------------------
    # 3. DYNAMIC MULTI-COLUMN DROPDOWN FILTERS & RUNTIME FILTER BUILDER
    # --------------------------------------------------------------------------
    st.markdown("---")
    st.subheader("🔍 2. Dynamic Column Filters & Custom Filter Builder")

    # Default filterable columns (excluding system blobs/hashes)
    eligible_cols = [
        c for c in all_columns 
        if c not in ["id", "file_blob", "input_file_blob", "file_data", "file_hash", "upload_batch_id"]
    ]

    # Runtime Custom Filter Selector: User can add/remove filter fields on the fly
    default_active_filters = eligible_cols[:min(4, len(eligible_cols))]
    user_chosen_filter_cols = st.multiselect(
        "➕ Choose columns to generate Filter Dropdowns:",
        eligible_cols,
        default=default_active_filters,
        key=f"{key_scope}_chosen_flts_{sel_tbl}"
    )

    if user_chosen_filter_cols:
        grid_width = min(len(user_chosen_filter_cols), 3)
        flt_grid_cols = st.columns(grid_width)

        for idx_f, col_f in enumerate(user_chosen_filter_cols):
            with flt_grid_cols[idx_f % grid_width]:
                distinct_options = sorted([str(x) for x in df_hub_raw[col_f].dropna().unique() if str(x).strip() != ""])
                
                # Check if column is numeric or categorical
                if len(distinct_options) > 0 and len(distinct_options) <= 200:
                    picked_vals = st.multiselect(
                        f"Filter by {col_f.replace('_', ' ').title()}:",
                        distinct_options,
                        key=f"{key_scope}_flt_{sel_tbl}_{col_f}"
                    )
                    if picked_vals:
                        df_hub_active = df_hub_active[df_hub_active[col_f].astype(str).isin(picked_vals)]
                else:
                    manual_match = st.text_input(f"Match {col_f.replace('_', ' ').title()}:", "", key=f"{key_scope}_manual_{sel_tbl}_{col_f}")
                    if manual_match:
                        df_hub_active = df_hub_active[df_hub_active[col_f].astype(str).str.contains(manual_match, case=False)]

    # --------------------------------------------------------------------------
    # 4. ADVANCED NUMERIC RANGE & OPERATOR FILTER
    # --------------------------------------------------------------------------
    numeric_columns = df_hub_raw.select_dtypes(include=['float', 'int']).columns.tolist()
    numeric_columns = [c for c in numeric_columns if c not in ['id', 'delivery_seq', 'version_no']]

    if numeric_columns:
        with st.expander("🧮 Numeric Range & Operator Filter (> , < , =)", expanded=False):
            n_c1, n_c2, n_c3 = st.columns(3)
            with n_c1:
                target_num_col = st.selectbox("Select Numeric Column:", numeric_columns, key=f"{key_scope}_num_col_{sel_tbl}")
            with n_c2:
                num_operator = st.selectbox("Operator:", ["Greater than or equal (>=)", "Less than or equal (<=)", "Exactly equal (==)"], key=f"{key_scope}_op_{sel_tbl}")
            with n_c3:
                threshold_val = st.number_input("Threshold Value:", value=0.0, step=10.0, key=f"{key_scope}_thresh_{sel_tbl}")

            if threshold_val > 0:
                if "Greater" in num_operator:
                    df_hub_active = df_hub_active[df_hub_active[target_num_col] >= threshold_val]
                elif "Less" in num_operator:
                    df_hub_active = df_hub_active[df_hub_active[target_num_col] <= threshold_val]
                else:
                    df_hub_active = df_hub_active[df_hub_active[target_num_col] == threshold_val]

    # --------------------------------------------------------------------------
    # 5. GLOBAL KEYWORD SEARCH
    # --------------------------------------------------------------------------
    st.markdown("---")
    keyword_query = st.text_input("🔎 Search text / keyword across all columns:", "", key=f"{key_scope}_kw_search_{sel_tbl}")
    if keyword_query:
        df_hub_active = df_hub_active[df_hub_active.apply(
            lambda row: row.astype(str).str.contains(keyword_query, case=False).any(), axis=1
        )]

    # --------------------------------------------------------------------------
    # 6. DYNAMIC SUMMARY METRICS BAR
    # --------------------------------------------------------------------------
    st.markdown("---")
    st.subheader("📊 3. Live Aggregation Metrics")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Filtered Records", f"{len(df_hub_active):,} of {len(df_hub_raw):,}")

    if "dispatched_bags" in df_hub_active.columns:
        m2.metric("Dispatched Bags", f"{df_hub_active['dispatched_bags'].sum():,.0f}")
        m3.metric("Dispatched Weight (MT)", f"{df_hub_active['dispatched_weight_mt'].sum():,.2f} MT")
    elif "plant_stock_qty" in df_hub_active.columns:
        tot_bags = df_hub_active['plant_stock_qty'].sum()
        tot_mt = (df_hub_active['plant_stock_qty'] * df_hub_active['unit_weight_mt']).sum() if 'unit_weight_mt' in df_hub_active.columns else (tot_bags * 0.05)
        m2.metric("Plant Stock Bags", f"{tot_bags:,.0f}")
        m3.metric("Stock Tonnage (MT)", f"{tot_mt:,.2f} MT")
    elif "bags_qty" in df_hub_active.columns:
        m2.metric("Demand Bags", f"{df_hub_active['bags_qty'].sum():,.0f}")
        m3.metric("Demand Weight (MT)", f"{df_hub_active['weight_mt'].sum():,.2f} MT")
    elif "remaining_bags" in df_hub_active.columns:
        m2.metric("Remaining Bags", f"{df_hub_active['remaining_bags'].sum():,.0f}")
        m3.metric("Dispatched Bags", f"{df_hub_active['dispatched_bags'].sum():,.0f}")
    else:
        match_rate = (len(df_hub_active) / len(df_hub_raw) * 100) if len(df_hub_raw) > 0 else 0
        m2.metric("Match Rate", f"{match_rate:.1f}%")
        m3.metric("Columns Present", len(all_columns))

    if "route_no" in df_hub_active.columns:
        m4.metric("Active Routes", f"{df_hub_active['route_no'].nunique():,}")
    elif "vehicle_no" in df_hub_active.columns:
        m4.metric("Active Vehicles", f"{df_hub_active['vehicle_no'].nunique():,}")
    else:
        m4.metric("Status", "🟢 Real-time Synced")

    # --------------------------------------------------------------------------
    # 7. LIVE DATA VIEW & EXPORT ACTIONS
    # --------------------------------------------------------------------------
    st.markdown(f"##### 📋 Live Filtered Result Table (`{sel_tbl}`):")
    st.dataframe(df_hub_active, use_container_width=True)

    exp_col1, exp_col2 = st.columns(2)
    with exp_col1:
        st.download_button(
            "📥 Export Filtered Excel (.xlsx)",
            to_excel_download_bytes(df_hub_active, sel_tbl),
            f"{sel_tbl}_Filtered_{get_ist_date_str()}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{key_scope}_btn_excel_{sel_tbl}"
        )
    with exp_col2:
        st.download_button(
            "📄 Export Filtered CSV (.csv)",
            df_hub_active.to_csv(index=False).encode('utf-8'),
            f"{sel_tbl}_Filtered_{get_ist_date_str()}.csv",
            "text/csv",
            key=f"{key_scope}_btn_csv_{sel_tbl}"
        )

    conn_hub.close()


# ------------------------------------------------------------------------------
# TRIGGER 1: DEDICATED FULL-PAGE VIEW (WHEN SELECTED IN SIDEBAR NAVIGATION)
# ------------------------------------------------------------------------------
if any(term in str(main_menu).lower() for term in ["universal date", "filter center", "filter engine", "multi-field"]):
    render_advanced_universal_data_hub(is_full_page=True, key_scope="fullpage_hub")

# ------------------------------------------------------------------------------
# TRIGGER 2: EMBEDDED GLOBAL FILTER PANEL (AVAILABLE AT BOTTOM OF ALL MODULES)
# ------------------------------------------------------------------------------
st.markdown("---")
with st.expander("🔍 Global Database Filter Hub (Auto-Linked to All Tables)", expanded=False):
    render_advanced_universal_data_hub(is_full_page=False, key_scope="bottom_hub")
    # ==============================================================================
# SECTION 20: ULTIMATE SUPREME ENTERPRISE RECOVERY & FLEXIBLE BACKUP SUITE
# ==============================================================================

def init_ultimate_supreme_master_db():
    try:
        conn = sqlite3.connect("enterprise_logistics_sales_hub.db", timeout=10)
        cursor = conn.cursor()
        cursor.execute("PRAGMA journal_mode=WAL;")
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS dynamic_modules_ledger (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE,
                category TEXT,
                icon TEXT,
                code TEXT,
                is_trashed INTEGER DEFAULT 0,
                is_permanent_deleted INTEGER DEFAULT 0,
                created_at TEXT,
                deleted_at TEXT
            )
        """)
        cursor.execute("PRAGMA table_info(dynamic_modules_ledger)")
        cols = [c[1] for c in cursor.fetchall()]
        if "is_permanent_deleted" not in cols:
            cursor.execute("ALTER TABLE dynamic_modules_ledger ADD COLUMN is_permanent_deleted INTEGER DEFAULT 0")
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Supreme Master Init Error: {str(e)}")

init_ultimate_supreme_master_db()

# --- SAFE PERSISTENT FETCH PRIOR TO SIDEBAR (PAGE REFRESH PROOF) ---
try:
    conn = sqlite3.connect("enterprise_logistics_sales_hub.db", timeout=10)
    df_sup = pd.read_sql("SELECT * FROM dynamic_modules_ledger", conn)
    conn.close()
except:
    df_sup = pd.DataFrame()

active_sup_mods = df_sup[(df_sup['is_trashed'] == 0) & (df_sup['is_permanent_deleted'] == 0)].to_dict('records') if not df_sup.empty and 'is_trashed' in df_sup.columns else []
trash_sup_mods = df_sup[(df_sup['is_trashed'] == 1) & (df_sup['is_permanent_deleted'] == 0)].to_dict('records') if not df_sup.empty and 'is_trashed' in df_sup.columns else []

# --- PERSISTENT MODULE BUILDER & FLEXIBLE BACKUP / RESTORE HUB ---
st.markdown("---")
with st.expander("🔌 Add New Dynamic Module & Flexible Backup Hub (Supreme Suite)", expanded=False):
    tab_bld1, tab_bld2 = st.tabs(["➕ Add New Module", "📦 Flexible Backup & Restore Hub"])
    
    with tab_bld1:
        st.markdown("Yahan aap naya module create kar sakte hain. Duplicate name conflict aur delete recovery 100% automated hai.")
        with st.form("supreme_dynamic_form_persistent"):
            col_s1, col_s2 = st.columns(2)
            with col_s1:
                s_name = st.text_input("Module / Feature Name", placeholder="e.g., NextGen Enterprise Tonnage Pro")
                s_cat = st.selectbox("Module Category", ["Analytics", "Automation", "Reporting", "Integration", "Custom Utility"])
            with col_s2:
                s_icon = st.text_input("Module Icon (Emoji)", placeholder="⚡")
            
            s_code = st.text_area(
                "Module Python Logic (Streamlit Code)", 
                placeholder="st.subheader('My Custom Feature')\nst.write('Operational and fully linked!')",
                height=120
            )
            
            s_submit = st.form_submit_button("⚡ Create & Permanently Save Feature")
            
            if s_submit:
                if s_name and s_code:
                    try:
                        conn = sqlite3.connect("enterprise_logistics_sales_hub.db", timeout=10)
                        cursor = conn.cursor()
                        clean_name = s_name.strip()
                        now_ts = get_ist_now().strftime("%Y-%m-%d %H:%M:%S")
                        icon_val = s_icon if s_icon else "🧩"
                        
                        cursor.execute("""
                            INSERT INTO dynamic_modules_ledger (name, category, icon, code, is_trashed, is_permanent_deleted, created_at, deleted_at)
                            VALUES (?, ?, ?, ?, 0, 0, ?, NULL)
                            ON CONFLICT(name) DO UPDATE SET 
                                category = excluded.category,
                                icon = excluded.icon,
                                code = excluded.code,
                                is_trashed = 0,
                                is_permanent_deleted = 0,
                                deleted_at = NULL
                        """, (clean_name, s_cat, icon_val, s_code, now_ts))
                        
                        conn.commit()
                        conn.close()
                        st.success(f"✅ Module '{clean_name}' successfully saved and locked for persistence!")
                        st.rerun()
                    except Exception as ex:
                        st.error(f"❌ Error: {str(ex)}")
                else:
                    st.warning("⚠️ Kripya Module Name aur Code dono enter karein.")

    with tab_bld2:
        st.markdown("##### 💾 Flexible Backup & Restore Center")
        st.markdown("Aap chahe toh **Single Module** ka backup lein, ya **Saare Modules (All / Multiple)** ka ek sath backup download karein.")
        
        sub_bk_tab1, sub_bk_tab2 = st.tabs(["📤 Export / Backup", "📥 Import / Restore"])
        
        with sub_bk_tab1:
            st.markdown("###### 1. Single Module Backup (1 by 1)")
            if active_sup_mods:
                single_mod_names = [m['name'] for m in active_sup_mods]
                sel_single_exp = st.selectbox("Select Module to Export:", single_mod_names, key="sel_single_exp_mod")
                
                # Find selected module dict
                match_mod = next((m for m in active_sup_mods if m['name'] == sel_single_exp), None)
                if match_mod:
                    single_json_str = json.dumps([match_mod], indent=4)
                    st.download_button(
                        f"📥 Download '{sel_single_exp}' Backup (.json)",
                        data=single_json_str.encode('utf-8'),
                        file_name=f"Module_{sel_single_exp.replace(' ', '_')}_{get_ist_date_str()}.json",
                        mime="application/json",
                        key=f"dl_single_mod_{sel_single_exp}"
                    )
            else:
                st.info("ℹ️ Koi active module available nahi hai.")
                
            st.markdown("---")
            st.markdown("###### 2. All / Multiple Modules Backup (Bulk Export)")
            if not df_sup.empty:
                all_backup_data = df_sup[['name', 'category', 'icon', 'code', 'is_trashed', 'is_permanent_deleted', 'created_at']].to_dict('records')
                all_json_str = json.dumps(all_backup_data, indent=4)
                
                st.download_button(
                    "📥 Download All Modules Bulk Backup (.json)",
                    data=all_json_str.encode('utf-8'),
                    file_name=f"All_Dynamic_Modules_Bulk_Backup_{get_ist_date_str()}.json",
                    mime="application/json",
                    type="primary",
                    key="dl_all_bulk_mods"
                )
            else:
                st.info("ℹ️ Abhi database me koi module save nahi hai.")
                
        with sub_bk_tab2:
            st.markdown("###### 3. Restore Modules from Backup File")
            up_backup_file = st.file_uploader("Upload Single or Bulk Backup JSON File:", type=["json"], key="up_sup_flexible_json_backup")
            if up_backup_file and st.button("🚀 Restore Modules From Uploaded File", key="btn_flexible_restore_mods"):
                try:
                    file_content = json.loads(up_backup_file.getvalue().decode('utf-8'))
                    # Handle both single dict or list of dicts
                    mods_to_import = file_content if isinstance(file_content, list) else [file_content]
                    
                    if len(mods_to_import) > 0:
                        conn = sqlite3.connect("enterprise_logistics_sales_hub.db", timeout=10)
                        cursor = conn.cursor()
                        now_ts = get_ist_now().strftime("%Y-%m-%d %H:%M:%S")
                        
                        restored_count = 0
                        for m_item in mods_to_import:
                            m_name = m_item.get('name', '').strip()
                            m_code = m_item.get('code', '')
                            m_cat = m_item.get('category', 'Custom Utility')
                            m_icon = m_item.get('icon', '🧩')
                            
                            if m_name and m_code:
                                cursor.execute("""
                                    INSERT INTO dynamic_modules_ledger (name, category, icon, code, is_trashed, is_permanent_deleted, created_at, deleted_at)
                                    VALUES (?, ?, ?, ?, 0, 0, ?, NULL)
                                    ON CONFLICT(name) DO UPDATE SET 
                                        category = excluded.category,
                                        icon = excluded.icon,
                                        code = excluded.code,
                                        is_trashed = 0,
                                        is_permanent_deleted = 0,
                                        deleted_at = NULL
                                """, (m_name, m_cat, m_icon, m_code, now_ts))
                                restored_count += 1
                                
                        conn.commit()
                        conn.close()
                        st.success(f"🎉 Successfully restored and activated {restored_count} module(s) from backup!")
                        st.rerun()
                    else:
                        st.error("❌ Invalid backup file format.")
                except Exception as ex:
                    st.error(f"Restore Error: {str(ex)}")

# --- SIDEBAR ROUTER ---
if True:
    st.sidebar.markdown("---")
    st.sidebar.markdown("### ⚡ Dynamic Modules Hub")
    
    sup_options = ["🏠 Main Dashboard"]
    if active_sup_mods:
        sup_options += [f"{m['icon']} {m['name']}" for m in active_sup_mods]
    if trash_sup_mods:
        sup_options += ["🗑️ Recycle Bin (Trash)"]
    sup_options += ["🛠️ Supreme Control & Recovery Center"]
    
    sup_nav = st.sidebar.radio("Select Workspace", sup_options, key="supreme_sidebar_nav_persistent")
    
    if sup_nav == "🗑️ Recycle Bin (Trash)":
        st.markdown("---")
        st.markdown("# 🗑️ Deleted Modules & Recycle Bin")
        st.markdown("Yahan se aap deleted modules ko **Restore** kar sakte hain.")
        st.markdown("---")

        if trash_sup_mods:
            for tm in trash_sup_mods:
                col_i, col_r, col_p = st.columns([3, 1, 1])
                with col_i:
                    st.markdown(f"**{tm['icon']} {tm['name']}** (`{tm['category']}`)")
                    st.caption(f"Deleted at: {tm.get('deleted_at', 'N/A')}")
                with col_r:
                    if st.button("♻️ Restore", key=f"sup_res_p_{tm['id']}"):
                        conn = sqlite3.connect("enterprise_logistics_sales_hub.db", timeout=10)
                        cursor = conn.cursor()
                        cursor.execute("UPDATE dynamic_modules_ledger SET is_trashed = 0, deleted_at = NULL WHERE id = ?", (tm['id'],))
                        conn.commit()
                        conn.close()
                        st.success(f"Module '{tm['name']}' restored successfully!")
                        st.rerun()
                with col_p:
                    if st.button("🔥 Delete Forever", key=f"sup_per_p_{tm['id']}"):
                        conn = sqlite3.connect("enterprise_logistics_sales_hub.db", timeout=10)
                        cursor = conn.cursor()
                        cursor.execute("UPDATE dynamic_modules_ledger SET is_permanent_deleted = 1 WHERE id = ?", (tm['id'],))
                        conn.commit()
                        conn.close()
                        st.warning(f"Module '{tm['name']}' moved to archive!")
                        st.rerun()
                st.markdown("---")
        else:
            st.info("Recycle Bin bilkul khali hai.")

    elif sup_nav == "🛠️ Supreme Control & Recovery Center":
        st.markdown("---")
        st.markdown("# 🛠️ Supreme Enterprise Control & Recovery Suite")
        st.markdown("Yahan se aap auto-detect karke saari operational tables wipe kar sakte hain, Master Factory Reset chala sakte hain, aur Custom Colors manage kar sakte hain.")
        st.markdown("---")

        tab_sc1, tab_sc2 = st.tabs([
            "🧹 Auto-Detect Wiping & Factory Reset", 
            "🎨 Advanced Custom Colors & Button Text"
        ])

        with tab_sc1:
            st.markdown("#### 🔍 Safe Auto-Detected Database Tables & Error-Free Wiping")
            try:
                conn_s = sqlite3.connect("enterprise_logistics_sales_hub.db", timeout=10)
                cur_s = conn_s.cursor()
                cur_s.execute("SELECT name FROM sqlite_master WHERE type='table';")
                all_s_tbls = [row[0] for row in cur_s.fetchall()]
                conn_s.close()
            except:
                all_s_tbls = []
            
            safe_s_tbls = [t for t in all_s_tbls if t != "dynamic_modules_ledger"]

            if safe_s_tbls:
                st.markdown("##### 📋 Detected Operational Tables in Database:")
                for stbl in safe_s_tbls:
                    st.markdown(f"- 📁 `{stbl}`")
                
                st.markdown("---")
                col_sw1, col_sw2 = st.columns(2)
                with col_sw1:
                    st.markdown("##### Wipe Specific Operational Table")
                    sel_s_tbl = st.selectbox("Detected Tables", ["-- Select Table --"] + safe_s_tbls, key="sel_sup_tbl_p")
                    if st.button("🗑️ Wipe Selected Table Data", key="btn_wipe_sel_p"):
                        if sel_s_tbl != "-- Select Table --":
                            try:
                                conn = sqlite3.connect("enterprise_logistics_sales_hub.db", timeout=10)
                                cursor = conn.cursor()
                                cursor.execute("PRAGMA foreign_keys = OFF;")
                                cursor.execute(f"DELETE FROM {sel_s_tbl}")
                                try:
                                    cursor.execute(f"DELETE FROM sqlite_sequence WHERE name='{sel_s_tbl}'")
                                except:
                                    pass
                                cursor.execute("PRAGMA foreign_keys = ON;")
                                conn.commit()
                                conn.close()
                                st.success(f"✅ Table '{sel_s_tbl}' successfully cleaned!")
                                st.rerun()
                            except Exception as ex:
                                st.error(f"Error: {str(ex)}")
                        else:
                            st.warning("⚠️ Kripya table select karein.")

                with col_sw2:
                    st.markdown("##### Master Factory Reset")
                    if st.button("⚡ EXECUTE SUPREME FACTORY RESET", type="primary", key="btn_sup_reset_p"):
                        try:
                            conn = sqlite3.connect("enterprise_logistics_sales_hub.db", timeout=10)
                            cursor = conn.cursor()
                            cursor.execute("PRAGMA foreign_keys = OFF;")
                            for t in safe_s_tbls:
                                try:
                                    cursor.execute(f"DELETE FROM {t}")
                                    cursor.execute(f"DELETE FROM sqlite_sequence WHERE name='{t}'")
                                except:
                                    pass
                            cursor.execute("PRAGMA foreign_keys = ON;")
                            cursor.execute("UPDATE dynamic_modules_ledger SET is_trashed = 0, is_permanent_deleted = 0, deleted_at = NULL")
                            conn.commit()
                            conn.close()
                            for c_key in ["sup_bg", "sup_txt", "sup_btn_bg", "sup_btn_txt"]:
                                if c_key in st.session_state:
                                    del st.session_state[c_key]
                            st.success("🚀 Supreme Factory Reset Completed!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Reset Error: {str(e)}")
            else:
                st.info("ℹ️ Koi bhi operational table abhi database me maujood nahi hai. Nayi table create karne ke baad yahan data wipe ya reset options dikhenge.")

        with tab_sc2:
            st.markdown("#### 🎨 Custom Colors & Button Text Overrider")
            sc1, sc2 = st.columns(2)
            with sc1:
                sup_bg = st.color_picker("Background Color", st.session_state.get("sup_bg", "#f4f6f9"), key="pick_sup_bg_p")
                sup_txt = st.color_picker("Main Text Color", st.session_state.get("sup_txt", "#1f2937"), key="pick_sup_txt_p")
            with sc2:
                sup_btn_bg = st.color_picker("Button Background Color", st.session_state.get("sup_btn_bg", "#1e3a8a"), key="pick_sup_btn_bg_p")
                sup_btn_txt = st.color_picker("Button Text Color", st.session_state.get("sup_btn_txt", "#ffffff"), key="pick_sup_btn_txt_p")

            sa1, sa2 = st.columns(2)
            with sa1:
                if st.button("✨ Apply Custom Colors", type="primary", key="apply_sup_colors_p"):
                    st.session_state["sup_bg"] = sup_bg
                    st.session_state["sup_txt"] = sup_txt
                    st.session_state["sup_btn_bg"] = sup_btn_bg
                    st.session_state["sup_btn_txt"] = sup_btn_txt
                    st.success("✅ Custom colors applied successfully!")
                    st.rerun()
            with sa2:
                if st.button("🔄 Reset Custom Colors", key="reset_sup_colors_p"):
                    for c_key in ["sup_bg", "sup_txt", "sup_btn_bg", "sup_btn_txt"]:
                        if c_key in st.session_state:
                            del st.session_state[c_key]
                    st.success("✅ Colors reset to default!")
                    st.rerun()

    elif sup_nav != "🏠 Main Dashboard":
        active_sup_names = [f"{m['icon']} {m['name']}" for m in active_sup_mods]
        if sup_nav in active_sup_names:
            idx_sm = active_sup_names.index(sup_nav)
            cur_sup_mod = active_sup_mods[idx_sm]
            
            st.empty()
            st.markdown(f"# {cur_sup_mod['icon']} {cur_sup_mod['name']}")
            st.markdown(f"**Category:** `{cur_sup_mod['category']}` | **Created At:** `{cur_sup_mod['created_at']}`")
            st.markdown("---")
            
            try:
                local_scope_sup = {"st": st, "pd": pd, "io": io, "sqlite3": sqlite3, "datetime": datetime}
                exec(cur_sup_mod['code'], globals(), local_scope_sup)
            except Exception as e_err:
                st.error(f"❌ Error in module code: {str(e_err)}")
                
            st.markdown("<br><br>", unsafe_allow_html=True)
            if st.button(f"🗑️ Move Module '{cur_sup_mod['name']}' to Trash", key=f"del_sup_mod_p_{cur_sup_mod['id']}"):
                conn = sqlite3.connect("enterprise_logistics_sales_hub.db", timeout=10)
                cursor = conn.cursor()
                cursor.execute("UPDATE dynamic_modules_ledger SET is_trashed = 1, deleted_at = ? WHERE id = ?", (get_ist_now().strftime("%Y-%m-%d %H:%M:%S"), cur_sup_mod['id']))
                conn.commit()
                conn.close()
                st.warning(f"Module '{cur_sup_mod['name']}' moved to Recycle Bin!")
                st.rerun()

# --- GLOBAL STYLING INJECTION ---
if "sup_bg" in st.session_state:
    su_bg = st.session_state.get("sup_bg", "#f4f6f9")
    su_txt = st.session_state.get("sup_txt", "#1f2937")
    su_btn_bg = st.session_state.get("sup_btn_bg", "#1e3a8a")
    su_btn_txt = st.session_state.get("sup_btn_txt", "#ffffff")
    
    st.markdown(
        f"""
        <style>
            .stApp {{
                background-color: {su_bg} !important;
                color: {su_txt} !important;
            }}
            h1, h2, h3, h4, h5, h6, p, span, label, .stMarkdown {{
                color: {su_txt} !important;
            }}
            .stButton>button {{
                background-color: {su_btn_bg} !important;
                color: {su_btn_txt} !important;
            }}
            .stButton>button p {{
                color: {su_btn_txt} !important;
            }}
        </style>
        """,
        unsafe_allow_html=True
    )
import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import datetime
import pytz
import io
import sqlite3
import smtplib
from email.message import EmailMessage
import streamlit.components.v1 as components

# ==============================================================================
# PAGE CONFIGURATION & TIMEZONE
# ==============================================================================
try:
    st.set_page_config(
        page_title="SAP Inventory Expiry Intelligence Hub",
        page_icon="⏳",
        layout="wide",
        initial_sidebar_state="expanded"
    )
except:
    pass

IST = pytz.timezone('Asia/Kolkata')
def get_ist_now():
    return datetime.datetime.now(IST)

# ==============================================================================
# DATABASE INITIALIZATION (PERMANENT BLOB & RULES STORAGE)
# ==============================================================================
def init_db():
    try:
        conn = sqlite3.connect("inventory_master_hub.db", check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS saved_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                upload_date TEXT,
                file_name TEXT,
                file_blob BLOB
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS material_shelf_mapping (
                material_code TEXT PRIMARY KEY,
                shelf_days INTEGER
            )
        """)
        conn.commit()
        conn.close()
    except:
        pass

init_db()

def get_saved_shelf_mappings():
    try:
        conn = sqlite3.connect("inventory_master_hub.db", check_same_thread=False)
        df_map = pd.read_sql("SELECT material_code, shelf_days FROM material_shelf_mapping", conn)
        conn.close()
        return dict(zip(df_map['material_code'].astype(str).str.strip().str.lower(), df_map['shelf_days']))
    except:
        return {}

def save_multiple_mappings(m_codes_list, s_days):
    try:
        conn = sqlite3.connect("inventory_master_hub.db", check_same_thread=False)
        cursor = conn.cursor()
        for m_code in m_codes_list:
            cursor.execute(
                "INSERT OR REPLACE INTO material_shelf_mapping (material_code, shelf_days) VALUES (?, ?)",
                (str(m_code).strip().lower(), int(s_days))
            )
        conn.commit()
        conn.close()
        return True
    except:
        return False

# ==============================================================================
# SESSION STATE MANAGEMENT
# ==============================================================================
if "active_df" not in st.session_state:
    st.session_state.active_df = None
if "selected_file_id" not in st.session_state:
    st.session_state.selected_file_id = None

# ==============================================================================
# UI STYLING & LIVE RUNNING DIGITAL CLOCK HEADER
# ==============================================================================
st.markdown("""
    <style>
        .stApp {
            background-color: #0f172a;
            color: #f8fafc;
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        }
        .main-hero {
            background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%);
            padding: 24px;
            border-radius: 16px;
            border: 1px solid #334155;
            box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.4);
            margin-bottom: 20px;
        }
        .live-clock-box {
            background: #1e293b;
            color: #ffffff !important;
            padding: 14px 22px;
            border-radius: 12px;
            font-size: 16px;
            font-weight: 700;
            border: 2px solid #38bdf8;
            display: inline-block;
            text-align: center;
            box-shadow: 0 4px 15px rgba(56, 189, 248, 0.3);
            width: 100%;
        }
    </style>
""", unsafe_allow_html=True)

# Live Running Clock Component (Seconds auto-syncing)
clock_html = """
    <div style="width: 100%;">
        <div class="live-clock-box" id="liveClock">🕒 Initializing Clock...</div>
    </div>
    <script>
        function updateClock() {
            const now = new Date();
            const options = { timeZone: 'Asia/Kolkata', hour12: true, year: 'numeric', month: '2-digit', day: '2-digit', hour: '2-digit', minute: '2-digit', second: '2-digit' };
            document.getElementById('liveClock').innerHTML = '🕒 ' + now.toLocaleString('en-IN', options) + ' IST';
        }
        setInterval(updateClock, 1000);
        updateClock();
    </script>
"""

col_h1, col_h2 = st.columns([2, 1])
with col_h1:
    st.markdown("""
        <div class="main-hero" style="margin-bottom:0px; padding:18px;">
            <h2 style="color: #f8fafc; margin: 0;">⏳ SAP Production & Expiry Intelligence Hub</h2>
            <p style="color: #94a3b8; margin: 6px 0 0 0; font-size: 13px;">
                Permanent BLOB Storage, ID Reset, Live Clock, Multi-Select Rules & HTML Email Dispatch.
            </p>
        </div>
    """, unsafe_allow_html=True)
with col_h2:
    st.markdown("<br>", unsafe_allow_html=True)
    components.html(clock_html, height=75)

st.markdown("<br>", unsafe_allow_html=True)

default_shelf_days = st.number_input(
    "⏱️ Default Global Shelf-Life Period (in Days)", 
    min_value=1, 
    max_value=1095, 
    value=180, 
    step=30, 
    key="global_def_days_input"
)

# ==============================================================================
# CORE PROCESSING FUNCTION
# ==============================================================================
def process_dataframe(df_raw):
    try:
        date_keywords = ["production date", "mfg date", "production_date", "mfg_date", "mfg", "production"]
        mfg_col_found = None
        for col_name in df_raw.columns:
            c_str = str(col_name).strip().lower()
            if any(kw in c_str for kw in date_keywords):
                mfg_col_found = col_name
                break

        mat_col_found = None
        for col_name in df_raw.columns:
            c_low = str(col_name).strip().lower()
            if any(k in c_low for k in ["material", "sku", "item", "code", "product"]):
                mat_col_found = col_name
                break

        if mfg_col_found is not None:
            today_dt = pd.Timestamp(get_ist_now().date())
            df_raw['Parsed_Mfg_Date'] = pd.to_datetime(df_raw[mfg_col_found], errors='coerce')
            
            db_rules = get_saved_shelf_mappings()
            
            shelf_days_list = []
            for idx, row in df_raw.iterrows():
                assigned_days = default_shelf_days
                if mat_col_found is not None:
                    m_val = str(row[mat_col_found]).strip().lower()
                    if m_val in db_rules:
                        assigned_days = db_rules[m_val]
                shelf_days_list.append(assigned_days)
                
            df_raw['Assigned_Shelf_Days'] = shelf_days_list
            df_raw['Calculated_Expiry_Date'] = df_raw['Parsed_Mfg_Date'] + pd.to_timedelta(df_raw['Assigned_Shelf_Days'], unit='d')
            df_raw['Remaining_Shelf_Life_Days'] = (df_raw['Calculated_Expiry_Date'] - today_dt).dt.days

            conditions = [
                df_raw['Remaining_Shelf_Life_Days'].isna(),
                df_raw['Remaining_Shelf_Life_Days'] < 0,
                df_raw['Remaining_Shelf_Life_Days'] <= 30
            ]
            choices = ["Unknown Date", "🔴 Expired", "🟡 Critical (<30 Days)"]
            df_raw['Shelf_Life_Status'] = np.select(conditions, choices, default="🟢 Fresh Stock")
            
            def make_remark(val):
                if pd.isna(val):
                    return "Review Date"
                elif val < 0:
                    return "Immediate Action: Stock Expired!"
                elif val <= 30:
                    return "Priority Dispatch / Clearance Required"
                else:
                    return "Stock Condition Healthy"
            
            df_raw['Inventory_Remarks'] = df_raw['Remaining_Shelf_Life_Days'].apply(make_remark)
            
            return df_raw, mat_col_found
        else:
            st.warning("⚠️ File mein 'Production Date' column auto-detect nahi ho paya.")
            return None, None
    except Exception as e_proc:
        st.error(f"❌ Processing error: {str(e_proc)}")
        return None, None

# ==============================================================================
# LEFT SIDEBAR: COLLAPSIBLE MASTER SUITE TAB, UPLOADER & TABLE NAVIGATION
# ==============================================================================
with st.sidebar:
    with st.expander("📦 Logistics Master Suite", expanded=True):
        st.markdown("🔹 Inbound Demand Engine")
        st.markdown("🔹 Route Dispatch Planner")
        st.markdown("🔹 Live ERP Matcher")
        st.markdown("🟢 **SAP Expiry Intelligence Hub**")

    st.markdown("---")
    st.markdown("### 📂 Upload New SAP Export")
    uploaded_sap_file = st.file_uploader("Upload .xlsx or .csv", type=["xlsx", "csv"], key="sidebar_uploader")

    if uploaded_sap_file is not None:
        try:
            file_bytes = uploaded_sap_file.getvalue()
            if uploaded_sap_file.name.endswith('.csv'):
                raw_df = pd.read_csv(io.BytesIO(file_bytes))
            else:
                excel_obj = pd.ExcelFile(io.BytesIO(file_bytes))
                sheet_target = 'SAPUI5 Export' if 'SAPUI5 Export' in excel_obj.sheet_names else excel_obj.sheet_names[0]
                raw_df = pd.read_excel(excel_obj, sheet_name=sheet_target)

            processed_df, _ = process_dataframe(raw_df)
            if processed_df is not None:
                if st.button("💾 Save & Replace in Permanent DB", type="primary"):
                    conn = sqlite3.connect("inventory_master_hub.db", check_same_thread=False)
                    upload_timestamp = get_ist_now().strftime('%Y-%m-%d %H:%M:%S')
                    cursor = conn.cursor()
                    
                    # Auto-replace old records and reset ID sequence
                    cursor.execute("DELETE FROM saved_files")
                    try:
                        cursor.execute("DELETE FROM sqlite_sequence WHERE name='saved_files'")
                    except:
                        pass

                    cursor.execute(
                        "INSERT INTO saved_files (upload_date, file_name, file_blob) VALUES (?, ?, ?)",
                        (upload_timestamp, uploaded_sap_file.name, sqlite3.Binary(file_bytes))
                    )
                    conn.commit()
                    conn.close()
                    st.session_state.active_df = processed_df
                    st.success("✅ File saved permanently! Old data replaced & IDs reset.")
                    st.rerun()
        except Exception as err_file:
            st.error(f"❌ Upload error: {str(err_file)}")

    st.markdown("---")
    st.markdown("### 🗂️ Select Table from Database")
    
    try:
        conn = sqlite3.connect("inventory_master_hub.db", check_same_thread=False)
        saved_records = pd.read_sql("SELECT id, upload_date, file_name FROM saved_files", conn)
        conn.close()
    except:
        saved_records = pd.DataFrame()

    if not saved_records.empty:
        for _, row in saved_records.iterrows():
            btn_label = f"📁 [{row['id']}] {row['file_name']}"
            if st.button(btn_label, key=f"tbl_btn_{row['id']}"):
                st.session_state.selected_file_id = row['id']
                try:
                    conn = sqlite3.connect("inventory_master_hub.db", check_same_thread=False)
                    cursor = conn.cursor()
                    cursor.execute("SELECT file_blob, file_name FROM saved_files WHERE id = ?", (row['id'],))
                    row_data = cursor.fetchone()
                    conn.close()
                    if row_data:
                        blob_data, fname = row_data
                        df_from_db = pd.read_csv(io.BytesIO(blob_data)) if fname.endswith('.csv') else pd.read_excel(io.BytesIO(blob_data))
                        processed_df, _ = process_dataframe(df_from_db)
                        st.session_state.active_df = processed_df
                        st.rerun()
                except Exception as load_err:
                    st.error(f"Load error: {load_err}")
        
        if st.button("🗑️ Delete Table & Reset IDs", type="secondary"):
            conn = sqlite3.connect("inventory_master_hub.db", check_same_thread=False)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM saved_files")
            try:
                cursor.execute("DELETE FROM sqlite_sequence WHERE name='saved_files'")
            except:
                pass
            conn.commit()
            conn.close()
            st.session_state.active_df = None
            st.success("🗑️ Table vanished and database IDs reset!")
            st.rerun()
    else:
        st.info("No saved tables found. Upload a file above.")

# ==============================================================================
# AUTO-LOAD FROM DB IF SESSION IS EMPTY
# ==============================================================================
if st.session_state.active_df is None and not saved_records.empty:
    try:
        latest_id = saved_records.iloc[-1]['id']
        conn = sqlite3.connect("inventory_master_hub.db", check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("SELECT file_blob, file_name FROM saved_files WHERE id = ?", (latest_id,))
        row_data = cursor.fetchone()
        conn.close()
        if row_data:
            blob_data, fname = row_data
            df_from_db = pd.read_csv(io.BytesIO(blob_data)) if fname.endswith('.csv') else pd.read_excel(io.BytesIO(blob_data))
            processed_df, _ = process_dataframe(df_from_db)
            st.session_state.active_df = processed_df
    except:
        pass

# ==============================================================================
# MAIN DASHBOARD & ADVANCED FILTER PANEL
# ==============================================================================
if st.session_state.active_df is not None:
    st.markdown("### 📊 Inventory Intelligence Dashboard & Filters")

    working_df = st.session_state.active_df.copy()

    mat_col_target = None
    for col_name in working_df.columns:
        if any(k in str(col_name).lower() for k in ["material", "sku", "item", "code", "product"]):
            mat_col_target = col_name
            break

    col_f1, col_f2 = st.columns(2)
    
    with col_f1:
        if mat_col_target is not None:
            unique_mats = sorted(working_df[mat_col_target].dropna().astype(str).unique().tolist())
            selected_materials = st.multiselect("🔍 Multi-Select Material Code(s):", options=unique_mats)
            if selected_materials:
                working_df = working_df[working_df[mat_col_target].astype(str).isin(selected_materials)]

    with col_f2:
        all_headers = list(working_df.columns)
        chosen_header_filter = st.selectbox("📌 Select Header to Filter by Value:", options=["-- Select Header --"] + all_headers)
        if chosen_header_filter != "-- Select Header --":
            unique_vals = sorted(working_df[chosen_header_filter].dropna().astype(str).unique().tolist())
            selected_header_vals = st.multiselect(f"Select value(s) for `{chosen_header_filter}`:", options=unique_vals)
            if selected_header_vals:
                working_df = working_df[working_df[chosen_header_filter].astype(str).isin(selected_header_vals)]

    if mat_col_target is not None:
        with st.expander("🛠️ Update Shelf-Life Rule for Multiple Materials (Permanent Save)"):
            with st.form("rule_update_form_final"):
                mat_list_exp = sorted(working_df[mat_col_target].dropna().astype(str).unique().tolist())
                sel_mat_codes = st.multiselect("Select Material Code(s)", options=mat_list_exp)
                
                shelf_choices = {
                    "30 Days": 30,
                    "60 Days": 60,
                    "90 Days": 90,
                    "180 Days": 180,
                    "365 Days (1 Year)": 365,
                    "548 Days (1.5 Years)": 548,
                    "730 Days (2 Years)": 730
                }
                sel_shelf_label = st.selectbox("Select Shelf Life Period", options=list(shelf_choices.keys()))
                submit_rule_btn = st.form_submit_button("💾 Save Rule for Selected Materials", type="primary")

                if submit_rule_btn and sel_mat_codes:
                    chosen_days = shelf_choices[sel_shelf_label]
                    save_multiple_mappings(sel_mat_codes, chosen_days)
                    
                    full_df = st.session_state.active_df.copy()
                    today_dt = pd.Timestamp(get_ist_now().date())
                    for idx, row in full_df.iterrows():
                        if str(row[mat_col_target]).strip() in sel_mat_codes:
                            full_df.loc[idx, 'Assigned_Shelf_Days'] = chosen_days
                    
                    full_df['Calculated_Expiry_Date'] = full_df['Parsed_Mfg_Date'] + pd.to_timedelta(full_df['Assigned_Shelf_Days'], unit='d')
                    full_df['Remaining_Shelf_Life_Days'] = (full_df['Calculated_Expiry_Date'] - today_dt).dt.days
                    
                    conditions = [full_df['Remaining_Shelf_Life_Days'].isna(), full_df['Remaining_Shelf_Life_Days'] < 0, full_df['Remaining_Shelf_Life_Days'] <= 30]
                    choices = ["Unknown Date", "🔴 Expired", "🟡 Critical (<30 Days)"]
                    full_df['Shelf_Life_Status'] = np.select(conditions, choices, default="🟢 Fresh Stock")
                    
                    st.session_state.active_df = full_df
                    st.success(f"✅ Rules updated successfully for selected materials!")
                    st.rerun()

    if "Shelf_Life_Status" in working_df.columns:
        status_counts = working_df["Shelf_Life_Status"].value_counts()
        m1, m2, m3 = st.columns(3)
        m1.metric("🟢 Fresh Stock", int(status_counts.get("🟢 Fresh Stock", 0)))
        m2.metric("🟡 Critical (<30 Days)", int(status_counts.get("🟡 Critical (<30 Days)", 0)))
        m3.metric("🔴 Expired", int(status_counts.get("🔴 Expired", 0)))

    st.markdown("<br>", unsafe_allow_html=True)
    st.dataframe(working_df, use_container_width=True)

    # ==========================================================================
    # ACTION BUTTONS: PRINT (FITTED COLUMNS), EXCEL & PROFESSIONAL HTML EMAIL
    # ==========================================================================
    st.markdown("---")
    st.markdown("### 🚀 Export, Print & Email Options")
    
    col_act1, col_act2, col_act3 = st.columns(3)

    with col_act1:
        # Fully responsive print HTML layout preventing column cut
        html_table_string = working_df.to_html(classes='table table-striped', index=False, border=0)
        print_html_code = f"""
            <html>
                <head>
                    <title>Inventory Report Print View</title>
                    <style>
                        body {{ font-family: 'Segoe UI', Arial, sans-serif; padding: 20px; color: #1e293b; }}
                        h2 {{ color: #0f172a; margin-bottom: 5px; }}
                        p {{ color: #475569; font-size: 13px; }}
                        table {{ width: 100%; border-collapse: collapse; margin-top: 20px; font-size: 11px; table-layout: auto; }}
                        th, td {{ border: 1px solid #cbd5e1; padding: 6px 8px; text-align: left; word-wrap: break-word; }}
                        th {{ background-color: #1e293b; color: white; font-weight: 600; }}
                        tr:nth-child(even) {{ background-color: #f8fafc; }}
                        @media print {{
                            body {{ padding: 0; }}
                            table {{ page-break-inside: auto; }}
                            tr {{ page-break-inside: avoid; page-break-after: auto; }}
                        }}
                    </style>
                </head>
                <body>
                    <h2>SAP Inventory Expiry & Stock Intelligence Report</h2>
                    <p><b>Generated on:</b> {get_ist_now().strftime('%d-%m-%Y %H:%M:%S IST')}</p>
                    {html_table_string}
                    <script>
                        window.onload = function() {{ window.print(); }}
                    </script>
                </body>
            </html>
        """
        encoded_html = io.BytesIO(print_html_code.encode('utf-8'))
        st.download_button(
            "🖨️ Print / Save Table as PDF",
            data=encoded_html,
            file_name=f"Printable_Inventory_Report_{get_ist_now().strftime('%Y%m%d_%H%M%S')}.html",
            mime="text/html",
            type="secondary",
            help="Downloads a responsive formatted HTML view that opens cleanly in print layout without cutting columns."
        )

    with col_act2:
        excel_buffer = io.BytesIO()
        working_df.to_excel(excel_buffer, index=False)
        excel_buffer.seek(0)
        st.download_button(
            "📥 Download Filtered Excel (.xlsx)",
            data=excel_buffer.getvalue(),
            file_name=f"Inventory_Report_{get_ist_now().strftime('%Y-%m-%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )

    with col_act3:
        with st.expander("✉️ Send Report via Email (HTML Table Body & Attachment)"):
            with st.form("email_dispatch_form"):
                st.markdown("**SMTP Server Settings (Gmail / Corporate Hub):**")
                smtp_host = st.text_input("SMTP Server", "smtp.gmail.com")
                smtp_port = st.number_input("SMTP Port", value=587, step=1)
                sender_email = st.text_input("Sender Email (From)", "your_email@gmail.com")
                sender_pass = st.text_input("Sender App Password", type="password", help="Use Gmail App Password")
                
                st.markdown("---")
                email_to = st.text_input("Recipient Email(s) separated by comma (,):", "recipient@example.com")
                email_sub = st.text_input("Email Subject", "🚨 SAP Inventory Expiry & Critical Stock Report")
                
                default_mail_body = (
                    "Dear Team,\n\n"
                    "Please find below the summary table and attached Excel report containing critical, expired, "
                    "and fresh stock details from SAP.\n\n"
                    "Kindly review the stock remarks and prioritize clearance for items with critical shelf-life.\n\n"
                    "Best Regards,\n"
                    "Supply Chain Management Hub"
                )
                email_body = st.text_area("Email Message", default_mail_body, height=120)
                
                send_email_btn = st.form_submit_button("📨 Send Live Email Now", type="primary")

                if send_email_btn:
                    recipients = [e.strip() for e in email_to.split(",") if e.strip()]
                    if not recipients:
                        st.error("❌ Kripya kam se kam ek valid recipient email enter karein.")
                    else:
                        try:
                            msg = EmailMessage()
                            msg['Subject'] = email_sub
                            msg['From'] = sender_email
                            msg['To'] = ", ".join(recipients)
                            
                            # Professional HTML Body formatting embedding the table
                            table_html_snippet = working_df.head(50).to_html(index=False, border=1, classes='styled-table')
                            html_content = f"""
                            <html>
                            <body style="font-family: Arial, sans-serif; color: #333; line-height: 1.6;">
                                <p>{email_body.replace(chr(10), '<br>')}</p>
                                <h3>📊 Inventory Intelligence Summary Table (Top Rows Preview):</h3>
                                <style>
                                    .styled-table {{ border-collapse: collapse; width: 100%; font-size: 11px; }}
                                    .styled-table th {{ background-color: #1e293b; color: white; padding: 6px; text-align: left; border: 1px solid #ddd; }}
                                    .styled-table td {{ padding: 6px; border: 1px solid #ddd; text-align: left; }}
                                    .styled-table tr:nth-child(even) {{ background-color: #f2f2f2; }}
                                </style>
                                {table_html_snippet}
                                <p><br><i>Note: Full comprehensive dataset is attached as an Excel sheet to this email.</i></p>
                                <p><b>Supply Chain Automated Intelligence Hub</b></p>
                            </body>
                            </html>
                            """
                            msg.set_content(email_body) # Fallback text
                            msg.add_alternative(html_content, subtype='html')

                            # Attach Excel report
                            excel_data = io.BytesIO()
                            working_df.to_excel(excel_data, index=False)
                            excel_data.seek(0)
                            msg.add_attachment(
                                excel_data.getvalue(),
                                maintype='application',
                                subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                filename=f"Inventory_Report_{get_ist_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                            )

                            server = smtplib.SMTP(smtp_host, int(smtp_port))
                            server.starttls()
                            server.login(sender_email, sender_pass)
                            server.send_message(msg)
                            server.quit()

                            st.success(f"✅ Live email with professional HTML table and Excel attachment successfully dispatched to: **{', '.join(recipients)}**!")
                        except Exception as mail_err:
                            st.error(f"❌ Email sending failed. Error details: {str(mail_err)}")
else:
    st.info("ℹ️ Kripya left sidebar se apni SAP stock export file upload karein ya saved table select karein.")
