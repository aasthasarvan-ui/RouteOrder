import streamlit as st
import pandas as pd
import openpyxl
import datetime
import pytz
import io
import re
import zipfile
import sqlite3
import smtplib
import json
import urllib.parse
from email.message import EmailMessage
from fpdf import FPDF
import streamlit.components.v1 as components

# Page Configuration & Styling
st.set_page_config(
    page_title="Enterprise Sales Order & Dispatch Hub", 
    page_icon="💼", 
    layout="wide",
    initial_sidebar_state="collapsed"
)

# 8 Professional Enterprise Themes with Unique Icons & Color Palettes
THEMES = {
    "💼 Classic Enterprise Navy": {
        "icon": "💼", "bg": "#f4f6f9", "text": "#1f2937", "card_bg": "#ffffff", "border": "#cbd5e1",
        "btn_bg": "#1e3a8a", "btn_hover": "#1d4ed8", "primary": "#2563eb", "input_bg": "#ffffff", "input_text": "#1f2937"
    },
    "🌙 Modern Dark ERP": {
        "icon": "🌙", "bg": "#0b0f19", "text": "#f3f4f6", "card_bg": "#1f2937", "border": "#374151",
        "btn_bg": "#374151", "btn_hover": "#4b5563", "primary": "#3b82f6", "input_bg": "#111827", "input_text": "#f3f4f6"
    },
    "📊 Corporate Slate": {
        "icon": "📊", "bg": "#eef2f5", "text": "#0f172a", "card_bg": "#ffffff", "border": "#94a3b8",
        "btn_bg": "#475569", "btn_hover": "#334155", "primary": "#0284c7", "input_bg": "#ffffff", "input_text": "#0f172a"
    },
    "☀️ Clean Light Minimal": {
        "icon": "☀️", "bg": "#ffffff", "text": "#111827", "card_bg": "#f9fafb", "border": "#d1d5db",
        "btn_bg": "#0f172a", "btn_hover": "#1e293b", "primary": "#10b981", "input_bg": "#ffffff", "input_text": "#111827"
    },
    "⚡ Cyber Blue": {
        "icon": "⚡", "bg": "#f0fdfa", "text": "#042f2e", "card_bg": "#ccfbf1", "border": "#5eead4",
        "btn_bg": "#0d9488", "btn_hover": "#0f766e", "primary": "#14b8a6", "input_bg": "#ffffff", "input_text": "#042f2e"
    },
    "🌲 Emerald Corporate": {
        "icon": "🌲", "bg": "#f0fdf4", "text": "#14532d", "card_bg": "#dcfce7", "border": "#86efac",
        "btn_bg": "#16a34a", "btn_hover": "#15803d", "primary": "#22c55e", "input_bg": "#ffffff", "input_text": "#14532d"
    },
    "🍇 Executive Burgundy": {
        "icon": "🍇", "bg": "#fdf2f8", "text": "#500724", "card_bg": "#fce7f3", "border": "#f472b6",
        "btn_bg": "#db2777", "btn_hover": "#be185d", "primary": "#ec4899", "input_bg": "#ffffff", "input_text": "#500724"
    },
    "🪙 Titanium Charcoal": {
        "icon": "🪙", "bg": "#18181b", "text": "#fafafa", "card_bg": "#27272a", "border": "#52525b",
        "btn_bg": "#52525b", "btn_hover": "#71717a", "primary": "#e4e4e7", "input_bg": "#09090b", "input_text": "#fafafa"
    }
}

# IST Timezone Helper via pytz
IST = pytz.timezone('Asia/Kolkata')

def get_ist_now():
    return datetime.datetime.now(IST)

# --- SQLite Database Initialization with Extended Logistics & Dispatch Tables ---
def init_db():
    conn = sqlite3.connect("sales_history.db")
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS history_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            files_count INTEGER,
            total_qty REAL,
            status TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS unique_routes_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name TEXT,
            route_no TEXT,
            agency_no TEXT,
            dr_code TEXT,
            created_at TEXT,
            UNIQUE(route_no, agency_no, dr_code)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS output_files_ledger (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name TEXT UNIQUE,
            file_type TEXT,
            file_data BLOB,
            created_at TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS unmapped_missing_dr_ledger (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name TEXT,
            route_no TEXT,
            agency_no TEXT,
            dr_code TEXT,
            created_at TEXT,
            UNIQUE(route_no, agency_no)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS input_output_traceability (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_timestamp TEXT,
            input_file_name TEXT,
            input_file_blob BLOB,
            total_input_qty REAL,
            generated_output_file TEXT,
            output_type TEXT,
            version_no INTEGER,
            created_at TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS discrepancy_audit_ledger (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_timestamp TEXT,
            file_name TEXT,
            agency_no TEXT,
            dr_code TEXT,
            fg_code TEXT,
            input_qty REAL,
            generated_qty REAL,
            difference REAL,
            logged_at TEXT
        )
    """)
    # --- ENHANCED DISPATCH MODULE TABLES ---
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS vehicles_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vehicle_no TEXT UNIQUE,
            vehicle_type TEXT,
            capacity_weight REAL,
            driver_name TEXT,
            driver_phone TEXT,
            status TEXT,
            created_at TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS dispatch_plans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dispatch_id TEXT UNIQUE,
            route_no TEXT,
            vehicle_no TEXT,
            driver_name TEXT,
            total_orders INTEGER,
            total_quantity REAL,
            dispatch_status TEXT,
            scheduled_date TEXT,
            created_at TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS dispatch_item_mapping (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dispatch_id TEXT,
            agency_no TEXT,
            dr_code TEXT,
            order_qty REAL,
            file_name TEXT,
            created_at TEXT,
            FOREIGN KEY(dispatch_id) REFERENCES dispatch_plans(dispatch_id)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS pending_dispatch_queue (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            route_no TEXT,
            agency_no TEXT,
            dr_code TEXT,
            pending_qty REAL,
            file_name TEXT,
            status TEXT,
            created_at TEXT
        )
    """)
    conn.commit()
    conn.close()

init_db()

# --- Session State Defaults ---
DEFAULTS = {
    "fg_code": "FG500014",
    "col_map": "36:FG500014AJ\n37:FG500014AK",
    "agency_override": "101:36:FG500014N01\n101:37:FG500014N02",
    "route": "22",
    "email_user": st.secrets.get("email", {}).get("sender_email", ""),
    "email_pass": st.secrets.get("email", {}).get("app_password", ""),
    "recipient": st.secrets.get("email", {}).get("recipient_email", ""),
    "whatsapp": "",
    "selected_theme": "💼 Classic Enterprise Navy",
    "processed_files": [],
    "comparison_summary": [],
    "skipped_rows_log": [],
    "anomaly_logs": [],
    "unmapped_current_batch": [],
    "kpi_data": {"input_qty": 0, "gen_qty": 0, "valid_count": 0, "missing_count": 0, "skipped_count": 0}
}

for key, val in DEFAULTS.items():
    if key not in st.session_state:
        st.session_state[key] = val

t = THEMES[st.session_state.selected_theme]

# Professional CSS Injection
st.markdown(
    f"""
    <style>
        #GithubIcon {{ visibility: hidden !important; display: none !important; }}
        .stAppHeader {{ background-color: transparent !important; }}
        header[data-testid="stHeader"] {{ display: none !important; }}
        
        .stApp {{
            background-color: {t['bg']};
            color: {t['text']};
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        }}
        h1, h2, h3, h4, h5, h6, p, span, label, .stMarkdown {{ color: {t['text']} !important; }}
        input, textarea, select {{
            background-color: {t['input_bg']} !important;
            color: {t['input_text']} !important;
            border: 1px solid {t['border']} !important;
        }}
        .stButton>button {{
            width: 100%;
            height: 38px;
            background-color: {t['btn_bg']} !important;
            color: #ffffff !important;
            font-size: 13px !important;
            font-weight: 600 !important;
            border-radius: 4px;
            border: 1px solid {t['border']};
        }}
        .stButton>button:hover {{ background-color: {t['btn_hover']} !important; }}
        div[data-testid="stExpander"] {{
            background-color: {t['card_bg']};
            border: 1px solid {t['border']};
            border-radius: 4px;
        }}
    </style>
    """,
    unsafe_allow_html=True
)

# --- TOP COLLAPSIBLE CONTROL PANEL ---
with st.expander("⚙️ Enterprise Control Panel & Theme Settings", expanded=False):
    def on_theme_change():
        st.session_state.selected_theme = st.session_state.theme_selectbox

    st.selectbox("Select Interface Theme", list(THEMES.keys()), key="theme_selectbox", index=list(THEMES.keys()).index(st.session_state.selected_theme), on_change=on_theme_change)
    
    col_set1, col_set2, col_set3 = st.columns(3)
    with col_set1:
        st.session_state.fg_code = st.text_input("Default Fallback FG Code", value=st.session_state.fg_code)
        st.session_state.route = st.text_input("Default Route Fallback", value=st.session_state.route)
    with col_set2:
        st.session_state.col_map = st.text_area("Direct Column Mapping", value=st.session_state.col_map, height=80)
    with col_set3:
        st.session_state.agency_override = st.text_area("Agency FG Override", value=st.session_state.agency_override, height=80)

# Parsing Config
default_fg_code = st.session_state.fg_code
col_mapping_input = st.session_state.col_map
agency_fg_override = st.session_state.agency_override
default_fallback_route = st.session_state.route
email_user = st.session_state.email_user
email_pass = st.session_state.email_pass
recipient_email = st.session_state.recipient
whatsapp_num = st.session_state.whatsapp

direct_col_mapping = {}
for line in col_mapping_input.split('\n'):
    if ':' in line:
        parts = line.split(':')
        if parts[0].strip().isdigit():
            direct_col_mapping[int(parts[0].strip())] = parts[1].strip()

agency_col_override_map = {}
for line in agency_fg_override.split('\n'):
    parts = line.split(':')
    if len(parts) == 3 and parts[0].strip().isdigit() and parts[1].strip().isdigit():
        agency_col_override_map[(int(parts[0].strip()), int(parts[1].strip()))] = parts[2].strip()

st.title(f"💼 Enterprise Sales Order & Dispatch Hub ({st.session_state.selected_theme})")
st.markdown("Upload inbound demand spreadsheets, execute capacity-optimized dispatches, and audit demand vs actual fulfillment.")
st.markdown("---")

uploaded_inputs = st.file_uploader("Upload Multiple Demand Excel Files", type=["xlsx", "xls"], accept_multiple_files=True, key="inputs")

if st.button("🚀 Process Batch Orders & Update Master DB", type="primary"):
    if uploaded_inputs:
        st.session_state.processed_files = []
        st.session_state.comparison_summary = []
        st.session_state.skipped_rows_log = []
        st.session_state.anomaly_logs = []
        st.session_state.unmapped_current_batch = []
        
        total_input_qty = 0
        total_gen_qty = 0
        total_valid_orders = 0
        total_missing_orders = 0
        total_skipped_rows = 0
        
        db_records_to_insert = []
        unmapped_records_to_insert = []
        output_files_to_store = []
        traceability_records = []
        discrepancy_records = []
        pending_queue_inserts = []
        
        with st.spinner("⚡ Processing batch demand files & generating structured outputs..."):
            try:
                try:
                    with open("Output.xlsx", "rb") as f:
                        template_bytes = f.read()
                except FileNotFoundError:
                    st.error("❌ 'Output.xlsx' template file missing in root repository.")
                    st.stop()
                
                ist_now = get_ist_now()
                today_date = ist_now.strftime("%Y-%m-%d")
                timestamp = ist_now.strftime("%H%M%S")
                batch_ts = ist_now.strftime("%Y-%m-%d %H:%M:%S")

                for uploaded_file in uploaded_inputs:
                    short_filename = uploaded_file.name
                    if short_filename.lower() == "output.xlsx":
                        continue

                    file_bytes = uploaded_file.getvalue()
                    df_input = pd.read_excel(io.BytesIO(file_bytes), header=None)

                    fg_row, fg_col = -1, -1
                    for r in range(df_input.shape[0]):
                        for c in range(df_input.shape[1]):
                            if "FG" in str(df_input.iloc[r, c]).strip().upper():
                                fg_row, fg_col = r, c
                                break
                        if fg_row != -1:
                            break

                    if fg_row == -1:
                        continue

                    total_col = df_input.shape[1]
                    for cSearch in range(fg_col, df_input.shape[1]):
                        is_total = False
                        for scan_r in range(max(0, fg_row - 10), min(fg_row + 3, df_input.shape[0])):
                            if any(kw in str(df_input.iloc[scan_r, cSearch]).strip().upper() for kw in ["TOTAL", "SUM", "TOT", "TTL", "NET"]):
                                is_total = True
                                break
                        if is_total:
                            total_col = cSearch
                            break

                    route_num = default_fallback_route if default_fallback_route != "" else "22"
                    for r in range(fg_row):
                        for c in range(min(total_col, 30)):
                            cell_val = str(df_input.iloc[r, c]).strip()
                            if cell_val != "" and 1 <= len(cell_val) <= 3 and any(ch.isdigit() for ch in cell_val):
                                route_num = cell_val
                                break

                    agency_col = fg_col - 1 if fg_col > 0 else 0
                    dr_code_col = -1
                    for cSearch in range(fg_col - 1, -1, -1):
                        if any(re.match(r'^DR\d+', str(df_input.iloc[fg_row + offset, cSearch]).strip().upper()) for offset in range(1, min(4, df_input.shape[0] - fg_row))):
                            dr_code_col = cSearch
                            break

                    valid_cols = []
                    for c in range(fg_col, total_col):
                        fg_code = str(df_input.iloc[fg_row, c]).strip()
                        if any(kw in fg_code.upper() for kw in ["TOTAL", "SUM", "TOT", "TTL"]):
                            break
                        valid_cols.append((c, fg_code))

                    wb_valid = openpyxl.load_workbook(io.BytesIO(template_bytes))
                    ws_valid = wb_valid["Order Data"] if "Order Data" in wb_valid.sheetnames else wb_valid.active

                    valid_row, valid_order_num = 6, 1
                    agency_counts_valid = {}
                    file_comparison_rows = []
                    file_input_qty = 0

                    for r in range(fg_row + 1, df_input.shape[0]):
                        agency = df_input.iloc[r, agency_col] if agency_col >= 0 else None
                        if pd.isna(agency) or str(agency).strip() in ["", "nan", "None"]:
                            continue
                        
                        agency_str = str(agency).replace('.0','').strip()
                        if not agency_str.isdigit():
                            continue
                        agency_val = int(agency_str)
                        
                        row_has_items = False
                        valid_row_quantities = []
                        row_total_qty = 0
                        for c, fg_code in valid_cols:
                            sku_qty = df_input.iloc[r, c]
                            if pd.notna(sku_qty) and str(sku_qty).strip() != "":
                                try:
                                    qty_val = float(sku_qty)
                                    if qty_val > 0:
                                        row_has_items = True
                                        row_total_qty += qty_val
                                        file_input_qty += qty_val
                                        valid_row_quantities.append((c, fg_code, qty_val))
                                except ValueError:
                                    pass

                        if not row_has_items:
                            continue

                        has_dr_code, clean_dr = False, ""
                        if dr_code_col >= 0:
                            raw_dr = df_input.iloc[r, dr_code_col]
                            if pd.notna(raw_dr) and "DR" in str(raw_dr).upper():
                                has_dr_code, clean_dr = True, str(raw_dr).replace('.0','').strip()

                        if not has_dr_code:
                            conn_lookup = sqlite3.connect("sales_history.db")
                            cursor_lookup = conn_lookup.cursor()
                            cursor_lookup.execute("SELECT dr_code FROM unique_routes_master WHERE route_no = ? AND agency_no = ? LIMIT 1", (str(route_num), str(agency_val)))
                            db_match = cursor_lookup.fetchone()
                            conn_lookup.close()
                            if db_match:
                                has_dr_code, clean_dr = True, db_match[0]

                        final_dr = clean_dr if has_dr_code else f"NEW_CUST_{agency_val}"
                        if not has_dr_code:
                            pending_queue_inserts.append((str(route_num), str(agency_val), final_dr, row_total_qty, short_filename, "Pending Unmapped", batch_ts))

                        agency_counts_valid[agency_val] = agency_counts_valid.get(agency_val, 0) + 1
                        ref_number = f"RT-{route_num}-{agency_val}-{today_date}"

                        item_id = 10
                        for c, fg_code, qty_val in valid_row_quantities:
                            current_fg = agency_col_override_map.get((agency_val, c), fg_code if str(fg_code).upper().startswith("FG") else direct_col_mapping.get(c, default_fg_code))
                            total_input_qty += qty_val
                            total_gen_qty += qty_val
                            
                            file_comparison_rows.append({
                                "File Name": short_filename, "Agency": agency_val, "DR Code": final_dr, "FG Code": current_fg, "Input Qty": qty_val, "Generated Qty": qty_val
                            })

                            ws_valid.cell(row=valid_row, column=2, value=valid_order_num)
                            ws_valid.cell(row=valid_row, column=3, value="OR")
                            ws_valid.cell(row=valid_row, column=4, value="SO20")
                            ws_valid.cell(row=valid_row, column=7, value=final_dr)
                            ws_valid.cell(row=valid_row, column=9, value=ref_number)
                            ws_valid.cell(row=valid_row, column=16, value=current_fg)
                            ws_valid.cell(row=valid_row, column=19, value=qty_val)
                            ws_valid.cell(row=valid_row, column=26, value=str(route_num))
                            ws_valid.cell(row=valid_row, column=27, value=agency_val)
                            
                            item_id += 10
                            valid_row += 1

                        valid_order_num += 1
                        total_valid_orders += 1

                    if valid_row > 6:
                        buf_valid = io.BytesIO()
                        wb_valid.save(buf_valid)
                        buf_valid.seek(0)
                        out_fname = f"{route_num}_{today_date}_{timestamp}_Processed.xlsx"
                        st.session_state.processed_files.append({"name": short_filename, "data": buf_valid.getvalue(), "filename": out_fname})
                        output_files_to_store.append((out_fname, "Processed Order", buf_valid.getvalue(), batch_ts))
                        traceability_records.append((batch_ts, short_filename, file_bytes, file_input_qty, out_fname, "Processed", 1, batch_ts))

                conn = sqlite3.connect("sales_history.db")
                cursor = conn.cursor()
                cursor.executemany("INSERT OR IGNORE INTO pending_dispatch_queue (route_no, agency_no, dr_code, pending_qty, file_name, status, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)", pending_queue_inserts)
                for fname, ftype, fdata, fdate in output_files_to_store:
                    cursor.execute("INSERT OR IGNORE INTO output_files_ledger (file_name, file_type, file_data, created_at) VALUES (?, ?, ?, ?)", (fname, ftype, fdata, fdate))
                conn.commit()
                conn.close()

                st.session_state.kpi_data = {"input_qty": total_input_qty, "gen_qty": total_gen_qty, "valid_count": total_valid_orders, "missing_count": total_missing_orders, "skipped_count": total_skipped_rows}
                st.success("✅ Batch processing and queue updates completed successfully!")
            except Exception as e:
                st.error(f"❌ Processing Error: {str(e)}")

# --- KPI DISPLAY ---
if st.session_state.processed_files:
    kpi = st.session_state.kpi_data
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Input Quantity", f"{kpi['input_qty']:,.0f}")
    c2.metric("Generated Quantity", f"{kpi['gen_qty']:,.0f}")
    c3.metric("Valid Orders Processed", kpi['valid_count'])
    c4.metric("Pending Queue Items", len(pending_queue_inserts) if 'pending_queue_inserts' in locals() else 0)

# --- NEW ADVANCED DISPATCH & LOGISTICS HUB ---
st.markdown("---")
st.subheader("🚚 Enterprise Capacity-Optimized Dispatch & Fleet Management Hub")

dispatch_tab1, dispatch_tab2, dispatch_tab3, dispatch_tab4 = st.tabs([
    "🚛 Vehicle Fleet Setup", 
    "📋 Capacity-Optimized Trip Planner", 
    "📊 Master Dispatch, Pending & Filter Hub",
    "📬 Export, Print & Email Manifests"
])

# --- TAB 1: VEHICLE FLEET SETUP ---
with dispatch_tab1:
    st.markdown("#### Register Delivery Trucks & Maximum Capacity (Bags/Units)")
    vc1, vc2 = st.columns(2)
    with vc1:
        with st.form("veh_form"):
            v_no = st.text_input("Vehicle Number (e.g., PB08AB1234)").upper()
            v_type = st.selectbox("Vehicle Type", ["10-Tyre Truck (Capacity: 300 Bags)", "6-Tyre Truck (Capacity: 150 Bags)", "Tata Ace (Capacity: 50 Bags)"])
            v_cap = st.number_input("Max Capacity (Units/Bags)", min_value=10.0, value=100.0, step=10.0)
            d_name = st.text_input("Driver Name")
            d_phone = st.text_input("Driver Contact Phone")
            if st.form_submit_button("➕ Register Vehicle"):
                if v_no and d_name:
                    conn = sqlite3.connect("sales_history.db")
                    cur = conn.cursor()
                    cur.execute("INSERT OR REPLACE INTO vehicles_master (vehicle_no, vehicle_type, capacity_weight, driver_name, driver_phone, status, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                                (v_no, v_type, v_cap, d_name, d_phone, "Available", get_ist_now().strftime("%Y-%m-%d %H:%M:%S")))
                    conn.commit()
                    conn.close()
                    st.success(f"✅ Vehicle {v_no} registered with capacity {v_cap} units!")
                    st.rerun()
                else:
                    st.warning("⚠️ Enter Vehicle Number and Driver Name.")
    with vc2:
        st.markdown("##### Active Fleet Registry")
        try:
            conn = sqlite3.connect("sales_history.db")
            df_v = pd.read_sql("SELECT * FROM vehicles_master", conn)
            conn.close()
            if not df_v.empty:
                st.dataframe(df_v, use_container_width=True)
            else:
                st.info("No vehicles registered.")
        except Exception as e:
            st.error(str(e))

# --- TAB 2: CAPACITY-OPTIMIZED TRIP PLANNER ---
with dispatch_tab2:
    st.markdown("#### Knapsack-Style Capacity Auto-Matching Trip Planner")
    st.markdown("Yeh system gadi ki capacity (jaise 100 bags) ke anusaar agency demands ko ek-ek karke fill karega. Agar gadi ki capacity se demand zyada hai, toh baki bachi demand **Pending Dispatch Queue** mein save ho jayegi agle trip ke liye.")

    try:
        conn = sqlite3.connect("sales_history.db")
        routes_list = pd.read_sql("SELECT DISTINCT route_no FROM unique_routes_master", conn)['route_no'].tolist()
        avail_vehicles = pd.read_sql("SELECT vehicle_no, capacity_weight FROM vehicles_master WHERE status='Available'", conn)
        conn.close()

        pc1, pc2 = st.columns(2)
        with pc1:
            selected_route = st.selectbox("Select Route", ["Select..."] + routes_list)
            selected_truck = st.selectbox("Select Available Vehicle", ["Select..."] + avail_vehicles['vehicle_no'].tolist() if not avail_vehicles.empty else ["Select..."])
        with pc2:
            trip_date = st.date_input("Dispatch Execution Date", value=get_ist_now())
            trip_notes = st.text_area("Dispatch Instructions", "Standard dispatch route order execution.")

        if st.button("⚡ Generate Capacity-Optimized Dispatch Plan", type="primary"):
            if selected_route != "Select..." and selected_truck != "Select...":
                conn = sqlite3.connect("sales_history.db")
                cur = conn.cursor()
                
                # Get vehicle capacity
                cur.execute("SELECT capacity_weight, driver_name FROM vehicles_master WHERE vehicle_no = ?", (selected_truck,))
                veh_info = cur.fetchone()
                max_capacity = veh_info[0] if veh_info else 100.0
                driver_name = veh_info[1] if veh_info else "Assigned Driver"

                # Get agencies for route
                cur.execute("SELECT agency_no, dr_code, file_name FROM unique_routes_master WHERE route_no = ?", (selected_route,))
                agencies = cur.fetchall()

                if agencies:
                    dispatch_id = f"DISP-{selected_route}-{get_ist_now().strftime('%Y%m%d%H%M%S')}"
                    allocated_qty = 0.0
                    allocated_items = 0
                    
                    cur.execute("INSERT INTO dispatch_plans (dispatch_id, route_no, vehicle_no, driver_name, total_orders, total_quantity, dispatch_status, scheduled_date, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                                (dispatch_id, selected_route, selected_truck, driver_name, 0, 0.0, "Planned", str(trip_date), get_ist_now().strftime("%Y-%m-%d %H:%M:%S")))

                    for ag_no, dr_c, f_name in agencies:
                        agency_demand = 150.0  # Estimated standard order size per agency baseline
                        if allocated_qty + agency_demand <= max_capacity:
                            cur.execute("INSERT INTO dispatch_item_mapping (dispatch_id, agency_no, dr_code, order_qty, file_name, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                                        (dispatch_id, ag_no, dr_c, agency_demand, f_name, get_ist_now().strftime("%Y-%m-%d %H:%M:%S")))
                            allocated_qty += agency_demand
                            allocated_items += 1
                        else:
                            # Push remaining unallocated demand to pending queue
                            cur.execute("INSERT INTO pending_dispatch_queue (route_no, agency_no, dr_code, pending_qty, file_name, status, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                                        (selected_route, ag_no, dr_c, agency_demand, f_name, "Pending Capacity Overflow", get_ist_now().strftime("%Y-%m-%d %H:%M:%S")))

                    # Update total planned quantities & orders
                    cur.execute("UPDATE dispatch_plans SET total_orders = ?, total_quantity = ? WHERE dispatch_id = ?", (allocated_items, allocated_qty, dispatch_id))
                    cur.execute("UPDATE vehicles_master SET status = 'Dispatched' WHERE vehicle_no = ?", (selected_truck,))
                    
                    conn.commit()
                    conn.close()
                    st.success(f"✅ Dispatch Plan **{dispatch_id}** created successfully! Allocated Qty: {allocated_qty} / Max Capacity: {max_capacity}")
                else:
                    conn.close()
                    st.warning("⚠️ No agencies found for this route.")
            else:
                st.warning("⚠️ Please select both a Route and a Vehicle.")
    except Exception as e:
        st.error(f"Error in trip planning: {str(e)}")

# --- TAB 3: MASTER DISPATCH, PENDING & MONTHLY FILTER HUB ---
with dispatch_tab3:
    st.markdown("#### Master Dispatch, Pending Queue & Monthly Filter Hub")
    
    try:
        conn = sqlite3.connect("sales_history.db")
        df_all_plans = pd.read_sql("SELECT * FROM dispatch_plans", conn)
        df_all_pending = pd.read_sql("SELECT * FROM pending_dispatch_queue", conn)
        df_all_mapping = pd.read_sql("SELECT * FROM dispatch_item_mapping", conn)
        conn.close()

        # Search & Filter Controls
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            filter_status = st.selectbox("Filter by Status", ["All", "Planned", "Dispatched", "Delivered", "Cancelled"])
        with fc2:
            filter_vehicle = st.selectbox("Filter by Vehicle", ["All"] + df_all_plans['vehicle_no'].unique().tolist() if not df_all_plans.empty else ["All"])
        with fc3:
            search_query = st.text_input("Search Dispatch ID / Route")

        filtered_plans = df_all_plans.copy()
        if filter_status != "All":
            filtered_plans = filtered_plans[filtered_plans['dispatch_status'] == filter_status]
        if filter_vehicle != "All":
            filtered_plans = filtered_plans[filtered_plans['vehicle_no'] == filter_vehicle]
        if search_query:
            filtered_plans = filtered_plans[filtered_plans['dispatch_id'].str.contains(search_query, case=False) | filtered_plans['route_no'].str.contains(search_query, case=False)]

        st.markdown("##### 📋 Master Dispatch Ledger")
        if not filtered_plans.empty:
            st.dataframe(filtered_plans, use_container_width=True)
        else:
            st.info("No matching dispatch records found.")

        st.markdown("##### ⏳ Pending Dispatch Queue (Unfulfilled / Overflow Demand)")
        if not df_all_pending.empty:
            st.dataframe(df_all_pending, use_container_width=True)
        else:
            st.success("No pending items in queue. All demand fulfilled!")

        st.markdown("##### 🔍 Demand vs Actual Fulfillment Audit")
        if not df_all_mapping.empty:
            merged_audit = df_all_mapping.merge(df_all_plans, on="dispatch_id", how="left")
            st.dataframe(merged_audit[['dispatch_id', 'route_no', 'vehicle_no', 'agency_no', 'dr_code', 'order_qty', 'dispatch_status', 'scheduled_date']], use_container_width=True)
        else:
            st.info("No mapping audit records available.")

    except Exception as e:
        st.error(f"Error loading filter hub: {str(e)}")

# --- TAB 4: EXPORT, PRINT & EMAIL MANIFESTS ---
with dispatch_tab4:
    st.markdown("#### Export, Print & Email Dispatch Manifests")
    
    try:
        conn = sqlite3.connect("sales_history.db")
        df_active_dispatches = pd.read_sql("SELECT dispatch_id FROM dispatch_plans", conn)
        conn.close()

        if not df_active_dispatches.empty:
            selected_disp_id = st.selectbox("Select Dispatch ID for Export", df_active_dispatches['dispatch_id'].tolist(), key="exp_disp_id")
            
            conn = sqlite3.connect("sales_history.db")
            df_manifest = pd.read_sql("SELECT * FROM dispatch_item_mapping WHERE dispatch_id = ?", conn, params=(selected_disp_id,))
            df_plan_info = pd.read_sql("SELECT * FROM dispatch_plans WHERE dispatch_id = ?", conn, params=(selected_disp_id,))
            conn.close()

            dc1, dc2, dc3, dc4 = st.columns(4)
            
            # Excel Export
            excel_buf = io.BytesIO()
            df_manifest.to_excel(excel_buf, index=False, sheet_name="Dispatch Manifest")
            excel_buf.seek(0)
            dc1.download_button("📥 Excel Manifest", data=excel_buf.getvalue(), file_name=f"Manifest_{selected_disp_id}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            # PDF Export
            try:
                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", "B", 14)
                pdf.cell(190, 10, f"Delivery Manifest: {selected_disp_id}", ln=True, align="C")
                pdf.set_font("Arial", "", 10)
                pdf.cell(190, 6, f"Scheduled Date: {df_plan_info['scheduled_date'].values[0]} | Vehicle: {df_plan_info['vehicle_no'].values[0]}", ln=True, align="C")
                pdf.ln(5)
                
                pdf.set_font("Arial", "B", 10)
                pdf.cell(50, 8, "Agency No", border=1)
                pdf.cell(50, 8, "DR Code", border=1)
                pdf.cell(40, 8, "Order Qty", border=1, ln=True)
                
                pdf.set_font("Arial", "", 10)
                for idx, row in df_manifest.iterrows():
                    pdf.cell(50, 8, str(row['agency_no']), border=1)
                    pdf.cell(50, 8, str(row['dr_code']), border=1)
                    pdf.cell(40, 8, str(row['order_qty']), border=1, ln=True)
                
                pdf_bytes = bytes(pdf.output())
                dc2.download_button("📄 PDF Manifest", data=pdf_bytes, file_name=f"Manifest_{selected_disp_id}.pdf", mime="application/pdf")
            except Exception as pdf_err:
                dc2.error(f"PDF Error: {str(pdf_err)}")

            # Print Button
            with dc3:
                print_html = f"""
                <button onclick="parent.window.print()" style="width:100%; height:38px; background:#2563eb; color:white; border:none; border-radius:4px; font-weight:600; cursor:pointer;">
                    🖨️ Print Manifest
                </button>
                """
                components.html(print_html, height=50)

            # Email Manifest Button
            with dc4:
                if st.button("📧 Email Manifest"):
                    if email_user and email_pass and recipient_email:
                        try:
                            msg = EmailMessage()
                            msg['Subject'] = f"🚚 Delivery Manifest Report - {selected_disp_id}"
                            msg['From'] = email_user
                            msg['To'] = recipient_email
                            msg.set_content(f"Please find attached delivery manifest for dispatch trip: {selected_disp_id}")
                            msg.add_attachment(excel_buf.getvalue(), maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=f"Manifest_{selected_disp_id}.xlsx")
                            
                            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                                smtp.login(email_user, email_pass)
                                smtp.send_message(msg)
                            st.success("✅ Manifest emailed successfully!")
                        except Exception as mail_err:
                            st.error(f"Email failed: {str(mail_err)}")
                    else:
                        st.warning("⚠️ Enter email credentials in control panel.")
        else:
            st.info("No dispatches available for export.")
    except Exception as e:
        st.error(f"Error in export module: {str(e)}")
