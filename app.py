import streamlit as st
import pandas as pd
import openpyxl
import datetime
import pytz
import io
import re
import zipfile
import sqlite3
import smtplib
import json
import urllib.parse
from email.message import EmailMessage
from fpdf import FPDF
import streamlit.components.v1 as components

# Page Configuration & Styling (Professional SAP ERP Layout with Top Control Panel)
st.set_page_config(
    page_title="Enterprise Sales Order Automation Hub (SAP ERP Edition)", 
    page_icon="💼", 
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Theme Dictionary with Fixed High-Contrast Typography & Button Text
THEMES = {
    "💼 SAP Classic Navy": {
        "bg": "#f4f6f9", "text": "#1f2937", "card_bg": "#ffffff", "border": "#cbd5e1",
        "btn_bg": "#1e3a8a", "btn_hover": "#1d4ed8", "btn_text": "#ffffff", "primary": "#2563eb", "input_bg": "#ffffff", "input_text": "#1f2937"
    },
    "🌙 Modern Dark ERP": {
        "bg": "#0b0f19", "text": "#f3f4f6", "card_bg": "#1f2937", "border": "#374151",
        "btn_bg": "#374151", "btn_hover": "#4b5563", "btn_text": "#ffffff", "primary": "#3b82f6", "input_bg": "#111827", "input_text": "#f3f4f6"
    },
    "📊 Corporate Slate": {
        "bg": "#eef2f5", "text": "#0f172a", "card_bg": "#ffffff", "border": "#94a3b8",
        "btn_bg": "#475569", "btn_hover": "#334155", "btn_text": "#ffffff", "primary": "#0284c7", "input_bg": "#ffffff", "input_text": "#0f172a"
    },
    "☀️ Clean Light Minimal": {
        "bg": "#ffffff", "text": "#111827", "card_bg": "#f9fafb", "border": "#d1d5db",
        "btn_bg": "#0f172a", "btn_hover": "#1e293b", "btn_text": "#ffffff", "primary": "#10b981", "input_bg": "#ffffff", "input_text": "#111827"
    }
}

# Theme Selector in Sidebar (Clean & Minimal)
st.sidebar.title("🎨 ERP Theme Engine")
selected_theme_name = st.sidebar.selectbox("Choose Interface Theme", list(THEMES.keys()), label_visibility="collapsed")
t = THEMES[selected_theme_name]

# Professional CSS Injection with Guaranteed Button Text Visibility
st.markdown(
    f"""
    <style>
        #GithubIcon {{ visibility: hidden !important; display: none !important; }}
        .stAppHeader {{ background-color: transparent !important; }}
        header[data-testid="stHeader"] {{ display: none !important; }}
        
        .stApp {{
            background-color: {t['bg']};
            color: {t['text']};
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        }}
        
        h1, h2, h3, h4, h5, h6, p, span, label, .stMarkdown {{
            color: {t['text']} !important;
        }}
        
        input, textarea, select {{
            background-color: {t['input_bg']} !important;
            color: {t['input_text']} !important;
            border: 1px solid {t['border']} !important;
        }}
        
        /* Force Button Text Visibility */
        .stButton>button {{
            width: 100%;
            height: 38px;
            background-color: {t['btn_bg']} !important;
            color: {t['btn_text']} !important;
            font-size: 13px;
            font-weight: 600;
            border-radius: 4px;
            border: 1px solid {t['border']};
            display: flex;
            align-items: center;
            justify-content: center;
            transition: all 0.2s ease;
        }}
        .stButton>button:hover {{
            background-color: {t['btn_hover']} !important;
            color: {t['btn_text']} !important;
        }}
        
        button[kind="primary"] {{
            background-color: {t['primary']} !important;
            color: #ffffff !important;
        }}
        
        div[data-testid="stExpander"] {{
            background-color: {t['card_bg']};
            border: 1px solid {t['border']};
            border-radius: 4px;
        }}
        
        div[data-testid="stDataFrame"] {{
            border: 1px solid {t['border']};
            border-radius: 4px;
            background-color: {t['card_bg']};
        }}
    </style>
    """,
    unsafe_allow_html=True
)

# IST Timezone Helper via pytz
IST = pytz.timezone('Asia/Kolkata')

def get_ist_now():
    return datetime.datetime.now(IST)

# --- UNIQUE CHECKPOINT: Core Logic & Integrity Verification Guard ---
def verify_core_integrity():
    try:
        conn = sqlite3.connect("sales_history.db")
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        existing_tables = [row[0] for row in cursor.fetchall()]
        conn.close()
        
        required_tables = ['history_logs', 'unique_routes_master', 'output_files_ledger', 'unmapped_missing_dr_ledger']
        for t_name in required_tables:
            if t_name not in existing_tables:
                return False, f"Missing critical database table: {t_name}"
        return True, "All Core Integrity Checkpoints Passed Successfully!"
    except Exception as e:
        return False, str(e)

# SQLite Database Initialization with Master, Output Ledger & Unmapped Missing DR Ledger
def init_db():
    conn = sqlite3.connect("sales_history.db")
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS history_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            files_count INTEGER,
            total_qty REAL,
            status TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS unique_routes_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name TEXT,
            route_no TEXT,
            agency_no TEXT,
            dr_code TEXT,
            created_at TEXT,
            UNIQUE(route_no, agency_no, dr_code)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS output_files_ledger (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name TEXT UNIQUE,
            file_type TEXT,
            file_data BLOB,
            created_at TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS unmapped_missing_dr_ledger (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name TEXT,
            route_no TEXT,
            agency_no TEXT,
            dr_code TEXT,
            created_at TEXT,
            UNIQUE(route_no, agency_no)
        )
    """)
    cursor.execute("PRAGMA table_info(unique_routes_master)")
    columns = [col[1] for col in cursor.fetchall()]
    if "file_name" not in columns:
        cursor.execute("ALTER TABLE unique_routes_master ADD COLUMN file_name TEXT")
        
    conn.commit()
    conn.close()

init_db()

is_healthy, health_msg = verify_core_integrity()
if not is_healthy:
    st.error(f"❌ **System Integrity Warning:** {health_msg}")
    st.stop()

# --- Session State Defaults for Reset/Clear/Restore ---
DEFAULTS = {
    "fg_code": "FG500014",
    "col_map": "36:FG500014AJ\n37:FG500014AK",
    "agency_override": "101:36:FG500014N01\n101:37:FG500014N02",
    "route": "22",
    "email_user": st.secrets.get("email", {}).get("sender_email", ""),
    "email_pass": st.secrets.get("email", {}).get("app_password", ""),
    "recipient": st.secrets.get("email", {}).get("recipient_email", ""),
    "whatsapp": "",
    "processed_files": [],
    "comparison_summary": [],
    "skipped_rows_log": [],
    "anomaly_logs": [],
    "unmapped_current_batch": [],
    "kpi_data": {"input_qty": 0, "gen_qty": 0, "valid_count": 0, "missing_count": 0, "skipped_count": 0}
}

for key, val in DEFAULTS.items():
    if key not in st.session_state:
        st.session_state[key] = val

# --- TOP COLLAPSIBLE ERP CONTROL PANEL ---
with st.expander("⚙️ Enterprise Control Panel & System Settings (Click to Expand)", expanded=False):
    col_set1, col_set2, col_set3 = st.columns(3)
    
    with col_set1:
        st.subheader("Default Fallback FG Code")
        st.session_state.fg_code = st.text_input("FG Code Input", value=st.session_state.fg_code, label_visibility="collapsed")
        c1, c2 = st.columns(2)
        if c1.button("Clear FG"): st.session_state.fg_code = ""; st.rerun()
        if c2.button("Restore FG"): st.session_state.fg_code = DEFAULTS["fg_code"]; st.rerun()
        
        st.subheader("Default Route Fallback")
        st.session_state.route = st.text_input("Route Input", value=st.session_state.route, label_visibility="collapsed")
        c1, c2 = st.columns(2)
        if c1.button("Clear Route"): st.session_state.route = ""; st.rerun()
        if c2.button("Restore Route"): st.session_state.route = DEFAULTS["route"]; st.rerun()

    with col_set2:
        st.subheader("Direct Column Index Mapping")
        st.session_state.col_map = st.text_area("Col Map Input", value=st.session_state.col_map, label_visibility="collapsed", help="ColIndex:Code", height=100)
        c1, c2 = st.columns(2)
        if c1.button("Clear Map"): st.session_state.col_map = ""; st.rerun()
        if c2.button("Restore Map"): st.session_state.col_map = DEFAULTS["col_map"]; st.rerun()

    with col_set3:
        st.subheader("Agency & Column-wise FG Override")
        st.session_state.agency_override = st.text_area("Agency Override Input", value=st.session_state.agency_override, label_visibility="collapsed", help="Agency:ColIndex:CustomFG", height=100)
        c1, c2 = st.columns(2)
        if c1.button("Clear Override"): st.session_state.agency_override = ""; st.rerun()
        if c2.button("Restore Override"): st.session_state.agency_override = DEFAULTS["agency_override"]; st.rerun()

    st.markdown("---")
    col_set4, col_set5 = st.columns(2)
    with col_set4:
        st.subheader("📧 Email Dispatch Settings")
        st.session_state.email_user = st.text_input("Sender Email ID", value=st.session_state.email_user)
        st.session_state.email_pass = st.text_input("Email App Password", type="password", value=st.session_state.email_pass)
        st.session_state.recipient = st.text_input("Recipient Email", value=st.session_state.recipient)
    with col_set5:
        st.subheader("📱 WhatsApp Notification")
        st.session_state.whatsapp = st.text_input("WhatsApp Number (e.g., 919876543210)", value=st.session_state.whatsapp)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Reset All Settings to Defaults"):
            for k, v in DEFAULTS.items():
                st.session_state[k] = v
            st.rerun()

# Mapping assignments
default_fg_code = st.session_state.fg_code
col_mapping_input = st.session_state.col_map
agency_fg_override = st.session_state.agency_override
default_fallback_route = st.session_state.route
email_user = st.session_state.email_user
email_pass = st.session_state.email_pass
recipient_email = st.session_state.recipient
whatsapp_num = st.session_state.whatsapp

direct_col_mapping = {}
for line in col_mapping_input.split('\n'):
    if ':' in line:
        parts = line.split(':')
        idx_str = parts[0].strip()
        if idx_str.isdigit():
            direct_col_mapping[int(idx_str)] = parts[1].strip()

agency_col_override_map = {}
for line in agency_fg_override.split('\n'):
    parts = line.split(':')
    if len(parts) == 3:
        ag, col_idx, fg = parts[0].strip(), parts[1].strip(), parts[2].strip()
        if ag.isdigit() and col_idx.isdigit():
            agency_col_override_map[(int(ag), int(col_idx))] = fg

st.title("💼 Enterprise Sales Order Automation Hub (SAP ERP Edition)")
st.markdown("Upload multiple **Inbound Demand Files** to process orders, auto-lookup missing DRs, log valid unmapped entries, and archive outputs.")
st.markdown("---")

uploaded_inputs = st.file_uploader("Upload Multiple Demand Excel Files", type=["xlsx", "xls"], accept_multiple_files=True, key="inputs")

# --- Pre-flight File Health Check ---
if uploaded_inputs:
    with st.expander("🔍 Pre-flight File Health Check Report", expanded=False):
        preflight_logs = []
        for uploaded_file in uploaded_inputs:
            short_filename = uploaded_file.name
            if short_filename.lower() == "output.xlsx":
                continue
            try:
                df_prev = pd.read_excel(io.BytesIO(uploaded_file.getvalue()), header=None)
                fg_found = any("FG" in str(df_prev.iloc[r, c]).strip().upper() for r in range(df_prev.shape[0]) for c in range(df_prev.shape[1]))
                if fg_found:
                    preflight_logs.append({"File Name": short_filename, "Health Status": "🟢 Healthy", "Details": "FG Header detected successfully"})
                else:
                    preflight_logs.append({"File Name": short_filename, "Health Status": "🔴 Warning", "Details": "'FG' header missing"})
            except Exception as e:
                preflight_logs.append({"File Name": short_filename, "Health Status": "❌ Corrupt", "Details": str(e)})
        if preflight_logs:
            st.dataframe(pd.DataFrame(preflight_logs), use_container_width=True)

st.markdown("<br>", unsafe_allow_html=True)

if st.button("🚀 Process Batch Orders & Update Master DB", type="primary"):
    if uploaded_inputs:
        st.session_state.processed_files = []
        st.session_state.comparison_summary = []
        st.session_state.skipped_rows_log = []
        st.session_state.anomaly_logs = []
        st.session_state.unmapped_current_batch = []
        
        total_input_qty = 0
        total_gen_qty = 0
        total_valid_orders = 0
        total_missing_orders = 0
        total_skipped_rows = 0
        
        db_records_to_insert = []
        unmapped_records_to_insert = []
        output_files_to_store = []
        
        with st.spinner("⚡ Reading files, auto-looking up missing DRs, logging valid unmapped entries... Please wait."):
            try:
                try:
                    with open("Output.xlsx", "rb") as f:
                        template_bytes = f.read()
                except FileNotFoundError:
                    st.error("❌ 'Output.xlsx' template file repository mein nahi mili. Kripya template file ko GitHub repo ke main folder mein upload karein.")
                    st.stop()
                
                ist_now = get_ist_now()
                today_date = ist_now.strftime("%Y-%m-%d")
                timestamp = ist_now.strftime("%H%M%S")

                for uploaded_file in uploaded_inputs:
                    short_filename = uploaded_file.name
                    if short_filename.lower() == "output.xlsx":
                        continue

                    file_bytes = uploaded_file.getvalue()
                    df_input = pd.read_excel(io.BytesIO(file_bytes), header=None)

                    # 1. Find FG Row & Col
                    fg_row, fg_col = -1, -1
                    for r in range(df_input.shape[0]):
                        for c in range(df_input.shape[1]):
                            val = str(df_input.iloc[r, c]).strip().upper()
                            if "FG" in val:
                                fg_row, fg_col = r, c
                                break
                        if fg_row != -1:
                            break

                    if fg_row == -1:
                        st.warning(f"⚠️ '{short_filename}' mein 'FG' header nahi mila. File skip ho rahi hai.")
                        continue

                    # 2. Strict Total/Sum Column Detection
                    total_col = df_input.shape[1]
                    for cSearch in range(fg_col, df_input.shape[1]):
                        is_total = False
                        for scan_r in range(max(0, fg_row - 10), min(fg_row + 3, df_input.shape[0])):
                            cell_val = str(df_input.iloc[scan_r, cSearch]).strip().upper()
                            if any(kw in cell_val for kw in ["TOTAL", "SUM", "TOTA", "TOT", "TTL", "NET"]):
                                is_total = True
                                break
                            if scan_r >= fg_row + 1 and ("SUM" in cell_val or "=" in cell_val):
                                is_total = True
                                break
                        if is_total:
                            total_col = cSearch
                            break

                    # 3. Route Number Finding Logic
                    route_num = default_fallback_route if default_fallback_route != "" else "22"
                    ignore_list = ["RT", "DR", "RT DR", "ROUTE", "SALES PERSON", "CONTACT NO:", "MATERIAL CODE"]
                    
                    for r in range(fg_row):
                        for c in range(min(total_col, 30)):
                            cell_val = str(df_input.iloc[r, c]).strip()
                            upper_val = cell_val.upper()
                            if upper_val in ignore_list:
                                continue
                            is_product_code = any(upper_val.startswith(p) for p in ["PC", "MS", "M", "GM", "DP", "SKU", "FG"])
                            if is_product_code:
                                continue
                            if cell_val != "" and 1 <= len(cell_val) <= 3:
                                if any(char.isdigit() for char in cell_val):
                                    route_num = cell_val
                                    break
                        if route_num != (default_fallback_route if default_fallback_route != "" else "22"):
                            break

                    if default_fallback_route != "" and default_fallback_route != "22":
                        route_num = default_fallback_route

                    safe_route_num = "".join(c if c.isalnum() or c in ('-', '_') else "-" for c in str(route_num))

                    # 4. Smart Agency Detection
                    agency_col = -1
                    for cSearch in range(fg_col - 1, -1, -1):
                        valid_count = 0
                        for rCheck in range(fg_row + 1, min(fg_row + 15, df_input.shape[0])):
                            v = df_input.iloc[rCheck, cSearch]
                            if pd.notna(v):
                                s_val = str(v).replace('.0', '').strip()
                                if s_val.isdigit() and 1 <= len(s_val) <= 5:
                                    valid_count += 1
                        if valid_count >= 3:
                            agency_col = cSearch
                            break

                    if agency_col == -1 and fg_col > 0:
                        agency_col = fg_col - 1

                    # 4.1 Strict DR Code Column Detection
                    dr_code_col = -1
                    for cSearch in range(fg_col - 1, -1, -1):
                        sample_val = str(df_input.iloc[fg_row + 1, cSearch] if fg_row + 1 < df_input.shape[0] else "").strip().upper()
                        if re.match(r'^DR\d+', sample_val):
                            dr_code_col = cSearch
                            break
                        matched_count = 0
                        for offset in range(1, min(4, df_input.shape[0] - fg_row)):
                            v = str(df_input.iloc[fg_row + offset, cSearch]).strip().upper()
                            if re.match(r'^DR\d+', v):
                                matched_count += 1
                        if matched_count > 0:
                            dr_code_col = cSearch
                            break

                    # 5. Pure Valid FG Columns Collection
                    valid_cols = []
                    for c in range(fg_col, total_col):
                        fg_code = str(df_input.iloc[fg_row, c] if fg_row >= 0 else "").strip()
                        if any(kw in fg_code.upper() for kw in ["TOTAL", "SUM", "TOTA", "TOT", "TTL", "NET"]):
                            break
                        valid_cols.append((c, fg_code))

                    wb_valid = openpyxl.load_workbook(io.BytesIO(template_bytes))
                    ws_valid = wb_valid["Order Data"] if "Order Data" in wb_valid.sheetnames else wb_valid.active

                    wb_missing = openpyxl.load_workbook(io.BytesIO(template_bytes))
                    ws_missing = wb_missing["Order Data"] if "Order Data" in wb_missing.sheetnames else wb_missing.active

                    valid_row, missing_row = 6, 6
                    valid_order_num, missing_order_num = 1, 1
                    agency_counts_valid, agency_counts_missing = {}, {}
                    valid_items_created, missing_items_created = 0, 0
                    file_comparison_rows = []

                    for r in range(fg_row + 1, df_input.shape[0]):
                        agency = df_input.iloc[r, agency_col] if agency_col >= 0 else None
                        if pd.isna(agency) or str(agency).strip() in ["", "nan", "None"]:
                            continue
                        
                        agency_str = str(agency).replace('.0','').strip()
                        if not agency_str.isdigit() or not (1 <= len(agency_str) <= 5):
                            st.session_state.skipped_rows_log.append({
                                "File Name": short_filename,
                                "Row Index": r + 1,
                                "Agency Value": str(agency),
                                "Reason": "Invalid or Non-numeric Agency Number"
                            })
                            total_skipped_rows += 1
                            continue

                        agency_val = int(agency_str)
                        
                        # Quantities Check First (Strictly check if quantity > 0)
                        row_has_items = False
                        valid_row_quantities = []
                        row_total_qty = 0
                        for c, fg_code in valid_cols:
                            if c >= total_col:
                                continue
                            sku_qty = df_input.iloc[r, c]
                            if pd.notna(sku_qty) and str(sku_qty).strip() != "":
                                try:
                                    qty_val = float(sku_qty)
                                    if qty_val > 0:
                                        row_has_items = True
                                        row_total_qty += qty_val
                                        valid_row_quantities.append((c, fg_code, qty_val))
                                except ValueError:
                                    pass

                        # Agar quantity 0 ya blank hai, toh row ko completely skip karo
                        if not row_has_items:
                            st.session_state.skipped_rows_log.append({
                                "File Name": short_filename,
                                "Row Index": r + 1,
                                "Agency Value": agency_val,
                                "Reason": "Skipped: Zero or Blank Quantities across all SKUs"
                            })
                            total_skipped_rows += 1
                            continue

                        # DR Code Detection
                        has_dr_code = False
                        clean_dr = ""
                        if dr_code_col >= 0 and dr_code_col < df_input.shape[1]:
                            raw_dr = df_input.iloc[r, dr_code_col]
                            if pd.notna(raw_dr):
                                val_str = str(raw_dr).replace('.0', '').strip()
                                if "DR" in val_str.upper() and val_str.upper() != "0":
                                    has_dr_code = True
                                    clean_dr = val_str

                        if not has_dr_code:
                            for c_scan in range(fg_col):
                                cell_val = df_input.iloc[r, c_scan]
                                if pd.notna(cell_val):
                                    val_str = str(cell_val).replace('.0', '').strip()
                                    if "DR" in val_str.upper() and val_str.upper() != "0":
                                        has_dr_code = True
                                        clean_dr = val_str
                                        break

                        # --- AUTO-LOOKUP MISSING DR FROM DATABASE MASTER ---
                        if not has_dr_code:
                            conn_lookup = sqlite3.connect("sales_history.db")
                            cursor_lookup = conn_lookup.cursor()
                            cursor_lookup.execute("""
                                SELECT dr_code FROM unique_routes_master 
                                WHERE route_no = ? AND agency_no = ? AND dr_code LIKE 'DR%' 
                                LIMIT 1
                            """, (str(route_num), str(agency_val)))
                            db_match = cursor_lookup.fetchone()
                            conn_lookup.close()
                            
                            if db_match:
                                has_dr_code = True
                                clean_dr = db_match[0]

                        # --- STRICT SINGLE UNIQUE ENTRY FOR UNMAPPED MISSING DR ---
                        if not has_dr_code:
                            unmapped_record = (short_filename, str(route_num), str(agency_val), f"NEW_CUST_{agency_val}", ist_now.strftime("%Y-%m-%d %H:%M:%S"))
                            if unmapped_record not in unmapped_records_to_insert:
                                unmapped_records_to_insert.append(unmapped_record)
                            
                            current_unmapped_dict = {
                                "File Name": short_filename,
                                "Route": str(route_num),
                                "Agency": agency_val,
                                "Status": "Generated via NEW_CUST (Missing DR in File and Master DB)"
                            }
                            if current_unmapped_dict not in st.session_state.unmapped_current_batch:
                                st.session_state.unmapped_current_batch.append(current_unmapped_dict)

                        final_dr = clean_dr if has_dr_code else f"NEW_CUST_{agency_val}"
                        
                        # Only insert into Route-Agency-DR Master if it is a genuine valid DR code starting with 'DR'
                        if has_dr_code and clean_dr.startswith("DR"):
                            db_record = (short_filename, str(route_num), str(agency_val), str(clean_dr), ist_now.strftime("%Y-%m-%d %H:%M:%S"))
                            if db_record not in db_records_to_insert:
                                db_records_to_insert.append(db_record)

                        if row_total_qty > 500:
                            st.session_state.anomaly_logs.append({
                                "File Name": short_filename,
                                "Agency": agency_val,
                                "Route": route_num,
                                "Total Qty": row_total_qty,
                                "Flag": "⚠️ High Volume Spike (>500)"
                            })

                        if has_dr_code:
                            agency_counts_valid[agency_val] = agency_counts_valid.get(agency_val, 0) + 1
                            current_seq = agency_counts_valid[agency_val]
                            ref_number = f"RT-{route_num}-{agency_val}-{today_date}" if current_seq == 1 else f"RT-{route_num}-{agency_val}-{today_date}-{current_seq}"
                            target_ws, current_r, order_num, dr_to_use, file_category = ws_valid, valid_row, valid_order_num, clean_dr, "Valid DR"
                        else:
                            agency_counts_missing[agency_val] = agency_counts_missing.get(agency_val, 0) + 1
                            current_seq = agency_counts_missing[agency_val]
                            ref_number = f"RT-{route_num}-{agency_val}-{today_date}-NEW" if current_seq == 1 else f"RT-{route_num}-{agency_val}-{today_date}-NEW-{current_seq}"
                            target_ws, current_r, order_num, dr_to_use, file_category = ws_missing, missing_row, missing_order_num, f"NEW_CUST_{agency_val}", "Missing DR"

                        item_id = 10
                        for c, fg_code, qty_val in valid_row_quantities:
                            cleaned_fg = str(fg_code).strip()
                            upper_fg = cleaned_fg.upper()
                            
                            if (agency_val, c) in agency_col_override_map:
                                current_fg = agency_col_override_map[(agency_val, c)]
                            elif upper_fg.startswith("FG"):
                                current_fg = cleaned_fg
                            elif upper_fg in ["", "NAN", "NONE"]:
                                current_fg = direct_col_mapping.get(c, default_fg_code)
                            else:
                                current_fg = direct_col_mapping.get(c, default_fg_code)
                            
                            total_input_qty += qty_val
                            total_gen_qty += qty_val
                            
                            file_comparison_rows.append({
                                "File Name": short_filename,
                                "Status": file_category,
                                "Agency": agency_val,
                                "DR Code": dr_to_use,
                                "FG Code": current_fg,
                                "Input Qty": qty_val,
                                "Generated Qty": qty_val
                            })
                            
                            target_ws.cell(row=current_r, column=2, value=order_num)
                            target_ws.cell(row=current_r, column=3, value="OR")
                            target_ws.cell(row=current_r, column=4, value="SO20")
                            target_ws.cell(row=current_r, column=5, value=10)
                            target_ws.cell(row=current_r, column=6, value=20)
                            target_ws.cell(row=current_r, column=7, value=dr_to_use)
                            target_ws.cell(row=current_r, column=8, value=dr_to_use)
                            target_ws.cell(row=current_r, column=9, value=ref_number)
                            target_ws.cell(row=current_r, column=10, value=today_date)
                            target_ws.cell(row=current_r, column=11, value=today_date)
                            target_ws.cell(row=current_r, column=15, value=item_id)
                            target_ws.cell(row=current_r, column=16, value=current_fg)
                            target_ws.cell(row=current_r, column=19, value=qty_val)
                            target_ws.cell(row=current_r, column=20, value="Bag")
                            target_ws.cell(row=current_r, column=22, value=2100)
                            target_ws.cell(row=current_r, column=26, value=str(route_num))
                            target_ws.cell(row=current_r, column=27, value=agency_val)
                            
                            item_id += 10
                            current_r += 1

                        if has_dr_code:
                            valid_row, valid_order_num, valid_items_created, total_valid_orders = current_r, valid_order_num + 1, valid_items_created + 1, total_valid_orders + 1
                        else:
                            missing_row, missing_order_num, missing_items_created, total_missing_orders = current_r, missing_order_num + 1, missing_items_created + 1, total_missing_orders + 1

                    if valid_items_created > 0:
                        buf_valid = io.BytesIO()
                        wb_valid.save(buf_valid)
                        buf_valid.seek(0)
                        out_fname = safe_route_num + "_" + today_date + "_" + timestamp + "_Valid.xlsx"
                        st.session_state.processed_files.append({
                            "name": short_filename + " (Valid DR)",
                            "data": buf_valid.getvalue(),
                            "filename": out_fname,
                            "orders": valid_items_created
                        })
                        output_files_to_store.append((out_fname, "Valid DR", buf_valid.getvalue(), ist_now.strftime("%Y-%m-%d %H:%M:%S")))

                    if missing_items_created > 0:
                        buf_missing = io.BytesIO()
                        wb_missing.save(buf_missing)
                        buf_missing.seek(0)
                        out_fname_miss = safe_route_num + "_" + today_date + "_" + timestamp + "_Missing_DR.xlsx"
                        st.session_state.processed_files.append({
                            "name": short_filename + " (Missing DR / New)",
                            "data": buf_missing.getvalue(),
                            "filename": out_fname_miss,
                            "orders": missing_items_created
                        })
                        output_files_to_store.append((out_fname_miss, "Missing DR", buf_missing.getvalue(), ist_now.strftime("%Y-%m-%d %H:%M:%S")))

                    if file_comparison_rows:
                        df_comp = pd.DataFrame(file_comparison_rows)
                        df_pivot = df_comp.pivot_table(
                            index=["File Name", "Status", "Agency", "DR Code", "FG Code"],
                            values=["Input Qty", "Generated Qty"],
                            aggfunc="sum"
                        ).reset_index()
                        df_pivot["Difference"] = df_pivot["Input Qty"] - df_pivot["Generated Qty"]
                        st.session_state.comparison_summary.append(df_pivot)

                # --- Update Master DB, Unmapped Ledger & Output Files Ledger ---
                conn = sqlite3.connect("sales_history.db")
                cursor = conn.cursor()
                cursor.executemany("""
                    INSERT OR IGNORE INTO unique_routes_master (file_name, route_no, agency_no, dr_code, created_at)
                    VALUES (?, ?, ?, ?, ?)
                """, db_records_to_insert)

                cursor.executemany("""
                    INSERT OR IGNORE INTO unmapped_missing_dr_ledger (file_name, route_no, agency_no, dr_code, created_at)
                    VALUES (?, ?, ?, ?, ?)
                """, unmapped_records_to_insert)
                
                for fname, ftype, fdata, fdate in output_files_to_store:
                    cursor.execute("""
                        INSERT OR IGNORE INTO output_files_ledger (file_name, file_type, file_data, created_at)
                        VALUES (?, ?, ?, ?)
                    """, (fname, ftype, fdata, fdate))
                
                cursor.execute(
                    "INSERT INTO history_logs (timestamp, files_count, total_qty, status) VALUES (?, ?, ?, ?)",
                    (get_ist_now().strftime("%Y-%m-%d %H:%M:%S"), len(uploaded_inputs), total_input_qty, "Success")
                )
                conn.commit()
                conn.close()

                st.session_state.kpi_data = {
                    "input_qty": total_input_qty,
                    "gen_qty": total_gen_qty,
                    "valid_count": total_valid_orders,
                    "missing_count": total_missing_orders,
                    "skipped_count": total_skipped_rows
                }

                st.success("✅ Batch Processing, Auto-Lookup, Unmapped Ledger & Output Archive Updated Successfully!")

            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
    else:
        st.warning("⚠️ Kripya pehle demand files upload karein!")

# Display KPI Summary Cards & Advanced Visual Analytics
if st.session_state.processed_files or st.session_state.skipped_rows_log:
    st.markdown("---")
    st.markdown("### 📈 Batch Performance & KPI Summary")
    kpi = st.session_state.kpi_data
    
    total_processed_orders = kpi['valid_count'] + kpi['missing_count']
    success_rate = (kpi['valid_count'] / total_processed_orders * 100) if total_processed_orders > 0 else 0
    
    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Total Input Qty", f"{kpi['input_qty']:,.0f}")
    col2.metric("Generated Qty", f"{kpi['gen_qty']:,.0f}")
    col3.metric("Valid Orders", kpi['valid_count'])
    col4.metric("Success Rate", f"{success_rate:.1f}%")
    col5.metric("Skipped Rows", kpi['skipped_count'], delta_color="inverse")

    if kpi['skipped_count'] > 5:
        st.warning(f"⚠️ **Smart Audit Alert:** {kpi['skipped_count']} rows skipped check exception logs.")

    # --- AI Sales Demand Forecasting & Velocity Health Scorecard ---
    st.markdown("---")
    st.markdown("### 🤖 AI Sales Demand Forecasting & Velocity Health Scorecard")
    
    demand_health_score = success_rate
    forecast_confidence = "🟢 High Confidence (Stable Batch Flow)" if demand_health_score >= 90 else ("🟡 Moderate Risk (Unmapped Fallbacks Detected)" if demand_health_score >= 70 else "🔴 Critical Review Needed (High Missing DR Ratio)")
    
    f_col1, f_col2, f_col3 = st.columns(3)
    f_col1.metric("Batch Demand Health Score", f"{demand_health_score:.1f}%", delta="Optimal Flow" if demand_health_score >= 90 else "Attention Needed")
    f_col2.metric("Forecast Confidence Status", forecast_confidence)
    f_col3.metric("Projected Next-Cycle Qty", f"{kpi['gen_qty'] * 1.05:,.0f} Units", delta="+5% Trend")

    # --- AI Anomaly & Unmapped Missing DR Alerts ---
    if st.session_state.anomaly_logs:
        st.markdown("---")
        st.markdown("### 🤖 AI Demand Spike & Anomaly Detector Alerts")
        st.warning("⚠️ System detected high-volume demand spikes (>500 units) in the following agencies:")
        st.dataframe(pd.DataFrame(st.session_state.anomaly_logs), use_container_width=True)

    if st.session_state.unmapped_current_batch:
        st.markdown("---")
        st.markdown("### 🚨 Unmapped Missing DRs Alert List")
        st.error("⚠️ The following agencies had valid quantity (>0) but no DR code in file or Master DB. They were successfully processed using `NEW_CUST` fallback and logged into the Unmapped Ledger:")
        st.dataframe(pd.DataFrame(st.session_state.unmapped_current_batch), use_container_width=True)

    # --- ADVANCED TABBED VISUAL ANALYTICS ---
    if st.session_state.comparison_summary:
        st.markdown("---")
        st.markdown("### 📊 Advanced Visual Analytics Dashboard")
        combined_df_chart = pd.concat(st.session_state.comparison_summary, ignore_index=True)
        
        tab1, tab2 = st.tabs(["📊 Agency-wise Breakdown", "📦 SKU-wise Share"])
        with tab1:
            st.bar_chart(combined_df_chart.groupby("Agency")["Generated Qty"].sum())
        with tab2:
            st.bar_chart(combined_df_chart.groupby("FG Code")["Generated Qty"].sum())

    st.markdown("---")
    st.markdown("### 📥 Bulk Download & Notifications")
    
    # --- Advanced Email Config Expander ---
    with st.expander("✉️ Advanced Email Dispatch Options (Custom Subject & Note)"):
        email_subject_custom = st.text_input("Custom Email Subject Line", f"🚀 Sales Orders Batch Execution Report (IST) - {get_ist_now().strftime('%Y-%m-%d')}")
        email_notes_custom = st.text_area("Custom Remarks / Notes to Include in Email Body", "All routes verified and processed successfully.")

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for item in st.session_state.processed_files:
            zip_file.writestr(item['filename'], item['data'])
    
    col_zip, col_pdf, col_summary, col_json, col_print, col_email, col_wa = st.columns(7)
    
    with col_zip:
        st.download_button(
            label="📦 ZIP",
            data=zip_buffer.getvalue(),
            file_name=f"Batch_Orders_{get_ist_now().strftime('%Y-%m-%d')}.zip",
            mime="application/zip",
            key="zip_download"
        )
        
    with col_pdf:
        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", "B", 16)
            pdf.cell(190, 10, "Enterprise Sales Order Summary Invoice", ln=True, align="C")
            pdf.set_font("Arial", "", 10)
            pdf.cell(190, 6, f"Generated On (IST): {get_ist_now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align="C")
            pdf.ln(10)
            
            pdf.set_font("Arial", "B", 11)
            pdf.cell(100, 8, "Metric Description", border=1)
            pdf.cell(90, 8, "Value", border=1, ln=True)
            
            pdf.set_font("Arial", "", 11)
            metrics_list = [
                ("Total Input Quantity", f"{kpi['input_qty']:,.0f}"),
                ("Total Generated Quantity", f"{kpi['gen_qty']:,.0f}"),
                ("Valid DR Orders", str(kpi['valid_count'])),
                ("Missing DR Orders (NEW_CUST)", str(kpi['missing_count'])),
                ("Skipped Rows Logged", str(kpi['skipped_count'])),
                ("Success Rate", f"{success_rate:.1f}%"),
                ("Demand Health Score", f"{demand_health_score:.1f}%")
            ]
            for m_desc, m_val in metrics_list:
                pdf.cell(100, 8, m_desc, border=1)
                pdf.cell(90, 8, m_val, border=1, ln=True)
                
            pdf_bytes = bytes(pdf.output())
            st.download_button(
                label="📄 PDF",
                data=pdf_bytes,
                file_name=f"Sales_Invoice_{get_ist_now().strftime('%Y-%m-%d')}.pdf",
                mime="application/pdf",
                key="pdf_invoice_download"
            )
        except Exception as e:
            st.error(f"PDF Error: {str(e)}")

    with col_summary:
        summary_txt = f"""=== ENTERPRISE SALES ORDER SUMMARY (IST) ===
Date/Time: {get_ist_now().strftime('%Y-%m-%d %H:%M:%S')}
----------------------------------------
Total Input Quantity : {kpi['input_qty']:,.0f}
Total Generated Qty  : {kpi['gen_qty']:,.0f}
Valid DR Orders      : {kpi['valid_count']}
Missing DR Orders    : {kpi['missing_count']}
Skipped Rows Logged  : {kpi['skipped_count']}
Success Rate         : {success_rate:.1f}%
Demand Health Score  : {demand_health_score:.1f}%
----------------------------------------
Generated Files Count: {len(st.session_state.processed_files)}
Status: Successfully Processed & Audited
========================================"""
        st.download_button(
            label="📄 TXT",
            data=summary_txt.encode('utf-8'),
            file_name=f"Summary_Report_{get_ist_now().strftime('%Y-%m-%d')}.txt",
            mime="text/plain",
            key="summary_txt_download"
        )
        
    with col_json:
        json_data = json.dumps({
            "timestamp": get_ist_now().strftime('%Y-%m-%d %H:%M:%S'),
            "metrics": kpi,
            "success_rate": f"{success_rate:.1f}%",
            "demand_health_score": f"{demand_health_score:.1f}%"
        }, indent=4)
        st.download_button(
            label="💾 JSON",
            data=json_data.encode('utf-8'),
            file_name=f"Audit_Backup_{get_ist_now().strftime('%Y-%m-%d')}.json",
            mime="application/json",
            key="json_backup_download"
        )
        
    with col_print:
        print_html = """
        <div style="width:100%; margin:0; padding:0;">
            <button onclick="parent.window.print()" style="width:100%; height:38px; background:#2563eb; color:white; border:none; border-radius:4px; font-weight:600; cursor:pointer; font-family:sans-serif; display:flex; align-items:center; justify-content:center;">
                🖨️ Print
            </button>
        </div>
        """
        components.html(print_html, height=50)
        
    with col_email:
        if st.button("📧 Email"):
            if email_user and email_pass and recipient_email:
                try:
                    conn = sqlite3.connect("sales_history.db")
                    df_master_email = pd.read_sql("SELECT * FROM unique_routes_master", conn)
                    conn.close()
                    
                    excel_buffer = io.BytesIO()
                    df_master_email.to_excel(excel_buffer, index=False, sheet_name="Master Routes")
                    excel_buffer.seek(0)

                    unmapped_email_html = ""
                    if st.session_state.unmapped_current_batch:
                        unmapped_email_html = "<h3 style='color: #dc2626;'>🚨 Unmapped Missing DRs Alert List</h3><table style='border-collapse: collapse; width: 100%; margin-top: 10px; border-radius: 6px; overflow: hidden;'><tr style='background-color: #ef4444; color: white;'><th style='padding: 8px; text-align: left;'>File Name</th><th style='padding: 8px; text-align: left;'>Route</th><th style='padding: 8px; text-align: left;'>Agency</th><th style='padding: 8px; text-align: left;'>Status</th></tr>"
                        for item in st.session_state.unmapped_current_batch:
                            unmapped_email_html += f"<tr style='background-color: #fef2f2;'><td style='padding: 8px; border-bottom: 1px solid #fee2e2;'>{item['File Name']}</td><td style='padding: 8px; border-bottom: 1px solid #fee2e2;'>{item['Route']}</td><td style='padding: 8px; border-bottom: 1px solid #fee2e2;'>{item['Agency']}</td><td style='padding: 8px; border-bottom: 1px solid #fee2e2; color: #dc2626;'>{item['Status']}</td></tr>"
                        unmapped_email_html += "</table>"

                    msg = EmailMessage()
                    msg['Subject'] = email_subject_custom
                    msg['From'] = email_user
                    msg['To'] = recipient_email
                    
                    html_content = f"""
                    <html>
                      <body style="font-family: Arial, sans-serif; color: #333; background-color: #f9fafb; padding: 20px;">
                        <div style="max-width: 600px; background: #ffffff; padding: 25px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.05);">
                          <h2 style="color: #10b981; border-bottom: 2px solid #e5e7eb; padding-bottom: 10px;">📊 Sales Order Batch Automation Hub</h2>
                          <p>Hello Team,</p>
                          <p>The daily inbound demand batch has been processed successfully on <b>{get_ist_now().strftime('%Y-%m-%d %H:%M:%S')} IST</b>.</p>
                          <p style="background: #f0fdf4; border-left: 4px solid #10b981; padding: 10px; margin: 15px 0;"><b>Remarks:</b> {email_notes_custom}</p>
                          <table style="border-collapse: collapse; width: 100%; margin-top: 15px; border-radius: 6px; overflow: hidden;">
                            <tr style="background-color: #10b981; color: white;">
                              <th style="padding: 10px; text-align: left;">Metric</th>
                              <th style="padding: 10px; text-align: left;">Value</th>
                            </tr>
                            <tr style="background-color: #f3f4f6;">
                              <td style="padding: 10px; border-bottom: 1px solid #e5e7eb;">Total Input Qty</td>
                              <td style="padding: 10px; border-bottom: 1px solid #e5e7eb;"><b>{kpi['input_qty']:,.0f}</b></td>
                            </tr>
                            <tr>
                              <td style="padding: 10px; border-bottom: 1px solid #e5e7eb;">Valid Orders</td>
                              <td style="padding: 10px; border-bottom: 1px solid #e5e7eb;">{kpi['valid_count']}</td>
                            </tr>
                            <tr style="background-color: #f3f4f6;">
                              <td style="padding: 10px; border-bottom: 1px solid #e5e7eb;">Missing DR Orders (NEW_CUST)</td>
                              <td style="padding: 10px; border-bottom: 1px solid #e5e7eb;"><b>{kpi['missing_count']}</b></td>
                            </tr>
                            <tr>
                              <td style="padding: 10px; border-bottom: 1px solid #e5e7eb;">Demand Health Score</td>
                              <td style="padding: 10px; border-bottom: 1px solid #e5e7eb;"><b>{demand_health_score:.1f}%</b></td>
                            </tr>
                          </table>
                          {unmapped_email_html}
                          <p style="margin-top: 25px; color: #666; font-size: 12px; border-top: 1px solid #e5e7eb; paddingTop: 10px;">Master Route-Agency-DR Database attached herewith.</p>
                        </div>
                      </body>
                    </html>
                    """
                    msg.set_content("Please enable HTML to view this report.")
                    msg.add_alternative(html_content, subtype='html')
                    
                    for item in st.session_state.processed_files:
                        msg.add_attachment(item['data'], maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=item['filename'])
                    
                    msg.add_attachment(excel_buffer.getvalue(), maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=f"Unique_Routes_Master_{get_ist_now().strftime('%Y-%m-%d')}.xlsx")
                    
                    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                        smtp.login(email_user, email_pass)
                        smtp.send_message(msg)
                    st.success("✅ Email dispatched with Master DB attached!")
                except Exception as e:
                    st.error(f"❌ Email failed: {str(e)}")
            else:
                st.warning("⚠️ Enter email credentials!")

    with col_wa:
        if whatsapp_num:
            wa_text = f"Sales Order Batch Ready! Total Qty: {kpi['input_qty']}, Health Score: {demand_health_score:.1f}%."
            wa_link = f"https://wa.me/{whatsapp_num}?text={urllib.parse.quote(wa_text)}"
            st.markdown(f'<a href="{wa_link}" target="_blank" style="text-decoration:none;"><button style="width:100%; height:38px; background:#25D366; color:white; border:none; border-radius:4px; font-weight:600; cursor:pointer; display:flex; align-items:center; justify-content:center;">📱 WhatsApp</button></a>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("##### Individual File Downloads:")
    for i, item in enumerate(st.session_state.processed_files):
        if st.download_button(
            label=f"📥 Download {item['name']}",
            data=item['data'],
            file_name=item['filename'],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_file_{i}_{item['filename']}"
        ):
            st.toast(f"🎉 '{item['filename']}' downloaded!", icon="📥")

# --- ALL THREE DATABASES MANAGEMENT PANEL (DIRECT DELETION) ---
st.markdown("---")
with st.expander("🗄️ View, Export & Manage All Databases (Master, Unmapped, Outputs) & Upload DR Codes"):
    st.markdown("Yahan aap teeno databases ke records dekh sakte hain, manual/bulk DR code upload kar sakte hain, aur record delete karne par ID auto-reset kar sakte hain.")
    try:
        conn = sqlite3.connect("sales_history.db")
        df_master = pd.read_sql("SELECT * FROM unique_routes_master ORDER BY id DESC", conn)
        df_unmapped = pd.read_sql("SELECT * FROM unmapped_missing_dr_ledger ORDER BY id DESC", conn)
        df_outputs = pd.read_sql("SELECT id, file_name, file_type, created_at FROM output_files_ledger ORDER BY id DESC", conn)
        conn.close()
        
        tab_m1, tab_m2, tab_m3 = st.tabs(["📋 Route-Agency-DR Master", "🚨 Unmapped Missing DR Ledger", "📦 Archived Output Files"])
        
        # --- TAB 1: MASTER DATABASE MANAGEMENT & MANUAL/BULK DR UPLOAD ---
        with tab_m1:
            if not df_master.empty:
                st.markdown("#### 📊 Master Database Health & Analytics")
                h_col1, h_col2, h_col3, h_col4 = st.columns(4)
                h_col1.metric("Total Master Records", len(df_master))
                h_col2.metric("Unique Routes", df_master['route_no'].nunique() if 'route_no' in df_master.columns else 0)
                h_col3.metric("Unique Agencies", df_master['agency_no'].nunique() if 'agency_no' in df_master.columns else 0)
                h_col4.metric("Tracked Files", df_master['file_name'].nunique() if 'file_name' in df_master.columns else 0)
                
                st.markdown("---")

                # Manual & Bulk DR Code Upload Feature
                with st.expander("📤 Manual Entry / Bulk Upload DR Codes into Master Database"):
                    col_man1, col_man2, col_man3 = st.columns(3)
                    with col_man1:
                        manual_route = st.text_input("Route No", "10", key="man_route")
                    with col_man2:
                        manual_agency = st.text_input("Agency No", "", key="man_agency")
                    with col_man3:
                        manual_dr = st.text_input("DR Code (e.g., DR12345)", "", key="man_dr")
                    
                    if st.button("➕ Add Single DR Code to Master DB"):
                        if manual_route and manual_agency and manual_dr:
                            try:
                                conn_add = sqlite3.connect("sales_history.db")
                                cur_add = conn_add.cursor()
                                cur_add.execute("""
                                    INSERT OR REPLACE INTO unique_routes_master (file_name, route_no, agency_no, dr_code, created_at)
                                    VALUES (?, ?, ?, ?, ?)
                                """, ("Manual_Entry", manual_route, manual_agency, manual_dr, get_ist_now().strftime("%Y-%m-%d %H:%M:%S")))
                                conn_add.commit()
                                conn_add.close()
                                st.success(f"✅ Successfully added Route: {manual_route}, Agency: {manual_agency}, DR: {manual_dr}!")
                                st.rerun()
                            except Exception as ex:
                                st.error(f"Error adding record: {str(ex)}")
                        else:
                            st.warning("⚠️ Kripya Route, Agency aur DR Code sabhi enter karein.")

                    st.markdown("---")
                    st.markdown("##### Or Bulk Upload CSV / Excel")
                    bulk_upload_file = st.file_uploader("Upload Master DR CSV/Excel", type=["csv", "xlsx"], key="bulk_dr_up")
                    if bulk_upload_file:
                        try:
                            if bulk_upload_file.name.endswith('.csv'):
                                df_bulk = pd.read_csv(bulk_upload_file)
                            else:
                                df_bulk = pd.read_excel(bulk_upload_file)
                            
                            if all(col in df_bulk.columns for col in ['route_no', 'agency_no', 'dr_code']):
                                bulk_records = []
                                for _, row in df_bulk.iterrows():
                                    bulk_records.append((str(row.get('file_name', 'Manual_Upload')), str(row['route_no']), str(row['agency_no']), str(row['dr_code']), get_ist_now().strftime("%Y-%m-%d %H:%M:%S")))
                                
                                conn_b = sqlite3.connect("sales_history.db")
                                cur_b = conn_b.cursor()
                                cur_b.executemany("""
                                    INSERT OR REPLACE INTO unique_routes_master (file_name, route_no, agency_no, dr_code, created_at)
                                    VALUES (?, ?, ?, ?, ?)
                                """, bulk_records)
                                conn_b.commit()
                                conn_b.close()
                                st.success(f"✅ Successfully imported {len(bulk_records)} records into Master Database!")
                                st.rerun()
                            else:
                                st.error("❌ File columns must contain: 'route_no', 'agency_no', 'dr_code'.")
                        except Exception as ex:
                            st.error(f"Error importing file: {str(ex)}")

                db_search = st.text_input("🔍 Search Master Database (Filter by File, Route, Agency or DR)", "", key="db_search")
                filtered_master = df_master
                if db_search:
                    q = db_search.lower()
                    filtered_master = df_master[
                        df_master['file_name'].astype(str).str.lower().str.contains(q) |
                        df_master['route_no'].astype(str).str.lower().str.contains(q) |
                        df_master['agency_no'].astype(str).str.lower().str.contains(q) |
                        df_master['dr_code'].astype(str).str.lower().str.contains(q)
                    ]
                
                st.dataframe(filtered_master, use_container_width=True)
                
                st.markdown("#### 🗑️ Advanced Deletion, Rollback & ID Reset Tools")
                del_col1, del_col2, del_col3, del_col4 = st.columns(4)
                
                with del_col1:
                    row_id_to_del = st.number_input("Enter Master Record ID", min_value=1, step=1, key="row_id_input")
                    if st.button("🗑️ Delete Master Row & Reset ID"):
                        conn = sqlite3.connect("sales_history.db")
                        cursor = conn.cursor()
                        cursor.execute("DELETE FROM unique_routes_master WHERE id = ?", (row_id_to_del,))
                        cursor.execute("DELETE FROM sqlite_sequence WHERE name='unique_routes_master'")
                        conn.commit()
                        conn.close()
                        st.success(f"✅ Master Record ID {row_id_to_del} deleted & ID sequence reset!")
                        st.rerun()

                with del_col2:
                    unique_files = df_master['file_name'].dropna().unique().tolist() if 'file_name' in df_master.columns else []
                    file_to_purge = st.selectbox("Select File to Purge", ["Select..."] + unique_files, key="purge_file_select")
                    if st.button("🔥 Delete Master File Data"):
                        if file_to_purge != "Select...":
                            conn = sqlite3.connect("sales_history.db")
                            cursor = conn.cursor()
                            cursor.execute("DELETE FROM unique_routes_master WHERE file_name = ?", (file_to_purge,))
                            cursor.execute("DELETE FROM sqlite_sequence WHERE name='unique_routes_master'")
                            conn.commit()
                            conn.close()
                            st.success(f"✅ File '{file_to_purge}' data deleted & ID sequence reset!")
                            st.rerun()
                        else:
                            st.warning("⚠️ Kripya purge karne ke liye file select karein.")

                with del_col3:
                    route_to_delete = st.text_input("Enter Route No to Purge", "", key="purge_route")
                    if st.button("🔥 Delete Master Route Data"):
                        if route_to_delete:
                            conn = sqlite3.connect("sales_history.db")
                            cursor = conn.cursor()
                            cursor.execute("DELETE FROM unique_routes_master WHERE route_no = ?", (route_to_delete,))
                            cursor.execute("DELETE FROM sqlite_sequence WHERE name='unique_routes_master'")
                            conn.commit()
                            conn.close()
                            st.success(f"✅ Route '{route_to_delete}' data deleted & ID sequence reset!")
                            st.rerun()
                        else:
                            st.warning("⚠️ Kripya delete karne ke liye Route No enter karein.")
                            
                with del_col4:
                    st.markdown("##### Master Wipe")
                    if st.button("🚨 Wipe Master DB & Reset IDs", type="secondary"):
                        conn = sqlite3.connect("sales_history.db")
                        cursor = conn.cursor()
                        cursor.execute("DELETE FROM unique_routes_master")
                        cursor.execute("DELETE FROM sqlite_sequence WHERE name='unique_routes_master'")
                        conn.commit()
                        conn.close()
                        st.success("✅ Master DB wiped & IDs reset!")
                        st.rerun()

                st.markdown("---")
                master_excel_buf = io.BytesIO()
                df_master.to_excel(master_excel_buf, index=False, sheet_name="Master Routes")
                master_excel_buf.seek(0)
                
                dl_col1, dl_col2 = st.columns(2)
                with dl_col1:
                    st.download_button(
                        label="📥 Export Master Database to Excel (.xlsx)",
                        data=master_excel_buf.getvalue(),
                        file_name=f"Unique_Routes_Master_{get_ist_now().strftime('%Y-%m-%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="export_master_excel"
                    )
                with dl_col2:
                    if st.button("📧 Send Master DB via Email"):
                        if email_user and email_pass and recipient_email:
                            try:
                                msg = EmailMessage()
                                msg['Subject'] = f"📊 Master Database Export Report (IST) - {get_ist_now().strftime('%Y-%m-%d')}"
                                msg['From'] = email_user
                                msg['To'] = recipient_email
                                
                                msg.set_content("Hello Team,\n\nPlease find attached the latest Unique Route-Agency-DR Master Database export.\n\nAutomated via Sales Order Hub (IST)")
                                msg.add_attachment(master_excel_buf.getvalue(), maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=f"Unique_Routes_Master_{get_ist_now().strftime('%Y-%m-%d')}.xlsx")
                                
                                with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                                    smtp.login(email_user, email_pass)
                                    smtp.send_message(msg)
                                st.success("✅ Master Database successfully emailed to recipient!")
                            except Exception as e:
                                st.error(f"❌ Email failed: {str(e)}")
                        else:
                            st.warning("⚠️ Kripya sidebar mein Email credentials enter karein!")
            else:
                st.info("No master records found yet.")

        # --- TAB 2: UNMAPPED LEDGER MANAGEMENT (DIRECT DELETION) ---
        with tab_m2:
            st.markdown("#### 🚨 Unmapped Missing DR Ledger (Generated via Fallback)")
            if not df_unmapped.empty:
                st.dataframe(df_unmapped, use_container_width=True)
                
                st.markdown("##### 🗑️ Unmapped Ledger Deletion & ID Reset Tools")
                um_col1, um_col2 = st.columns(2)
                with um_col1:
                    unmap_del_id = st.number_input("Enter Unmapped Record ID", min_value=1, step=1, key="unmap_del_id")
                    if st.button("🗑️ Delete Unmapped Record & Reset ID"):
                        conn = sqlite3.connect("sales_history.db")
                        cursor = conn.cursor()
                        cursor.execute("DELETE FROM unmapped_missing_dr_ledger WHERE id = ?", (unmap_del_id,))
                        cursor.execute("DELETE FROM sqlite_sequence WHERE name='unmapped_missing_dr_ledger'")
                        conn.commit()
                        conn.close()
                        st.success(f"✅ Unmapped Record ID {unmap_del_id} deleted & ID sequence reset!")
                        st.rerun()
                with um_col2:
                    st.markdown("##### Wipe Unmapped Ledger")
                    if st.button("🚨 Wipe Unmapped Ledger & Reset IDs"):
                        conn = sqlite3.connect("sales_history.db")
                        cursor = conn.cursor()
                        cursor.execute("DELETE FROM unmapped_missing_dr_ledger")
                        cursor.execute("DELETE FROM sqlite_sequence WHERE name='unmapped_missing_dr_ledger'")
                        conn.commit()
                        conn.close()
                        st.success("✅ Unmapped Ledger wiped & IDs reset!")
                        st.rerun()

                unmapped_buf = io.BytesIO()
                df_unmapped.to_excel(unmapped_buf, index=False, sheet_name="Unmapped DR Ledger")
                unmapped_buf.seek(0)
                st.download_button(
                    label="📥 Export Unmapped Ledger to Excel (.xlsx)",
                    data=unmapped_buf.getvalue(),
                    file_name=f"Unmapped_Missing_DR_Ledger_{get_ist_now().strftime('%Y-%m-%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="export_unmapped_excel"
                )
            else:
                st.info("No unmapped missing DR records logged yet.")

        # --- TAB 3: ARCHIVED OUTPUT FILES MANAGEMENT (DIRECT DELETION) ---
        with tab_m3:
            st.markdown("#### 📦 Archived Output Files (Saved per file without duplication)")
            if not df_outputs.empty:
                st.dataframe(df_outputs, use_container_width=True)
                
                st.markdown("##### 🗑️ Output Ledger Deletion & ID Reset Tools")
                out_col1, out_col2 = st.columns(2)
                with out_col1:
                    selected_output_id = st.number_input("Enter Archived File ID", min_value=1, step=1, key="out_file_id")
                    if st.button("📥 Download Archived File"):
                        conn = sqlite3.connect("sales_history.db")
                        cursor = conn.cursor()
                        cursor.execute("SELECT file_name, file_data FROM output_files_ledger WHERE id = ?", (selected_output_id,))
                        row_res = cursor.fetchone()
                        conn.close()
                        if row_res:
                            fname_res, fdata_res = row_res[0], row_res[1]
                            st.download_button(
                                label=f"💾 Click to save '{fname_res}'",
                                data=fdata_res,
                                file_name=fname_res,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"dl_archived_{selected_output_id}"
                            )
                        else:
                            st.warning("⚠️ Invalid ID or File not found.")
                    
                    delete_arch_id = st.number_input("Enter Archived File ID to Delete", min_value=1, step=1, key="del_arch_id_input")
                    if st.button("🗑️ Delete Archived File & Reset ID"):
                        conn = sqlite3.connect("sales_history.db")
                        cursor = conn.cursor()
                        cursor.execute("DELETE FROM output_files_ledger WHERE id = ?", (delete_arch_id,))
                        cursor.execute("DELETE FROM sqlite_sequence WHERE name='output_files_ledger'")
                        conn.commit()
                        conn.close()
                        st.success(f"✅ Archived File ID {delete_arch_id} deleted & ID sequence reset!")
                        st.rerun()

                with out_col2:
                    st.markdown("##### Wipe Output Ledger")
                    if st.button("🚨 Wipe Output Ledger & Reset IDs"):
                        conn = sqlite3.connect("sales_history.db")
                        cursor = conn.cursor()
                        cursor.execute("DELETE FROM output_files_ledger")
                        cursor.execute("DELETE FROM sqlite_sequence WHERE name='output_files_ledger'")
                        conn.commit()
                        conn.close()
                        st.success("✅ Output Ledger wiped & IDs reset!")
                        st.rerun()
            else:
                st.info("No output files archived yet.")

    except Exception as e:
        st.error(f"Error loading database: {str(e)}")

# Tables and Logs with Search & Filter Feature
if st.session_state.comparison_summary:
    st.markdown("---")
    st.markdown("### 📋 Agency-wise Material & Input Comparison")
    
    search_query = st.text_input("🔍 Search Table (Filter by Agency, DR Code, or FG Code)", "", key="table_search")
    
    combined_df = pd.concat(st.session_state.comparison_summary, ignore_index=True)
    summary_table = combined_df.groupby(["Agency", "DR Code", "FG Code"], as_index=False).agg({"Input Qty": "sum", "Generated Qty": "sum"})
    summary_table["Difference"] = summary_table["Input Qty"] - summary_table["Generated Qty"]
    
    if search_query:
        q = search_query.lower()
        summary_table = summary_table[
            summary_table['Agency'].astype(str).str.lower().str.contains(q) |
            summary_table['DR Code'].astype(str).str.lower().str.contains(q) |
            summary_table['FG Code'].astype(str).str.lower().str.contains(q)
        ]
        
    st.dataframe(summary_table, use_container_width=True)

if st.session_state.skipped_rows_log:
    st.markdown("---")
    st.markdown("### ⚠️ Skipped / Invalid Rows Exception Log")
    df_skipped = pd.DataFrame(st.session_state.skipped_rows_log)
    
    skip_search = st.text_input("🔍 Search Skipped Log (Filter by File Name or Agency)", "", key="skip_search")
    if skip_search:
        sq = skip_search.lower()
        df_skipped = df_skipped[
            df_skipped['File Name'].astype(str).str.lower().str.contains(sq) |
            df_skipped['Agency Value'].astype(str).str.lower().str.contains(sq)
        ]
    st.dataframe(df_skipped, use_container_width=True)

if st.session_state.comparison_summary:
    st.markdown("---")
    st.markdown("### 📊 Audit Reconciliation & Comparison Pivot")
    combined_pivot = pd.concat(st.session_state.comparison_summary, ignore_index=True)
    st.dataframe(combined_pivot, use_container_width=True)
    audit_csv = combined_pivot.to_csv(index=False).encode('utf-8')
    st.download_button(
        label="📥 Download Audit Reconciliation Report (CSV)",
        data=audit_csv,
        file_name=f"Audit_Reconciliation_Report_{get_ist_now().strftime('%Y-%m-%d')}.csv",
        mime="text/csv",
        key="audit_csv_download"
    )

# --- Historical Trend Analysis View ---
st.markdown("---")
with st.expander("🕒 View Historical Trend Analysis (SQLite Database - IST)"):
    try:
        conn = sqlite3.connect("sales_history.db")
        df_history = pd.read_sql("SELECT * FROM history_logs ORDER BY id DESC", conn)
        conn.close()
        if not df_history.empty:
            st.dataframe(df_history, use_container_width=True)
            st.markdown("##### Day-over-Day Total Quantity Trend")
            st.line_chart(df_history.set_index("timestamp")["total_qty"])
        else:
            st.info("No historical logs available yet.")
    except Exception as e:
        st.error(f"Error loading history: {str(e)}")
